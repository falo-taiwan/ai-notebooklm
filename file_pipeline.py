# v2.01版 Falo x Force Cheng 2026/6/14
"""Local file pipeline helpers.

The original script mixed folder scanning, Excel conversion, upload, and GUI
work in one file. This module keeps local file handling in one small place.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List

import openpyxl


SKIP_EXTENSIONS = {".tmp", ".ini", ".lnk"}


def get_valid_files(source_dir: Path) -> List[Path]:
    if not source_dir.is_dir():
        return []

    files: List[Path] = []
    for path in sorted(source_dir.iterdir(), key=lambda item: item.name.lower()):
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() in SKIP_EXTENSIONS:
            continue
        if path.is_file():
            files.append(path)
    return files


def xlsx_to_csvs(filepath: Path, temp_dir: Path) -> List[Path]:
    temp_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    stem = filepath.stem
    csvs: List[Path] = []

    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            name = f"{stem}__{sheet}.csv" if len(wb.sheetnames) > 1 else f"{stem}.csv"
            out = temp_dir / name
            with out.open("w", newline="", encoding="utf-8-sig") as handle:
                writer = csv.writer(handle)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if value is None else value for value in row])
            csvs.append(out)
    finally:
        wb.close()

    return csvs
