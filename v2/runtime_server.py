# v2.01版 Falo x Force Cheng 2026/6/14
"""Local HTML portal runtime for AI NotebookLM Runtime Lab."""

from __future__ import annotations

import argparse
import cgi
import os
import platform
import hashlib
import json
import shutil
import socket
import subprocess
import sys
import threading
import time
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Dict
from urllib.parse import parse_qs, quote_plus, urlparse
from zoneinfo import ZoneInfo

from app_config import AppConfig, load_or_create_config
from environment_check import check_environment
from file_pipeline import get_valid_files, xlsx_to_csvs

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent
APP_VERSION = "v2.19"
APP_WATERMARK = "v2.19版 Falo x Force Cheng 2026/6/14"
TAIPEI_TZ = ZoneInfo("Asia/Taipei")
DEFAULT_SIMPLE_TYPES = [".pdf", ".txt", ".md", ".csv", ".docx", ".xlsx", ".pptx", ".png", ".jpg", ".jpeg"]
_GAS_AUTO_WORKER_STARTED = False
_INCOMING_WATCH_WORKER_STARTED = False
_INCOMING_REALTIME_WORKER_STARTED = False
RUNTIME_BIND_HOST = "0.0.0.0"
RUNTIME_BIND_PORT = 8765
DEFAULT_GAS_POLL_INTERVAL_SECONDS = 600
DEFAULT_GAS_MAX_TASKS_PER_POLL = 3
MIN_GAS_POLL_INTERVAL_SECONDS = 30
DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS = 600
MIN_INCOMING_WATCH_INTERVAL_SECONDS = 30
INCOMING_WATCH_MODES = {"polling", "realtime"}
COMMAND_STAGES = ["inbox", "queued", "processing", "completed", "failed", "archived"]
ROLE_PERMISSIONS = {
    "user": {"upload_folder"},
    "document_manager": {"upload_folder", "adapter_execute", "create_notebook", "sync_projects"},
    "admin": {"upload_folder", "adapter_execute", "create_notebook", "sync_projects", "clear_runtime"},
}

SESSION_DB = {}


import queue
import uuid

class TaskQueueManager:
    def __init__(self, config: AppConfig):
        self.config = config
        self.tasks = {}
        self.q = queue.Queue()
        self.lock = threading.Lock()
        self.worker_threads = []
        self.running = False
        self.avg_nlm_duration = 20.0
        self.avg_gemini_duration = 10.0
        self.active_tasks = {}  # task_id -> start_time

    def add_task(self, platform: str, user_name: str, payload: dict) -> str:
        task_id = f"task_{uuid.uuid4().hex[:8]}"
        task = {
            "task_id": task_id,
            "platform": platform,
            "user_name": user_name,
            "status": "pending",
            "payload": payload,
            "result": None,
            "error": None,
            "created_at": time.time(),
            "started_at": None,
            "completed_at": None
        }
        with self.lock:
            self.tasks[task_id] = task
        self.q.put(task_id)
        return task_id

    def get_task(self, task_id: str) -> dict | None:
        with self.lock:
            return self.tasks.get(task_id)

    def clear_queue(self):
        with self.lock:
            for task_id, task in self.tasks.items():
                if task["status"] == "pending":
                    task["status"] = "failed"
                    task["error"] = "Cancelled by administrator"
            while not self.q.empty():
                try:
                    self.q.get_nowait()
                    self.q.task_done()
                except queue.Empty:
                    break

    def get_active_tasks(self) -> list:
        with self.lock:
            active = []
            for t_id, t in self.tasks.items():
                if t["status"] in {"pending", "processing"}:
                    active.append(t)
            active.sort(key=lambda x: x["created_at"])
            
            processing_tasks = [t for t in active if t["status"] == "processing"]
            pending_tasks = [t for t in active if t["status"] == "pending"]
            
            num_workers = 3
            workers = [0.0] * num_workers
            
            # Assign processing tasks to workers
            for idx, t in enumerate(processing_tasks):
                start_time = self.active_tasks.get(t["task_id"], t["started_at"] or t["created_at"])
                elapsed = time.time() - start_time
                base = self.avg_nlm_duration if t["platform"] == "notebooklm" else self.avg_gemini_duration
                rem = max(1.0, base - elapsed)
                
                worker_idx = idx % num_workers
                workers[worker_idx] = rem
                t["eta"] = rem
                
            # Assign pending tasks
            for t in pending_tasks:
                next_free_worker = min(range(num_workers), key=lambda i: workers[i])
                start_wait = workers[next_free_worker]
                base = self.avg_nlm_duration if t["platform"] == "notebooklm" else self.avg_gemini_duration
                completion_time = start_wait + base
                
                workers[next_free_worker] = completion_time
                t["eta"] = completion_time
                
            result = []
            for t in active:
                result.append({
                    "user_name": t["user_name"],
                    "platform": t["platform"],
                    "status": t["status"],
                    "task_id": t["task_id"],
                    "eta_seconds": round(t.get("eta", 0.0), 1)
                })
            return result

    def get_task_status_detail(self, task_id: str) -> dict | None:
        task = self.get_task(task_id)
        if not task:
            return None
        
        position = 0
        if task["status"] == "pending":
            with self.lock:
                pending_tasks = [t for t in self.tasks.values() if t["status"] == "pending"]
                pending_tasks.sort(key=lambda x: x["created_at"])
                try:
                    position = pending_tasks.index(task) + 1
                except ValueError:
                    position = 1
        
        active_list = self.get_active_tasks()
        task_info = next((t for t in active_list if t["task_id"] == task_id), None)
        eta = task_info["eta_seconds"] if task_info else 0.0

        return {
            "task_id": task["task_id"],
            "platform": task["platform"],
            "user_name": task["user_name"],
            "status": task["status"],
            "queue_position": position,
            "eta_seconds": eta,
            "result": task["result"],
            "error": task["error"]
        }

    def start(self, max_workers=3):
        with self.lock:
            if self.running:
                return
            self.running = True
            self.worker_threads = []
            for _ in range(max_workers):
                t = threading.Thread(target=self._worker_loop, daemon=True)
                t.start()
                self.worker_threads.append(t)

    def _worker_loop(self):
        while self.running:
            try:
                task_id = self.q.get(timeout=1.0)
            except queue.Empty:
                continue

            task = self.get_task(task_id)
            if not task:
                self.q.task_done()
                continue

            with self.lock:
                task["status"] = "processing"
                task["started_at"] = time.time()
                self.active_tasks[task_id] = time.time()

            try:
                platform = task["platform"]
                payload = task["payload"]
                
                if platform == "notebooklm":
                    result = self._run_notebooklm(payload)
                else:
                    result = self._run_gemini(payload)

                elapsed = time.time() - task["started_at"]
                
                with self.lock:
                    if result.get("ok"):
                        task["status"] = "completed"
                        task["result"] = result
                        if platform == "notebooklm":
                            self.avg_nlm_duration = self.avg_nlm_duration * 0.7 + elapsed * 0.3
                        else:
                            self.avg_gemini_duration = self.avg_gemini_duration * 0.7 + elapsed * 0.3
                    else:
                        task["status"] = "failed"
                        task["error"] = result.get("error", "Unknown error")
                    task["completed_at"] = time.time()

            except Exception as e:
                with self.lock:
                    task["status"] = "failed"
                    task["error"] = str(e)
                    task["completed_at"] = time.time()
            finally:
                with self.lock:
                    if task_id in self.active_tasks:
                        del self.active_tasks[task_id]
                self.q.task_done()

    def _run_notebooklm(self, payload: dict) -> dict:
        notebook_id = payload["notebook_id"]
        notebook_title = payload.get("notebook_title", "") or f"ID: {notebook_id[:8]}"
        user_name = payload["user_name"]
        user_id = payload.get("user_id", "")
        conversation_id = payload["conversation_id"]
        question = payload["question"]

        helper_path = PROJECT_ROOT / "ask_helper.py"
        cmd = [sys.executable, str(helper_path), "-n", notebook_id, "-q", question]
        if conversation_id and conversation_id != "new":
            cmd.extend(["-c", conversation_id])

        result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", check=False)
        if result.returncode != 0:
            return {"ok": False, "error": f"Helper execution failed: {result.stderr or result.stdout}"}

        try:
            result_data = json.loads(result.stdout)
        except Exception as je:
            return {"ok": False, "error": f"Failed to parse helper output: {result.stdout}, err: {str(je)}"}

        if not result_data.get("ok"):
            return {"ok": False, "error": result_data.get("error", "Unknown helper error")}

        answer = result_data["answer"]
        resolved_conv_id = result_data["conversation_id"]

        with self.lock:
            sessions_file = PROJECT_ROOT / "data" / "multichat_sessions.json"
            
            data = {"sessions": {}}
            if sessions_file.exists():
                try:
                    with open(sessions_file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                except Exception:
                    pass
            
            if "sessions" not in data:
                data["sessions"] = {}
                
            now_str = datetime.now(TAIPEI_TZ).strftime("%Y-%m-%d %H:%M:%S")
            
            if resolved_conv_id not in data["sessions"]:
                data["sessions"][resolved_conv_id] = {
                    "user_id": user_id,
                    "user_name": user_name,
                    "notebook_id": notebook_id,
                    "notebook_title": notebook_title,
                    "created_at": now_str,
                    "last_query_at": now_str,
                    "turns": []
                }
            else:
                data["sessions"][resolved_conv_id]["last_query_at"] = now_str
                data["sessions"][resolved_conv_id]["user_name"] = user_name
                data["sessions"][resolved_conv_id]["user_id"] = user_id
                data["sessions"][resolved_conv_id]["notebook_id"] = notebook_id
                data["sessions"][resolved_conv_id]["notebook_title"] = notebook_title
                
            data["sessions"][resolved_conv_id]["turns"].append({
                "role": "user",
                "content": question,
                "timestamp": now_str,
                "notebook_id": notebook_id,
                "notebook_title": notebook_title
            })
            data["sessions"][resolved_conv_id]["turns"].append({
                "role": "assistant",
                "content": answer,
                "timestamp": now_str
            })
            
            sessions_file.parent.mkdir(parents=True, exist_ok=True)
            with open(sessions_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        return {
            "ok": True,
            "answer": answer,
            "conversation_id": resolved_conv_id
        }

    def _run_gemini(self, payload: dict) -> dict:
        user_name = payload["user_name"]
        user_id = payload.get("user_id", "")
        question = payload["question"]
        metadata = payload["metadata"]
        model = payload.get("model", "").strip()
        thinking = payload.get("thinking", "").strip()

        helper_path = PROJECT_ROOT / "gemini_helper.py"
        cmd = [sys.executable, str(helper_path), "-q", question]
        if metadata and metadata != "new" and metadata != "[]" and metadata != "null":
            cmd.extend(["-m", metadata])
        if model:
            cmd.extend(["--model", model])
        if thinking:
            cmd.extend(["--thinking", thinking])

        result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", check=False)
        if result.returncode != 0:
            return {"ok": False, "error": f"Helper execution failed: {result.stderr or result.stdout}"}

        try:
            result_data = json.loads(result.stdout)
        except Exception as je:
            return {"ok": False, "error": f"Failed to parse helper output: {result.stdout}, err: {str(je)}"}
        
        if not result_data.get("ok"):
            return {"ok": False, "error": result_data.get("error", "Unknown helper error")}

        answer = result_data["answer"]
        resolved_metadata = result_data["metadata"]
        resolved_conv_id = resolved_metadata[0] if resolved_metadata else "unknown"

        with self.lock:
            sessions_file = PROJECT_ROOT / "data" / "gemini_sessions.json"
            
            data = {"sessions": {}}
            if sessions_file.exists():
                try:
                    with open(sessions_file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                except Exception:
                    pass
            
            if "sessions" not in data:
                data["sessions"] = {}
                
            now_str = datetime.now(TAIPEI_TZ).strftime("%Y-%m-%d %H:%M:%S")
            
            if resolved_conv_id not in data["sessions"]:
                data["sessions"][resolved_conv_id] = {
                    "user_id": user_id,
                    "user_name": user_name,
                    "created_at": now_str,
                    "last_query_at": now_str,
                    "metadata": resolved_metadata,
                    "turns": []
                }
            else:
                data["sessions"][resolved_conv_id]["last_query_at"] = now_str
                data["sessions"][resolved_conv_id]["user_name"] = user_name
                data["sessions"][resolved_conv_id]["metadata"] = resolved_metadata
                data["sessions"][resolved_conv_id]["user_id"] = user_id
                
            data["sessions"][resolved_conv_id]["turns"].append({
                "role": "user",
                "content": question,
                "timestamp": now_str
            })
            data["sessions"][resolved_conv_id]["turns"].append({
                "role": "assistant",
                "content": answer,
                "timestamp": now_str
            })
            
            sessions_file.parent.mkdir(parents=True, exist_ok=True)
            with open(sessions_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        return {
            "ok": True,
            "answer": answer,
            "metadata": json.dumps(resolved_metadata)
        }

task_queue_manager: TaskQueueManager | None = None


def now_taipei() -> datetime:
    return datetime.now(TAIPEI_TZ)


def now_iso() -> str:
    return now_taipei().isoformat()


def time_stamp(fmt: str = "%Y%m%d-%H%M%S") -> str:
    return now_taipei().strftime(fmt)


def logs_dir(config: AppConfig) -> Path:
    path = config.project_root / "logs"
    path.mkdir(parents=True, exist_ok=True)
    return path


def config_dir(config: AppConfig) -> Path:
    path = config.project_root / "config"
    path.mkdir(parents=True, exist_ok=True)
    return path


def projects_path(config: AppConfig) -> Path:
    return config_dir(config) / "projects.json"


def runtime_state_path(config: AppConfig) -> Path:
    return config_dir(config) / "runtime_state.json"


def source_pool_dir(config: AppConfig) -> Path:
    path = config.project_root / "data" / "source_pool" / "simple_upload"
    path.mkdir(parents=True, exist_ok=True)
    return path


def simple_incoming_dir(config: AppConfig) -> Path:
    path = source_pool_dir(config) / "incoming"
    path.mkdir(parents=True, exist_ok=True)
    return path


def simple_evidence_dir(config: AppConfig) -> Path:
    path = source_pool_dir(config) / "evidence"
    path.mkdir(parents=True, exist_ok=True)
    return path


def simple_watch_batches_dir(config: AppConfig) -> Path:
    path = source_pool_dir(config) / "watch_batches"
    path.mkdir(parents=True, exist_ok=True)
    return path


def incoming_watch_state_path(config: AppConfig) -> Path:
    path = source_pool_dir(config) / "incoming_watch_state.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def users_path(config: AppConfig) -> Path:
    return config_dir(config) / "users.json"


def command_packages_dir(config: AppConfig) -> Path:
    path = config.project_root / "data" / "command_packages"
    path.mkdir(parents=True, exist_ok=True)
    for stage in COMMAND_STAGES:
        (path / stage).mkdir(parents=True, exist_ok=True)
    return path


def command_stage_dir(config: AppConfig, stage: str) -> Path:
    if stage not in COMMAND_STAGES:
        raise ValueError(f"unsupported command stage: {stage}")
    return command_packages_dir(config) / stage


def command_audit_path(config: AppConfig) -> Path:
    return logs_dir(config) / "command_audit.jsonl"


def write_runtime_log(config: AppConfig, event: str, payload: Dict[str, object]) -> Path:
    path = logs_dir(config) / "runtime.jsonl"
    record = {
        "ts": now_iso(),
        "app": "AI NotebookLM Runtime Lab",
        "event": event,
        "payload": payload,
    }
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")
    return path


def write_command_audit(config: AppConfig, event: str, payload: Dict[str, object]) -> Path:
    path = command_audit_path(config)
    record = {
        "ts": now_iso(),
        "app": "AI NotebookLM Runtime Lab",
        "event": event,
        "payload": payload,
    }
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")
    return path


def load_or_create_users(config: AppConfig) -> Dict[str, object]:
    path = users_path(config)
    modified = False
    data = {}
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            data = {}

    if not data or "local_users" not in data:
        data = {
            "app": "AI NotebookLM Runtime Lab",
            "kind": "local_user_registry",
            "updated_at": now_iso(),
            "roles": {
                "user": {
                    "label": "一般使用者",
                    "description": "可以丟資料、建立本機上傳任務、查看自己的任務狀態。",
                    "allowed_actions": sorted(ROLE_PERMISSIONS["user"]),
                },
                "document_manager": {
                    "label": "文件管理者",
                    "description": "可以管理 Project、整理檔案、執行文件上傳與 ETL 任務。",
                    "allowed_actions": sorted(ROLE_PERMISSIONS["document_manager"]),
                },
                "admin": {
                    "label": "Admin",
                    "description": "可以管理 Runtime 設定、清儲、log、權限與所有任務。",
                    "allowed_actions": sorted(ROLE_PERMISSIONS["admin"]),
                },
            },
            "local_users": [
                {"user_id": "admin", "display_name": "Admin", "role": "admin", "password": "admin123456"},
                {"user_id": "doc_manager", "display_name": "Document Manager", "role": "document_manager", "password": "doc_manager"},
                {"user_id": "general_user", "display_name": "General User", "role": "user", "password": "general_user"},
            ],
            "anonymous_keys": []
        }
        modified = True

    if "roles" not in data:
        data["roles"] = {
            "user": {
                "label": "一般使用者",
                "description": "可以丟資料、建立本機上傳任務、查看自己的任務狀態。",
                "allowed_actions": sorted(ROLE_PERMISSIONS["user"]),
            },
            "document_manager": {
                "label": "文件管理者",
                "description": "可以管理 Project、整理檔案、執行文件上傳與 ETL 任務。",
                "allowed_actions": sorted(ROLE_PERMISSIONS["document_manager"]),
            },
            "admin": {
                "label": "Admin",
                "description": "可以管理 Runtime 設定、清儲、log、權限與所有任務。",
                "allowed_actions": sorted(ROLE_PERMISSIONS["admin"]),
            },
        }
        modified = True

    if "anonymous_keys" not in data:
        data["anonymous_keys"] = []
        modified = True

    for k in data.get("anonymous_keys", []):
        if "key_id" not in k:
            k["key_id"] = f"anon_key_{uuid.uuid4().hex[:8]}"
            modified = True

    for u in data.get("local_users", []):
        if "password" not in u:
            # Migration: set default password to user_id (predictable) or "admin123456" if it's admin
            if u.get("user_id") == "admin":
                u["password"] = "admin123456"
            else:
                u["password"] = u.get("user_id", "123456")
            modified = True

    if modified:
        data["updated_at"] = now_iso()
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return data


def save_registry_directly(config: AppConfig, data: Dict[str, object]) -> None:
    path = users_path(config)
    data["updated_at"] = now_iso()
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")



def read_runtime_log_records(config: AppConfig) -> list:
    path = logs_dir(config) / "runtime.jsonl"
    if not path.exists():
        return []
    records = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        if not line.strip():
            continue
        try:
            record = json.loads(line)
            record["_line"] = line_number
            records.append(record)
        except json.JSONDecodeError:
            records.append({"_line": line_number, "ts": "", "event": "invalid_json", "payload": {"raw": line}})
    return records


def query_runtime_logs(config: AppConfig, search: str = "", event: str = "", page: int = 1, page_size: int = 50) -> Dict[str, object]:
    records = list(reversed(read_runtime_log_records(config)))
    search_text = search.strip().lower()
    event_text = event.strip().lower()
    if event_text:
        records = [item for item in records if event_text in str(item.get("event", "")).lower()]
    if search_text:
        records = [item for item in records if search_text in json.dumps(item, ensure_ascii=False).lower()]

    page = max(1, page)
    page_size = max(1, min(page_size, 200))
    total = len(records)
    start = (page - 1) * page_size
    return {
        "mode": "log_cms",
        "search": search,
        "event": event,
        "page": page,
        "page_size": page_size,
        "total": total,
        "logs": records[start : start + page_size],
    }


def load_projects(config: AppConfig) -> list:
    path = projects_path(config)
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return data.get("projects", []) if isinstance(data, dict) else []


def save_projects(config: AppConfig, projects: list) -> Path:
    path = projects_path(config)
    payload = {
        "app": "AI NotebookLM Runtime Lab",
        "kind": "project_registry",
        "updated_at": now_iso(),
        "project_count": len(projects),
        "projects": projects,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def make_project_id(name: str, notebook_id: str) -> str:
    base = "".join(char.lower() if char.isalnum() else "-" for char in name.strip())[:32].strip("-")
    if not base:
        base = "notebook"
    suffix = notebook_id.replace("-", "")[:8] or "local"
    return f"{base}-{suffix}"


def sync_projects_from_notebooks(config: AppConfig, notebooks: list) -> Dict[str, object]:
    existing = load_projects(config)
    by_notebook = {item.get("notebook_id"): item for item in existing if item.get("notebook_id")}
    now = now_iso()
    changed = 0

    for notebook in notebooks:
        notebook_id = str(notebook.get("id", "")).strip()
        if not notebook_id:
            continue
        title = str(notebook.get("title", "") or "(untitled)").strip()
        if notebook_id in by_notebook:
            project = by_notebook[notebook_id]
            project["name"] = project.get("name") or title
            project["notebook_title"] = title
            project["is_owner"] = notebook.get("is_owner", "")
            project["notebook_created_at"] = notebook.get("created_at", "")
            project["updated_at"] = now
        else:
            project = {
                "project_id": make_project_id(title, notebook_id),
                "name": title,
                "notebook_id": notebook_id,
                "notebook_title": title,
                "status": "active",
                "tags": [],
                "is_owner": notebook.get("is_owner", ""),
                "notebook_created_at": notebook.get("created_at", ""),
                "created_at": now,
                "updated_at": now,
                "source": "notebooklm_sync",
            }
            existing.append(project)
            by_notebook[notebook_id] = project
            changed += 1

    save_projects(config, existing)
    write_runtime_log(config, "sync_projects_from_notebooks", {"project_count": len(existing), "added_count": changed})
    return {"mode": "project_sync", "project_count": len(existing), "added_count": changed, "projects": existing}


def list_projects(config: AppConfig, search: str = "", sort: str = "updated_at", page: int = 1, page_size: int = 25) -> Dict[str, object]:
    projects = load_projects(config)
    search_text = search.strip().lower()
    if search_text:
        projects = [
            item
            for item in projects
            if search_text in json.dumps(item, ensure_ascii=False).lower()
        ]

    sort_key = sort if sort in {"name", "notebook_id", "notebook_created_at", "updated_at", "status"} else "updated_at"
    projects = sorted(projects, key=lambda item: str(item.get(sort_key, "")).lower(), reverse=sort_key in {"updated_at", "notebook_created_at"})
    page = max(1, page)
    page_size = max(1, min(page_size, 500))
    total = len(projects)
    start = (page - 1) * page_size
    return {
        "mode": "project_registry",
        "search": search,
        "sort": sort_key,
        "page": page,
        "page_size": page_size,
        "total": total,
        "projects": projects[start : start + page_size],
    }


def find_project(config: AppConfig, project_id: str) -> Dict[str, object]:
    for project in load_projects(config):
        if project.get("project_id") == project_id:
            return project
    return {}


def read_runtime_settings(config: AppConfig) -> Dict[str, object]:
    defaults = {
        "app": "AI NotebookLM Runtime Lab",
        "kind": "runtime_settings",
        "active_project_id": "",
        "allow_network_access": True,
        "incoming_watch_enabled": False,
        "incoming_watch_mode": "polling",
        "incoming_watch_interval_seconds": DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS,
        "incoming_watch_folder": str(simple_incoming_dir(config)),
        "incoming_watch_recursive": False,
        "incoming_watch_file_types": list(DEFAULT_SIMPLE_TYPES),
        "incoming_watch_min_age_seconds": 10,
        "incoming_watch_max_files_per_scan": 10,
        "incoming_watch_auto_queue": True,
        "incoming_watch_auto_execute": False,
    }
    path = runtime_state_path(config)
    if not path.exists():
        return defaults
    settings = json.loads(path.read_text(encoding="utf-8"))
    for key, value in defaults.items():
        settings.setdefault(key, value)
    return settings


def write_runtime_settings(config: AppConfig, settings: Dict[str, object]) -> Path:
    settings["app"] = "AI NotebookLM Runtime Lab"
    settings["kind"] = "runtime_settings"
    settings["updated_at"] = now_iso()
    path = runtime_state_path(config)
    path.write_text(json.dumps(settings, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def is_loopback_client(client_ip: str) -> bool:
    return client_ip.startswith("127.") or client_ip == "::1" or client_ip == "localhost"


def network_access_enabled(config: AppConfig) -> bool:
    return bool(read_runtime_settings(config).get("allow_network_access", True))


def set_network_access(config: AppConfig, enabled: bool) -> Dict[str, object]:
    # Falo x Force 教學註解：
    # server 預設綁 0.0.0.0，代表技術上可被同網段連線。
    # 但是否接受非本機 IP，由這個 runtime setting 在線控制，不需要重啟 server。
    settings = read_runtime_settings(config)
    settings["allow_network_access"] = enabled
    write_runtime_settings(config, settings)
    write_runtime_log(config, "network_access_updated", {"allow_network_access": enabled})
    return {
        "ok": True,
        "mode": "network_access",
        "allow_network_access": enabled,
        "message": "Same-network access enabled." if enabled else "Same-network access disabled. Localhost remains available.",
    }


def set_active_project(config: AppConfig, project_id: str) -> Dict[str, object]:
    project = find_project(config, project_id)
    if not project:
        return {"ok": False, "active_project_id": "", "project": {}, "error": "project not found"}
    settings = read_runtime_settings(config)
    settings["active_project_id"] = project_id
    settings["active_notebook_id"] = project.get("notebook_id", "")
    settings["active_project_name"] = project.get("name", "")
    write_runtime_settings(config, settings)
    write_runtime_log(config, "set_active_project", {"project_id": project_id, "notebook_id": project.get("notebook_id", "")})
    return {"ok": True, "active_project_id": project_id, "active_notebook_id": project.get("notebook_id", ""), "project": project}


def get_active_project(config: AppConfig) -> Dict[str, object]:
    settings = read_runtime_settings(config)
    project_id = str(settings.get("active_project_id", ""))
    project = find_project(config, project_id) if project_id else {}
    return {"settings": settings, "active_project_id": project_id, "project": project}


def list_notebook_sources(config: AppConfig, notebook_id: str) -> Dict[str, object]:
    if not notebook_id:
        return {"ok": False, "sources": [], "count": 0, "error": "notebook_id is required"}
    result = subprocess.run(
        [config.notebooklm_command, "source", "list", "-n", notebook_id, "--json"],
        capture_output=True,
        timeout=180,
    )
    output = result.stdout + result.stderr
    if result.returncode != 0:
        return {
            "ok": False,
            "sources": [],
            "count": 0,
            "returncode": result.returncode,
            "error": output.decode("utf-8", errors="replace").strip()[:1000],
        }
    data = _parse_notebooklm_json_output(result.stdout)
    sources = data.get("sources", []) if isinstance(data, dict) else []
    return {"ok": True, "sources": sources, "count": len(sources), "notebook": data}


def create_or_use_notebook(config: AppConfig, title: str) -> Dict[str, object]:
    clean_title = title.strip()
    if not clean_title:
        return {"ok": False, "mode": "create_or_use_notebook", "error": "title is required", "project": {}}
    notebooks = list_notebooks(config)
    if not notebooks.get("ok"):
        return {"ok": False, "mode": "create_or_use_notebook", "error": notebooks.get("error", ""), "project": {}}
    exact = next((item for item in notebooks.get("notebooks", []) if str(item.get("title", "")).strip() == clean_title), None)
    if exact:
        sync_projects_from_notebooks(config, notebooks.get("notebooks", []))
        project_id = make_project_id(str(exact.get("title", "")), str(exact.get("id", "")))
        active = set_active_project(config, project_id)
        return {"ok": True, "mode": "use_existing_notebook", "notebook": exact, "project": active.get("project", {}), "message": "Existing notebook found and selected."}

    created = create_notebook(config, clean_title)
    if not created.get("ok"):
        return {"ok": False, "mode": "create_or_use_notebook", "error": created.get("error", ""), "project": {}}
    notebook = created.get("notebook", {})
    project_id = make_project_id(str(notebook.get("title", clean_title)), str(notebook.get("id", "")))
    active = set_active_project(config, project_id)
    return {"ok": True, "mode": "created_notebook", "notebook": notebook, "project": active.get("project", {}), "message": "Notebook created and selected."}


def parse_extensions(values: list) -> list:
    raw = []
    for value in values:
        raw.extend(str(value).split(","))
    extensions = []
    for item in raw:
        ext = item.strip().lower()
        if not ext:
            continue
        if not ext.startswith("."):
            ext = "." + ext
        extensions.append(ext)
    return extensions or list(DEFAULT_SIMPLE_TYPES)


def make_command_id(prefix: str = "cmd") -> str:
    return f"{prefix}_{time_stamp('%Y%m%d_%H%M%S')}"


def command_filename(command_id: str) -> str:
    safe = "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in command_id.strip())
    return f"{safe or make_command_id()}.json"


def command_package_template(config: AppConfig) -> Dict[str, object]:
    active = get_active_project(config)
    project_id = str(active.get("active_project_id", ""))
    return {
        "app": "AI NotebookLM Runtime Lab",
        "kind": "command_package",
        "version": "0.1",
        "command_id": make_command_id("cmd_upload_folder"),
        "command_type": "upload_folder",
        "submitter": "admin",
        "role": "admin",
        "target_project_id": project_id,
        "source": {
            "type": "folder",
            "path": str(simple_incoming_dir(config)),
            "recursive": False,
            "file_types": list(DEFAULT_SIMPLE_TYPES),
            "order": "name",
        },
        "options": {
            "duplicate_policy": "rename",
            "copy_evidence": True,
            "evidence_root": str(simple_evidence_dir(config)),
        },
        "execution": {
            "mode": "manual",
            "requires_confirm": True,
        },
        "created_at": now_iso(),
    }


def create_sample_command_package(config: AppConfig) -> Dict[str, object]:
    command = command_package_template(config)
    target = command_stage_dir(config, "inbox") / command_filename(str(command["command_id"]))
    target.write_text(json.dumps(command, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_command_audit(config, "create_sample_command_package", {"path": str(target), "command_id": command["command_id"]})
    write_runtime_log(config, "create_sample_command_package", {"path": str(target), "command_id": command["command_id"]})
    return {"ok": True, "mode": "create_sample_command_package", "path": str(target), "command": command}


def read_incoming_watch_state(config: AppConfig) -> Dict[str, object]:
    path = incoming_watch_state_path(config)
    if not path.exists():
        return {"kind": "incoming_watch_state", "processed": {}}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        data = {}
    if not isinstance(data, dict):
        data = {}
    data.setdefault("kind", "incoming_watch_state")
    data.setdefault("processed", {})
    return data


def write_incoming_watch_state(config: AppConfig, state: Dict[str, object]) -> Path:
    state["kind"] = "incoming_watch_state"
    state["updated_at"] = now_iso()
    path = incoming_watch_state_path(config)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def file_fingerprint(path: Path) -> str:
    stat = path.stat()
    return f"{path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}|{sha256_file(path)}"


def incoming_watch_settings_result(config: AppConfig) -> Dict[str, object]:
    settings = read_runtime_settings(config)
    keys = [
        "incoming_watch_enabled",
        "incoming_watch_mode",
        "incoming_watch_interval_seconds",
        "incoming_watch_folder",
        "incoming_watch_recursive",
        "incoming_watch_file_types",
        "incoming_watch_min_age_seconds",
        "incoming_watch_max_files_per_scan",
        "incoming_watch_auto_queue",
        "incoming_watch_auto_execute",
    ]
    return {
        "ok": True,
        "mode": "incoming_watch_settings_read",
        "realtime_engine": incoming_realtime_engine_status(),
        "settings": {key: settings.get(key) for key in keys},
    }


def incoming_realtime_engine_status() -> Dict[str, object]:
    try:
        import watchdog  # type: ignore

        return {"available": True, "engine": "watchdog", "path": str(getattr(watchdog, "__file__", ""))}
    except Exception as exc:
        return {"available": False, "engine": "watchdog", "error": str(exc)}


def update_incoming_watch_settings_from_params(config: AppConfig, params: Dict[str, list]) -> Dict[str, object]:
    # Falo x Force 教學註解：
    # incoming watcher 是本機資料夾 trigger。它不直接碰 NotebookLM，而是先產生 command package，
    # 讓 queue / audit / evidence 機制完整保留。
    settings = read_runtime_settings(config)
    settings["incoming_watch_enabled"] = params.get("incoming_watch_enabled", [""])[0].lower() in {"1", "yes", "on", "true"}
    settings["incoming_watch_auto_queue"] = params.get("incoming_watch_auto_queue", [""])[0].lower() in {"1", "yes", "on", "true"}
    settings["incoming_watch_auto_execute"] = params.get("incoming_watch_auto_execute", [""])[0].lower() in {"1", "yes", "on", "true"}
    mode = params.get("incoming_watch_mode", ["polling"])[0].strip().lower()
    settings["incoming_watch_mode"] = mode if mode in INCOMING_WATCH_MODES else "polling"
    folder = params.get("incoming_watch_folder", [""])[0].strip()
    if folder:
        settings["incoming_watch_folder"] = folder
    settings["incoming_watch_recursive"] = params.get("incoming_watch_recursive", [""])[0].lower() in {"1", "yes", "on", "true"}
    types = parse_extensions(params.get("types", []))
    settings["incoming_watch_file_types"] = types
    try:
        settings["incoming_watch_interval_seconds"] = max(
            MIN_INCOMING_WATCH_INTERVAL_SECONDS,
            int(params.get("incoming_watch_interval_seconds", [str(DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS)])[0] or DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS),
        )
    except ValueError:
        settings["incoming_watch_interval_seconds"] = DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS
    try:
        settings["incoming_watch_min_age_seconds"] = max(0, int(params.get("incoming_watch_min_age_seconds", ["10"])[0] or 10))
    except ValueError:
        settings["incoming_watch_min_age_seconds"] = 10
    try:
        settings["incoming_watch_max_files_per_scan"] = max(1, min(100, int(params.get("incoming_watch_max_files_per_scan", ["10"])[0] or 10)))
    except ValueError:
        settings["incoming_watch_max_files_per_scan"] = 10
    write_runtime_settings(config, settings)
    write_runtime_log(config, "incoming_watch_settings_updated", {key: value for key, value in settings.items() if key.startswith("incoming_watch_")})
    return incoming_watch_settings_result(config)


def create_incoming_watch_command(config: AppConfig, files: list, settings: Dict[str, object]) -> Dict[str, object]:
    active = get_active_project(config)
    project_id = str(active.get("active_project_id", ""))
    if not project_id:
        return {"ok": False, "mode": "incoming_watch_command", "error": "Active Project is required before watching incoming files."}
    stamp = time_stamp("%Y%m%d-%H%M%S")
    batch_dir = simple_watch_batches_dir(config) / f"watch_{stamp}"
    batch_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for item in files:
        source_path = Path(str(item["path"]))
        target_path = unique_upload_path(batch_dir, source_path.name)
        shutil.copy2(source_path, target_path)
        copied.append({**item, "batch_path": str(target_path)})

    command_id = make_command_id("watch_upload_folder")
    command = command_package_template(config)
    command.update(
        {
            "command_id": command_id,
            "submitter": "local_incoming_watcher",
            "role": "admin",
            "target_project_id": project_id,
            "source": {
                "type": "folder",
                "path": str(batch_dir),
                "recursive": False,
                "file_types": list(settings.get("incoming_watch_file_types") or DEFAULT_SIMPLE_TYPES),
                "order": "name",
            },
            "execution": {
                "mode": "incoming_watch",
                "requires_confirm": False,
            },
            "watch": {
                "source_folder": str(settings.get("incoming_watch_folder") or simple_incoming_dir(config)),
                "batch_dir": str(batch_dir),
                "files": copied,
            },
            "created_at": now_iso(),
        }
    )
    target = command_stage_dir(config, "inbox") / command_filename(command_id)
    write_command_file(target, command)
    write_command_audit(config, "incoming_watch_command_created", {"command_id": command_id, "path": str(target), "file_count": len(copied), "files": [item["name"] for item in copied]})
    write_runtime_log(config, "incoming_watch_command_created", {"command_id": command_id, "path": str(target), "file_count": len(copied), "batch_dir": str(batch_dir)})
    return {"ok": True, "mode": "incoming_watch_command", "command_id": command_id, "path": str(target), "batch_dir": str(batch_dir), "files": copied}


def incoming_watch_scan_once(config: AppConfig, execute: bool | None = None) -> Dict[str, object]:
    settings = read_runtime_settings(config)
    folder = Path(str(settings.get("incoming_watch_folder") or simple_incoming_dir(config))).expanduser()
    extensions = parse_extensions(list(settings.get("incoming_watch_file_types") or DEFAULT_SIMPLE_TYPES))
    recursive = bool(settings.get("incoming_watch_recursive", False))
    max_files = int(settings.get("incoming_watch_max_files_per_scan") or 10)
    min_age = int(settings.get("incoming_watch_min_age_seconds") or 0)
    scan = scan_source_folder(config, str(folder), extensions, recursive, "modified_asc")
    if not scan.get("ok"):
        write_runtime_log(config, "incoming_watch_scan_error", {"error": scan.get("error", ""), "folder": scan.get("folder", "")})
        return {"ok": False, "mode": "incoming_watch_scan", "error": scan.get("error", ""), "folder": scan.get("folder", "")}

    state = read_incoming_watch_state(config)
    processed = state.get("processed", {}) if isinstance(state.get("processed", {}), dict) else {}
    now_ts = time.time()
    new_files = []
    skipped_young = []
    for item in scan.get("files", []):
        path = Path(str(item.get("path", "")))
        if not path.exists():
            continue
        age = now_ts - float(item.get("mtime", now_ts))
        if age < min_age:
            skipped_young.append({"name": item.get("name", ""), "age_seconds": round(age, 1)})
            continue
        fingerprint = file_fingerprint(path)
        if fingerprint in processed:
            continue
        new_files.append({**item, "fingerprint": fingerprint, "sha256": fingerprint.split("|")[-1]})
        if len(new_files) >= max_files:
            break

    created = {}
    queued = {}
    executed = {}
    if new_files:
        created = create_incoming_watch_command(config, new_files, settings)
        if created.get("ok"):
            current_iso = now_iso()
            for item in new_files:
                processed[item["fingerprint"]] = {
                    "path": item["path"],
                    "name": item["name"],
                    "command_id": created.get("command_id", ""),
                    "first_seen_at": current_iso,
                    "sha256": item.get("sha256", ""),
                }
            state["processed"] = processed
            write_incoming_watch_state(config, state)
            if settings.get("incoming_watch_auto_queue", True):
                queued = queue_command_packages(config)
            should_execute = bool(settings.get("incoming_watch_auto_execute", False)) if execute is None else bool(execute)
            if should_execute:
                executed = execute_command_queue(config, command_id=str(created.get("command_id", "")), limit=1)

    result = {
        "ok": True,
        "mode": "incoming_watch_scan",
        "folder": scan.get("folder", ""),
        "scanned_count": scan.get("count", 0),
        "new_count": len(new_files),
        "skipped_young": skipped_young,
        "created": created,
        "queued": queued,
        "executed": executed,
    }
    write_runtime_log(config, "incoming_watch_scan", {"new_count": len(new_files), "scanned_count": scan.get("count", 0), "created_command": created.get("command_id", "") if isinstance(created, dict) else ""})
    return result


def read_command_file(path: Path) -> Dict[str, object]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, dict) else {}


def write_command_file(path: Path, command: Dict[str, object]) -> None:
    path.write_text(json.dumps(command, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def move_command_file(config: AppConfig, source: Path, stage: str, command: Dict[str, object]) -> Path:
    target = command_stage_dir(config, stage) / source.name
    if target.exists():
        target = target.with_name(f"{target.stem}__{time_stamp('%Y%m%d-%H%M%S')}{target.suffix}")
    write_command_file(source, command)
    shutil.move(str(source), str(target))
    return target


def validate_command_package(config: AppConfig, command: Dict[str, object]) -> Dict[str, object]:
    errors = []
    command_id = str(command.get("command_id", "")).strip()
    command_type = str(command.get("command_type") or command.get("action") or "").strip()
    role = str(command.get("role", "user")).strip() or "user"
    project_id = str(command.get("target_project_id", "")).strip()
    source = command.get("source", {}) if isinstance(command.get("source", {}), dict) else {}
    if not command_id:
        errors.append("command_id is required")
    if command_type not in {"upload_folder"}:
        errors.append("command_type must be upload_folder in this MVP")
    if role not in ROLE_PERMISSIONS:
        errors.append(f"unsupported role: {role}")
    elif command_type and command_type not in ROLE_PERMISSIONS[role]:
        errors.append(f"role {role} cannot execute {command_type}")
    if not project_id:
        errors.append("target_project_id is required")
    elif not find_project(config, project_id):
        errors.append(f"project not found: {project_id}")
    if str(source.get("type", "folder")) != "folder":
        errors.append("source.type must be folder")
    folder_path = str(source.get("path", "")).strip()
    if not folder_path:
        errors.append("source.path is required")
    else:
        folder = Path(folder_path).expanduser()
        if not folder.is_absolute():
            folder = (config.project_root / folder).resolve()
        if not folder.exists() or not folder.is_dir():
            errors.append(f"source folder not found: {folder}")
    return {"ok": not errors, "errors": errors}


def summarize_command_file(path: Path, stage: str) -> Dict[str, object]:
    try:
        command = read_command_file(path)
        validation = validate_command_package(load_or_create_config(PROJECT_ROOT), command)
        gas_meta = command.get("gas", {}) if isinstance(command.get("gas", {}), dict) else {}
        watch_meta = command.get("watch", {}) if isinstance(command.get("watch", {}), dict) else {}
        downloaded = gas_meta.get("downloaded", []) if isinstance(gas_meta.get("downloaded", []), list) else []
        watched = watch_meta.get("files", []) if isinstance(watch_meta.get("files", []), list) else []
        source_names = [str(item.get("name", "")) for item in downloaded if isinstance(item, dict) and item.get("name")]
        source_names.extend(str(item.get("name", "")) for item in watched if isinstance(item, dict) and item.get("name"))
        return {
            "stage": stage,
            "file": str(path),
            "filename": path.name,
            "command_id": command.get("command_id", ""),
            "cloud_task_id": command.get("cloud_task_id", gas_meta.get("task_id", "")),
            "trigger_source": command.get("trigger_source", gas_meta.get("trigger_source", "")),
            "trigger_mode": command.get("trigger_mode", gas_meta.get("trigger_mode", "")),
            "cloud_event_type": command.get("cloud_event_type", gas_meta.get("cloud_event_type", "")),
            "command_type": command.get("command_type", ""),
            "submitter": command.get("submitter", ""),
            "role": command.get("role", ""),
            "target_project_id": command.get("target_project_id", ""),
            "source_mode": command.get("execution", {}).get("mode", "") if isinstance(command.get("execution", {}), dict) else "",
            "source_names": source_names,
            "source_count": len(source_names),
            "created_at": command.get("created_at", ""),
            "status": command.get("status", stage),
            "updated_at": command.get("updated_at", command.get("created_at", "")),
            "valid": validation.get("ok", False),
            "errors": validation.get("errors", []),
        }
    except Exception as exc:
        return {
            "stage": stage,
            "file": str(path),
            "filename": path.name,
            "command_id": "",
            "trigger_source": "",
            "trigger_mode": "",
            "cloud_event_type": "",
            "command_type": "",
            "submitter": "",
            "role": "",
            "target_project_id": "",
            "status": "invalid_json",
            "updated_at": "",
            "valid": False,
            "errors": [str(exc)],
        }


def list_command_packages(config: AppConfig) -> Dict[str, object]:
    command_packages_dir(config)
    rows = []
    counts = {}
    for stage in COMMAND_STAGES:
        files = sorted(command_stage_dir(config, stage).glob("*.json"), key=lambda item: item.name.lower())
        counts[stage] = len(files)
        rows.extend(summarize_command_file(path, stage) for path in files)
    return {
        "mode": "command_queue",
        "counts": counts,
        "total": len(rows),
        "commands": rows,
        "users": load_or_create_users(config),
        "auto_run": bool(read_runtime_settings(config).get("command_auto_run", False)),
    }


def archive_command_package(config: AppConfig, command_id: str) -> Dict[str, object]:
    for stage in ["queued", "completed", "failed"]:
        path = command_stage_dir(config, stage) / f"{command_id}.json"
        if not path.exists():
            continue
        command = read_command_file(path)
        command["previous_stage"] = stage
        command["status"] = "archived"
        command["archived_at"] = now_iso()
        command["updated_at"] = command["archived_at"]
        target = move_command_file(config, path, "archived", command)
        write_command_audit(config, "command_archived", {"command_id": command_id, "from": stage, "path": str(target)})
        return {"ok": True, "mode": "archive_command_package", "command_id": command_id, "from_stage": stage, "path": str(target)}
    return {"ok": False, "mode": "archive_command_package", "command_id": command_id, "error": "command not found in queued/completed/failed"}


def queue_command_packages(config: AppConfig) -> Dict[str, object]:
    queued = []
    failed = []
    for path in sorted(command_stage_dir(config, "inbox").glob("*.json"), key=lambda item: item.name.lower()):
        try:
            command = read_command_file(path)
            validation = validate_command_package(config, command)
            now = now_iso()
            if validation.get("ok"):
                command["status"] = "queued"
                command["queued_at"] = now
                command["updated_at"] = now
                target = move_command_file(config, path, "queued", command)
                queued.append({"command_id": command.get("command_id", ""), "path": str(target)})
                write_command_audit(config, "command_queued", {
                    "command_id": command.get("command_id", ""),
                    "path": str(target),
                    "trigger_source": command.get("trigger_source", ""),
                    "trigger_mode": command.get("trigger_mode", ""),
                    "cloud_event_type": command.get("cloud_event_type", ""),
                })
            else:
                command["status"] = "failed_validation"
                command["validation_errors"] = validation.get("errors", [])
                command["updated_at"] = now
                target = move_command_file(config, path, "failed", command)
                failed.append({"command_id": command.get("command_id", ""), "path": str(target), "errors": validation.get("errors", [])})
                write_command_audit(config, "command_validation_failed", {"command_id": command.get("command_id", ""), "errors": validation.get("errors", [])})
        except Exception as exc:
            failed.append({"command_id": path.stem, "path": str(path), "errors": [str(exc)]})
    write_runtime_log(config, "queue_command_packages", {"queued_count": len(queued), "failed_count": len(failed)})
    return {"ok": not failed, "mode": "queue_command_packages", "queued_count": len(queued), "failed_count": len(failed), "queued": queued, "failed": failed}


def command_to_folder_upload_params(command: Dict[str, object]) -> Dict[str, list]:
    source = command.get("source", {}) if isinstance(command.get("source", {}), dict) else {}
    options = command.get("options", {}) if isinstance(command.get("options", {}), dict) else {}
    return {
        "folder_path": [str(source.get("path", ""))],
        "types": list(source.get("file_types", DEFAULT_SIMPLE_TYPES)),
        "recursive": ["yes" if source.get("recursive") else ""],
        "order": [str(source.get("order", "name"))],
        "project_id": [str(command.get("target_project_id", ""))],
        "conflict_policy": [str(options.get("duplicate_policy", "rename"))],
        "evidence_root": [str(options.get("evidence_root", ""))],
    }


def execute_command_file(config: AppConfig, path: Path) -> Dict[str, object]:
    command = read_command_file(path)
    validation = validate_command_package(config, command)
    now = now_iso()
    if not validation.get("ok"):
        command["status"] = "failed_validation"
        command["validation_errors"] = validation.get("errors", [])
        command["updated_at"] = now
        failed_path = move_command_file(config, path, "failed", command)
        return {"ok": False, "command_id": command.get("command_id", ""), "path": str(failed_path), "errors": validation.get("errors", [])}

    command["status"] = "processing"
    command["started_at"] = now
    command["updated_at"] = now
    processing_path = move_command_file(config, path, "processing", command)
    try:
        if command.get("command_type") == "upload_folder":
            result = execute_folder_upload(config, command_to_folder_upload_params(command))
        else:
            result = {"ok": False, "error": "unsupported command_type"}
        finished_at = now_iso()
        command["status"] = "completed" if result.get("ok") else "failed"
        command["finished_at"] = finished_at
        command["updated_at"] = finished_at
        command["result"] = result
        final_stage = "completed" if result.get("ok") else "failed"
        final_path = move_command_file(config, processing_path, final_stage, command)
        source_payload = {
            "command_id": command.get("command_id", ""),
            "status": command["status"],
            "path": str(final_path),
            "trigger_source": command.get("trigger_source", ""),
            "trigger_mode": command.get("trigger_mode", ""),
            "cloud_event_type": command.get("cloud_event_type", ""),
        }
        write_command_audit(config, "command_executed", source_payload)
        write_runtime_log(config, "command_executed", {**source_payload, "result_ok": result.get("ok", False)})
        return {"ok": bool(result.get("ok")), "command_id": command.get("command_id", ""), "status": command["status"], "path": str(final_path), "result": result}
    except Exception as exc:
        failed_at = now_iso()
        command["status"] = "failed"
        command["finished_at"] = failed_at
        command["updated_at"] = failed_at
        command["error"] = str(exc)
        failed_path = move_command_file(config, processing_path, "failed", command)
        write_command_audit(config, "command_execute_failed", {"command_id": command.get("command_id", ""), "error": str(exc)})
        write_runtime_log(config, "command_execute_failed", {"command_id": command.get("command_id", ""), "error": str(exc)})
        return {"ok": False, "command_id": command.get("command_id", ""), "status": "failed", "path": str(failed_path), "error": str(exc)}


def execute_command_queue(config: AppConfig, command_id: str = "", limit: int = 1) -> Dict[str, object]:
    queue_command_packages(config)
    queued_files = sorted(command_stage_dir(config, "queued").glob("*.json"), key=lambda item: item.name.lower())
    if command_id:
        queued_files = [path for path in queued_files if summarize_command_file(path, "queued").get("command_id") == command_id or path.name == command_id]
    limit = max(1, min(limit, 50))
    results = [execute_command_file(config, path) for path in queued_files[:limit]]
    return {"ok": not any(not item.get("ok") for item in results), "mode": "execute_command_queue", "executed_count": len(results), "results": results}


def set_command_auto_run(config: AppConfig, enabled: bool) -> Dict[str, object]:
    settings = read_runtime_settings(config)
    settings["command_auto_run"] = bool(enabled)
    write_runtime_settings(config, settings)
    write_command_audit(config, "set_command_auto_run", {"enabled": bool(enabled)})
    return {"ok": True, "auto_run": bool(enabled)}


def command_auto_tick(config: AppConfig, limit: int = 1) -> Dict[str, object]:
    if not read_runtime_settings(config).get("command_auto_run"):
        return {"ok": False, "mode": "command_auto_tick", "executed_count": 0, "error": "Auto Mode is OFF."}
    result = execute_command_queue(config, limit=limit)
    result["mode"] = "command_auto_tick"
    write_command_audit(config, "command_auto_tick", {"executed_count": result.get("executed_count", 0), "ok": result.get("ok", False)})
    return result


def scan_source_folder(config: AppConfig, folder_path: str, extensions: list, recursive: bool, order: str) -> Dict[str, object]:
    folder = Path(folder_path or str(simple_incoming_dir(config))).expanduser()
    if not folder.is_absolute():
        folder = (config.project_root / folder).resolve()
    if not folder.exists() or not folder.is_dir():
        return {"ok": False, "folder": str(folder), "files": [], "count": 0, "error": "Folder does not exist or is not a directory."}
    pattern = "**/*" if recursive else "*"
    files = []
    allowed = {item.lower() for item in extensions}
    for path in folder.glob(pattern):
        if not path.is_file():
            continue
        if path.name.startswith("."):
            continue
        if path.suffix.lower() not in allowed:
            continue
        stat = path.stat()
        files.append(
            {
                "name": path.name,
                "path": str(path),
                "suffix": path.suffix.lower(),
                "size_kb": round(stat.st_size / 1024, 1),
                "mtime": stat.st_mtime,
            }
        )
    if order == "modified_desc":
        files.sort(key=lambda item: item["mtime"], reverse=True)
    elif order == "modified_asc":
        files.sort(key=lambda item: item["mtime"])
    else:
        files.sort(key=lambda item: item["name"].lower())
    return {"ok": True, "folder": str(folder), "extensions": sorted(allowed), "recursive": recursive, "order": order, "count": len(files), "files": files}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def prepare_evidence_copy(config: AppConfig, project_id: str, original_path: Path, target_name: str, evidence_root: str = "") -> Dict[str, object]:
    root = Path(evidence_root).expanduser() if evidence_root else simple_evidence_dir(config)
    if not root.is_absolute():
        root = (config.project_root / root).resolve()
    stamp = time_stamp("%Y%m%d-%H%M%S")
    safe_project = project_id or "manual"
    target_dir = root / safe_project / stamp
    target_dir.mkdir(parents=True, exist_ok=True)
    evidence_path = unique_upload_path(target_dir, target_name)
    shutil.copy2(original_path, evidence_path)
    return {
        "original_path": str(original_path),
        "evidence_path": str(evidence_path),
        "sha256": sha256_file(evidence_path),
        "size_kb": round(evidence_path.stat().st_size / 1024, 1),
        "manifest_path": str(target_dir / "upload_manifest.json"),
    }


def plan_upload_action(filename: str, sources: list, conflict_policy: str) -> Dict[str, object]:
    existing = next((item for item in sources if str(item.get("title", "")) == filename), None)
    policy = conflict_policy if conflict_policy in {"skip", "rename", "replace", "upload_anyway"} else "rename"
    if not existing:
        return {"action": "upload", "target_name": filename, "existing": {}}
    if policy == "skip":
        return {"action": "skip", "target_name": filename, "existing": existing}
    if policy == "replace":
        return {"action": "replace", "target_name": filename, "existing": existing}
    if policy == "upload_anyway":
        return {"action": "upload_anyway", "target_name": filename, "existing": existing}

    path = Path(filename)
    stamp = time_stamp("%Y%m%d-%H%M%S")
    target_name = f"{path.stem}__copy_{stamp}{path.suffix}"
    return {"action": "rename", "target_name": target_name, "existing": existing}


def add_source_file(config: AppConfig, notebook_id: str, file_path: Path) -> Dict[str, object]:
    result = subprocess.run(
        [config.notebooklm_command, "source", "add", "-n", notebook_id, "--json", str(file_path)],
        capture_output=True,
        timeout=240,
    )
    output = result.stdout + result.stderr
    text = output.decode("utf-8", errors="replace").strip()
    parsed = {}
    if result.stdout.strip():
        try:
            parsed = _parse_notebooklm_json_output(result.stdout)
        except json.JSONDecodeError:
            parsed = {}
    return {
        "ok": result.returncode == 0,
        "returncode": result.returncode,
        "message": text[:1000],
        "source": parsed,
    }


def delete_source(config: AppConfig, notebook_id: str, source_id: str) -> Dict[str, object]:
    result = subprocess.run(
        [config.notebooklm_command, "source", "delete", "-n", notebook_id, "-y", source_id],
        capture_output=True,
        timeout=180,
    )
    output = (result.stdout + result.stderr).decode("utf-8", errors="replace").strip()
    return {"ok": result.returncode == 0, "returncode": result.returncode, "message": output[:1000]}


def execute_simple_upload(config: AppConfig, uploaded_files: list, conflict_policy: str, project_id: str = "", evidence_root: str = "") -> Dict[str, object]:
    target = resolve_notebook_target(config, notebook_id="", project_id=project_id)
    notebook_id = target["notebook_id"]
    project = target["project"]
    resolved_project_id = project_id or target.get("project_id", "")
    if not notebook_id:
        return {
            "ok": False,
            "mode": "simple_upload",
            "project_id": project_id,
            "notebook_id": "",
            "results": [],
            "error": "Select an Active Project before upload.",
        }

    source_result = list_notebook_sources(config, notebook_id)
    sources = source_result.get("sources", []) if source_result.get("ok") else []
    results = []
    manifest_records = []
    for file_info in uploaded_files:
        original_name = file_info["original_name"]
        saved_path = Path(file_info["path"])
        plan = plan_upload_action(original_name, sources, conflict_policy)
        action = plan["action"]
        target_name = str(plan["target_name"])
        target_path = saved_path

        if action == "skip":
            results.append({"filename": original_name, "status": "skipped", "action": action, "message": "Duplicate source title found.", "source": plan.get("existing", {})})
            continue
        if action == "replace":
            existing = plan.get("existing", {})
            source_id = str(existing.get("id", ""))
            delete_result = delete_source(config, notebook_id, source_id) if source_id else {"ok": False, "message": "existing source id not found"}
            if not delete_result.get("ok"):
                results.append({"filename": original_name, "status": "failed", "action": action, "message": delete_result.get("message", ""), "source": existing})
                continue

        evidence = prepare_evidence_copy(config, resolved_project_id, saved_path, target_name, evidence_root)
        target_path = Path(evidence["evidence_path"])
        upload_result = add_source_file(config, notebook_id, target_path)
        record = {
            "filename": original_name,
            "target_name": target_path.name,
            "status": "uploaded" if upload_result["ok"] else "failed",
            "action": action,
            "message": upload_result["message"],
            "source": upload_result.get("source", {}),
            "original_path": str(saved_path),
            "evidence_path": evidence["evidence_path"],
            "sha256": evidence["sha256"],
        }
        results.append(record)
        manifest_records.append(record)

    if manifest_records:
        manifest_path = Path(manifest_records[-1]["evidence_path"]).parent / "upload_manifest.json"
        manifest_payload = {
            "app": "AI NotebookLM Runtime Lab",
            "kind": "simple_upload_manifest",
            "created_at": now_iso(),
            "project_id": resolved_project_id,
            "notebook_id": notebook_id,
            "conflict_policy": conflict_policy,
            "records": manifest_records,
        }
        manifest_path.write_text(json.dumps(manifest_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    summary = {
        "ok": not any(item["status"] == "failed" for item in results),
        "mode": "simple_upload",
        "project_id": resolved_project_id,
        "project": project,
        "notebook_id": notebook_id,
        "conflict_policy": conflict_policy,
        "evidence_root": str(Path(evidence_root).expanduser()) if evidence_root else str(simple_evidence_dir(config)),
        "uploaded_count": sum(1 for item in results if item["status"] == "uploaded"),
        "skipped_count": sum(1 for item in results if item["status"] == "skipped"),
        "failed_count": sum(1 for item in results if item["status"] == "failed"),
        "results": results,
    }
    write_runtime_log(config, "simple_upload", summary)
    return summary


def read_queue_manifest(config: AppConfig) -> Dict[str, object]:
    path = config.temp_dir / "etl_queue.json"
    if not path.exists():
        return {"kind": "etl_queue_manifest", "task_count": 0, "tasks": []}
    return json.loads(path.read_text(encoding="utf-8"))


def build_runtime_state(config: AppConfig) -> Dict[str, object]:
    temp_files = [
        {
            "name": path.name,
            "path": str(path),
            "size_kb": round(path.stat().st_size / 1024, 1),
            "suffix": path.suffix.lower(),
        }
        for path in sorted(config.temp_dir.glob("*"), key=lambda item: item.name.lower())
        if path.is_file()
    ]
    log_path = logs_dir(config) / "runtime.jsonl"
    return {
        "app": "AI NotebookLM Runtime Lab",
        "kind": "runtime_state",
        "created_at": now_iso(),
        "status": build_status_payload(config),
        "temp_files": temp_files,
        "queue": read_queue_manifest(config),
        "command_queue": list_command_packages(config),
        "projects": list_projects(config, page_size=100).get("projects", []),
        "active_project": get_active_project(config),
        "log_path": str(log_path),
    }


def export_state_json(config: AppConfig) -> Path:
    path = logs_dir(config) / "runtime_state_export.json"
    path.write_text(json.dumps(build_runtime_state(config), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "export_state_json", {"path": str(path)})
    return path


def export_excel_report(config: AppConfig) -> Path:
    state = build_runtime_state(config)
    path = logs_dir(config) / "AI_ETL_Run_Report.xlsx"
    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "summary"
    _write_rows(
        summary,
        [
            ["key", "value"],
            ["app", state["app"]],
            ["created_at", state["created_at"]],
            ["project_root", state["status"]["project_root"]],
            ["queue_task_count", state["queue"].get("task_count", 0)],
        ],
    )

    env_ws = wb.create_sheet("environment")
    _write_rows(env_ws, [["item", "ok", "message", "path"]])
    for key, item in state["status"]["environment"].items():
        env_ws.append([key, item["ok"], item["message"], item["path"]])

    inbox_ws = wb.create_sheet("inbox_files")
    _write_rows(inbox_ws, [["name", "path", "size_kb", "suffix"]])
    for item in state["status"]["files"]:
        inbox_ws.append([item["name"], item["path"], item["size_kb"], item["suffix"]])

    temp_ws = wb.create_sheet("csv_outputs")
    _write_rows(temp_ws, [["name", "path", "size_kb", "suffix"]])
    for item in state["temp_files"]:
        temp_ws.append([item["name"], item["path"], item["size_kb"], item["suffix"]])

    queue_ws = wb.create_sheet("queue_tasks")
    headers = ["task_id", "target", "status", "csv_name", "csv_path", "size_kb", "created_at", "executed_at", "returncode", "error"]
    _write_rows(queue_ws, [headers])
    for task in state["queue"].get("tasks", []):
        queue_ws.append([task.get(header, "") for header in headers])

    project_ws = wb.create_sheet("projects")
    project_headers = ["project_id", "name", "notebook_id", "status", "notebook_title", "is_owner", "notebook_created_at", "updated_at", "source"]
    _write_rows(project_ws, [project_headers])
    for project in state.get("projects", []):
        project_ws.append([project.get(header, "") for header in project_headers])

    command_ws = wb.create_sheet("command_packages")
    command_headers = [
        "stage", "command_id", "cloud_task_id", "trigger_source", "trigger_mode",
        "cloud_event_type", "command_type", "submitter", "role", "target_project_id",
        "source_names", "created_at", "updated_at", "status", "valid", "errors",
    ]
    _write_rows(command_ws, [command_headers])
    for command in state.get("command_queue", {}).get("commands", []):
        row = []
        for header in command_headers:
            value = command.get(header, "")
            if header in {"source_names", "errors"}:
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)
        command_ws.append(row)

    log_ws = wb.create_sheet("runtime_logs")
    _write_rows(log_ws, [["ts", "event", "payload", "line"]])
    for record in query_runtime_logs(config, page_size=100).get("logs", []):
        log_ws.append([record.get("ts", ""), record.get("event", ""), json.dumps(record.get("payload", {}), ensure_ascii=False), record.get("_line", "")])

    rec_ws = wb.create_sheet("recommendations")
    _write_rows(
        rec_ws,
        [
            ["item", "recommendation"],
            ["notebook_id", "Portal should support list/create/select before execute."],
            ["clear_cache", "Do not clear inbox or auth state from the portal."],
            ["closed_loop", "Every run should be exportable first; import is intentionally paused for MVP governance."],
        ],
    )

    wb.save(path)
    write_runtime_log(config, "export_excel_report", {"path": str(path)})
    return path


def export_local_logs_json(config: AppConfig) -> Path:
    path = logs_dir(config) / f"local_logs_export_{time_stamp()}.json"
    payload = {
        "app": "AI NotebookLM Runtime Lab",
        "watermark": APP_WATERMARK,
        "kind": "local_logs_export",
        "exported_at": now_iso(),
        "runtime_logs": read_runtime_log_records(config),
        "command_audit_path": str(command_audit_path(config)),
    }
    if command_audit_path(config).exists():
        payload["command_audit_raw"] = command_audit_path(config).read_text(encoding="utf-8").splitlines()
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "export_local_logs_json", {"path": str(path)})
    return path


def export_local_tasks_json(config: AppConfig) -> Path:
    path = logs_dir(config) / f"local_tasks_export_{time_stamp()}.json"
    payload = {
        "app": "AI NotebookLM Runtime Lab",
        "watermark": APP_WATERMARK,
        "kind": "local_tasks_export",
        "exported_at": now_iso(),
        "queue_manifest": read_queue_manifest(config),
        "command_queue": list_command_packages(config),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "export_local_tasks_json", {"path": str(path)})
    return path


def export_local_logs_excel(config: AppConfig) -> Path:
    path = logs_dir(config) / f"local_logs_export_{time_stamp()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "runtime_logs"
    _write_rows(ws, [["ts", "event", "payload", "line"]])
    for record in read_runtime_log_records(config):
        ws.append([record.get("ts", ""), record.get("event", ""), json.dumps(record.get("payload", {}), ensure_ascii=False), record.get("_line", "")])
    audit = wb.create_sheet("command_audit")
    _write_rows(audit, [["raw_jsonl"]])
    if command_audit_path(config).exists():
        for line in command_audit_path(config).read_text(encoding="utf-8").splitlines():
            audit.append([line])
    wb.save(path)
    write_runtime_log(config, "export_local_logs_excel", {"path": str(path)})
    return path


def export_local_tasks_excel(config: AppConfig) -> Path:
    path = logs_dir(config) / f"local_tasks_export_{time_stamp()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "command_packages"
    headers = ["stage", "command_id", "cloud_task_id", "trigger_source", "trigger_mode", "cloud_event_type", "source_names", "updated_at", "valid", "errors"]
    _write_rows(ws, [headers])
    for command in list_command_packages(config).get("commands", []):
        ws.append([json.dumps(command.get(h, ""), ensure_ascii=False) if isinstance(command.get(h, ""), (list, dict)) else command.get(h, "") for h in headers])
    queue = wb.create_sheet("etl_queue")
    _write_rows(queue, [["raw_json"]])
    queue.append([json.dumps(read_queue_manifest(config), ensure_ascii=False)])
    wb.save(path)
    write_runtime_log(config, "export_local_tasks_excel", {"path": str(path)})
    return path


def clear_local_logs(config: AppConfig) -> Dict[str, object]:
    removed = []
    for path in logs_dir(config).glob("*.jsonl"):
        path.write_text("", encoding="utf-8")
        removed.append(str(path))
    return {"ok": True, "mode": "clear_local_logs", "cleared_count": len(removed), "cleared": removed}


def clear_local_tasks(config: AppConfig) -> Dict[str, object]:
    removed = []
    for stage in COMMAND_STAGES:
        for path in command_stage_dir(config, stage).glob("*.json"):
            path.unlink()
            removed.append(str(path))
    queue_path = config.temp_dir / "etl_queue.json"
    if queue_path.exists():
        queue_path.unlink()
        removed.append(str(queue_path))
    write_runtime_log(config, "clear_local_tasks", {"removed_count": len(removed)})
    return {"ok": True, "mode": "clear_local_tasks", "removed_count": len(removed), "removed": removed}


def import_queue_json(config: AppConfig, source_path: Path) -> Dict[str, object]:
    data = json.loads(source_path.read_text(encoding="utf-8"))
    tasks = data.get("tasks", [])
    imported = _normalize_imported_tasks(tasks)
    manifest = {
        "app": "AI NotebookLM Runtime Lab",
        "kind": "etl_queue_manifest",
        "created_at": now_iso(),
        "imported_from": str(source_path),
        "task_count": len(imported),
        "tasks": imported,
    }
    target = config.temp_dir / "etl_queue.json"
    target.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "import_queue_json", {"source": str(source_path), "task_count": len(imported)})
    return manifest


def import_queue_excel(config: AppConfig, source_path: Path) -> Dict[str, object]:
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        if "queue_tasks" not in wb.sheetnames:
            raise ValueError("queue_tasks sheet not found")
        ws = wb["queue_tasks"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            tasks = []
        else:
            headers = [str(value) if value is not None else "" for value in rows[0]]
            tasks = []
            for row in rows[1:]:
                item = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
                if item.get("task_id") or item.get("csv_path"):
                    tasks.append(item)
    finally:
        wb.close()
    manifest = {
        "app": "AI NotebookLM Runtime Lab",
        "kind": "etl_queue_manifest",
        "created_at": now_iso(),
        "imported_from": str(source_path),
        "task_count": len(tasks),
        "imported_count": len(tasks),
        "tasks": _normalize_imported_tasks(tasks),
    }
    target = config.temp_dir / "etl_queue.json"
    target.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "import_queue_excel", {"source": str(source_path), "task_count": len(tasks)})
    return manifest


def clear_local_working_data(config: AppConfig, scope: str) -> Dict[str, object]:
    removed = []
    allowed = {"temp", "logs", "queue", "working"}
    if scope not in allowed:
        return {"scope": scope, "removed_count": 0, "removed": [], "error": f"unsupported scope: {scope}"}

    targets = []
    if scope == "temp":
        targets = [path for path in config.temp_dir.glob("*") if path.is_file()]
    elif scope == "queue":
        targets = [config.temp_dir / "etl_queue.json"]
    elif scope == "logs":
        targets = [path for path in logs_dir(config).glob("*") if path.is_file()]
    elif scope == "working":
        targets = [path for path in config.temp_dir.glob("*") if path.is_file()]
        targets += [path for path in logs_dir(config).glob("*") if path.is_file()]

    for path in targets:
        if path.exists() and path.is_file():
            path.unlink()
            removed.append(str(path))
    if scope not in {"logs", "working"}:
        write_runtime_log(config, "clear_local_working_data", {"scope": scope, "removed_count": len(removed)})
    return {"scope": scope, "removed_count": len(removed), "removed": removed}


def _write_rows(ws, rows):
    for row in rows:
        ws.append(row)


def _normalize_imported_tasks(tasks):
    normalized = []
    for index, task in enumerate(tasks, 1):
        csv_path = str(task.get("csv_path", "") or "")
        normalized.append(
            {
                "task_id": str(task.get("task_id") or f"csv-{index:04d}"),
                "target": str(task.get("target") or "notebooklm_source_add"),
                "status": str(task.get("status") or "pending"),
                "csv_name": str(task.get("csv_name") or Path(csv_path).name),
                "csv_path": csv_path,
                "size_kb": task.get("size_kb", ""),
                "created_at": str(task.get("created_at") or now_iso()),
            }
        )
    return normalized


def build_status_payload(config: AppConfig) -> Dict[str, object]:
    files = []
    for path in get_valid_files(config.source_dir):
        files.append(
            {
                "name": path.name,
                "path": str(path),
                "size_kb": round(path.stat().st_size / 1024, 1),
                "suffix": path.suffix.lower(),
            }
        )
    return {
        "app": "AI NotebookLM Runtime Lab",
        "version": APP_VERSION,
        "mode": "local runtime lab",
        "strategy": "html_portal",
        "project_root": str(config.project_root),
        "runtime_identity": build_runtime_identity(config),
        "environment": check_environment(config),
        "files": files,
        "notes": [
            "HTML is the portal; Python is the runtime.",
            "Document Manager Windows notebooklm.exe remains the first-stage target.",
            "Mac notebooklm-py is compatibility research for later.",
        ],
    }


def build_runtime_identity(config: AppConfig) -> Dict[str, object]:
    # Falo x Force 教學註解：
    # Win / Mac 混合部署時，使用者最容易搞混「現在是哪台主機在跑」。
    # 所以 status 與 Portal 都要明確顯示 hostname、OS、執行路徑與 URL。
    hostname = socket.gethostname()
    computer_name = os.environ.get("COMPUTERNAME") or os.environ.get("HOSTNAME") or hostname
    os_name = platform.system() or "unknown"
    os_label = {
        "Darwin": "macOS",
        "Windows": "Windows",
        "Linux": "Linux",
    }.get(os_name, os_name)
    lan_ip = detect_lan_ip()
    local_url = f"http://127.0.0.1:{RUNTIME_BIND_PORT}"
    lan_url = f"http://{lan_ip}:{RUNTIME_BIND_PORT}" if RUNTIME_BIND_HOST == "0.0.0.0" else ""
    return {
        "watermark": APP_WATERMARK,
        "computer_name": computer_name,
        "hostname": hostname,
        "os": os_label,
        "platform": platform.platform(),
        "machine": platform.machine(),
        "python_version": platform.python_version(),
        "python_executable": sys.executable,
        "project_root": str(config.project_root),
        "bind_host": RUNTIME_BIND_HOST,
        "port": RUNTIME_BIND_PORT,
        "local_url": local_url,
        "lan_ip": lan_ip,
        "lan_url": lan_url,
        "worker_role": "local_runtime_worker",
        "pid": os.getpid(),
    }


def read_gas_settings(config: AppConfig) -> Dict[str, object]:
    defaults = {
        "enabled": False,
        "auto_poll_enabled": False,
        "web_app_url": "https://script.google.com/macros/s/AKfycbw9X3Y6MQ2XpvsS9BXuCZeZsVkrbT1VL0JkDkotrbs-omYG8OpuWpAl1fowiJa_QW1i/exec",
        "api_token": "123456",
        "poll_interval_seconds": DEFAULT_GAS_POLL_INTERVAL_SECONDS,
        "max_tasks_per_poll": DEFAULT_GAS_MAX_TASKS_PER_POLL,
        "auto_execute": False,
        "local_download_root": "data/gas_downloads",
        "default_file_types": [".pdf", ".md", ".csv", ".docx", ".xlsx", ".pptx", ".png", ".jpg", ".jpeg"],
        "default_duplicate_policy": "rename",
    }
    path = config.project_root / "config" / "gas_config.json"
    if not path.exists():
        return defaults
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        data = {}
    return {**defaults, **data} if isinstance(data, dict) else defaults


def write_gas_settings(config: AppConfig, settings: Dict[str, object]) -> Path:
    path = config.project_root / "config" / "gas_config.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    settings["updated_at"] = now_iso()
    path.write_text(json.dumps(settings, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "gas_settings_updated", {key: value for key, value in settings.items() if key != "api_token"})
    return path


def update_gas_settings_from_params(config: AppConfig, params: Dict[str, list]) -> Dict[str, object]:
    # Falo x Force 教學註解：
    # 本機 GUI 寫入 config/gas_config.json，讓非工程使用者不用改 JSON。
    # 真正的 token 仍只留在本機設定檔，不顯示在頁面上。
    settings = read_gas_settings(config)
    for key in ("enabled", "auto_poll_enabled", "auto_execute"):
        if key in params:
            settings[key] = params.get(key, [""])[0].lower() in {"1", "yes", "on", "true"}
    try:
        settings["poll_interval_seconds"] = max(
            MIN_GAS_POLL_INTERVAL_SECONDS,
            int(params.get("poll_interval_seconds", [str(DEFAULT_GAS_POLL_INTERVAL_SECONDS)])[0] or DEFAULT_GAS_POLL_INTERVAL_SECONDS),
        )
    except ValueError:
        settings["poll_interval_seconds"] = DEFAULT_GAS_POLL_INTERVAL_SECONDS
    try:
        settings["max_tasks_per_poll"] = max(
            1,
            min(50, int(params.get("max_tasks_per_poll", [str(DEFAULT_GAS_MAX_TASKS_PER_POLL)])[0] or DEFAULT_GAS_MAX_TASKS_PER_POLL)),
        )
    except ValueError:
        settings["max_tasks_per_poll"] = DEFAULT_GAS_MAX_TASKS_PER_POLL
    web_app_url = params.get("web_app_url", [""])[0].strip()
    if web_app_url:
        settings["web_app_url"] = web_app_url
    api_token = params.get("api_token", [""])[0].strip()
    if api_token:
        settings["api_token"] = api_token
    local_download_root = params.get("local_download_root", [""])[0].strip()
    if local_download_root:
        settings["local_download_root"] = local_download_root
    write_gas_settings(config, settings)
    return {"ok": True, "mode": "gas_settings", "settings": {key: value for key, value in settings.items() if key != "api_token"}}


def apply_gas_safe_default(config: AppConfig) -> Dict[str, object]:
    # Falo x Force 教學註解：
    # Safe Default 是「回到治理預設」的救援按鈕，不等於手動表單儲存。
    # 它保留 Web App URL / token / 下載資料夾，只重設節流與自動化核心參數。
    # 治理預設只開「可連線」，不主動 poll、不主動 execute，避免一啟動就批次跑任務。
    settings = read_gas_settings(config)
    settings["enabled"] = True
    settings["auto_poll_enabled"] = False
    settings["auto_execute"] = False
    settings["poll_interval_seconds"] = DEFAULT_GAS_POLL_INTERVAL_SECONDS
    settings["max_tasks_per_poll"] = DEFAULT_GAS_MAX_TASKS_PER_POLL
    write_gas_settings(config, settings)
    return {
        "ok": True,
        "mode": "gas_safe_default",
        "message": "Applied Falo safe default: GAS enabled, auto polling off, auto execute off, 600 seconds / 3 tasks.",
        "settings": {key: value for key, value in settings.items() if key != "api_token"},
    }


def toggle_gas_setting(config: AppConfig, key: str) -> Dict[str, object]:
    allowed = {
        "enabled": "GAS connection",
        "auto_poll_enabled": "Auto polling",
        "auto_execute": "Auto execute",
    }
    if key not in allowed:
        return {"ok": False, "mode": "gas_toggle", "error": f"unsupported GAS toggle: {key}"}
    settings = read_gas_settings(config)
    settings[key] = not bool(settings.get(key, False))
    write_gas_settings(config, settings)
    return {
        "ok": True,
        "mode": "gas_toggle",
        "key": key,
        "label": allowed[key],
        "enabled": bool(settings[key]),
        "settings": {item_key: value for item_key, value in settings.items() if item_key != "api_token"},
    }


def read_gas_settings_result(config: AppConfig) -> Dict[str, object]:
    settings = read_gas_settings(config)
    return {
        "ok": True,
        "mode": "gas_settings_read",
        "note": "read only; no config was changed",
        "settings": {key: value for key, value in settings.items() if key != "api_token"},
    }



import threading
gas_history_lock = threading.Lock()

def get_gas_url_history(config: AppConfig) -> list:
    history_file = Path(config.project_root) / "data" / "gas_url_history.json"
    history_file.parent.mkdir(parents=True, exist_ok=True)
    with gas_history_lock:
        if not history_file.exists():
            return []
        try:
            return json.loads(history_file.read_text(encoding="utf-8"))
        except Exception:
            return []

def save_gas_url_history(config: AppConfig, history: list) -> None:
    history_file = Path(config.project_root) / "data" / "gas_url_history.json"
    history_file.parent.mkdir(parents=True, exist_ok=True)
    with gas_history_lock:
        history_file.write_text(json.dumps(history, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def gas_auto_worker_loop(config: AppConfig) -> None:
    # Falo x Force 教學註解：
    # 這是地端自動輪詢 worker。它每 N 秒醒來一次，讀取最新設定。
    # 若 auto_poll_enabled 關閉，就只睡覺；若開啟，就執行一次 GAS poll。
    while True:
        settings = read_gas_settings(config)
        interval = max(
            MIN_GAS_POLL_INTERVAL_SECONDS,
            int(settings.get("poll_interval_seconds") or DEFAULT_GAS_POLL_INTERVAL_SECONDS),
        )
        if settings.get("enabled") and settings.get("auto_poll_enabled"):
            try:
                from gas_adapter import push_host_info_to_gas
                push_res = push_host_info_to_gas(config, "scheduled")
                write_runtime_log(
                    config,
                    "gas_auto_push_host_info",
                    {
                        "ok": push_res.get("ok", False),
                        "error": push_res.get("error", "")
                    }
                )
            except Exception as exc:
                write_runtime_log(config, "gas_auto_push_host_info_error", {"error": str(exc)})
            
            # Also run task poll
            try:
                from gas_adapter import poll_gas_once

                result = poll_gas_once(config)
                write_runtime_log(
                    config,
                    "gas_auto_poll_tick",
                    {
                        "ok": result.get("ok", False),
                        "task_count": result.get("task_count", 0),
                        "created_count": len(result.get("created_commands", [])),
                    },
                )
            except Exception as exc:
                write_runtime_log(config, "gas_auto_poll_error", {"error": str(exc)})
        time.sleep(interval)


def ensure_gas_auto_worker(config: AppConfig) -> None:
    global _GAS_AUTO_WORKER_STARTED
    if _GAS_AUTO_WORKER_STARTED:
        return
    thread = threading.Thread(target=gas_auto_worker_loop, args=(config,), daemon=True)
    thread.start()
    _GAS_AUTO_WORKER_STARTED = True
    write_runtime_log(config, "gas_auto_worker_started", {"thread": thread.name})


def incoming_watch_worker_loop(config: AppConfig) -> None:
    # Falo x Force 教學註解：
    # 這是本機資料夾 trigger。它每 N 秒掃 incoming folder，
    # 找到新檔案後只產生 command package，不直接跳過治理層。
    while True:
        settings = read_runtime_settings(config)
        interval = max(
            MIN_INCOMING_WATCH_INTERVAL_SECONDS,
            int(settings.get("incoming_watch_interval_seconds") or DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS),
        )
        if settings.get("incoming_watch_enabled") and str(settings.get("incoming_watch_mode", "polling")) == "polling":
            try:
                result = incoming_watch_scan_once(config)
                write_runtime_log(
                    config,
                    "incoming_watch_tick",
                    {
                        "ok": result.get("ok", False),
                        "new_count": result.get("new_count", 0),
                        "folder": result.get("folder", ""),
                    },
                )
            except Exception as exc:
                write_runtime_log(config, "incoming_watch_error", {"error": str(exc)})
        time.sleep(interval)


def ensure_incoming_watch_worker(config: AppConfig) -> None:
    global _INCOMING_WATCH_WORKER_STARTED
    if _INCOMING_WATCH_WORKER_STARTED:
        return
    thread = threading.Thread(target=incoming_watch_worker_loop, args=(config,), daemon=True)
    thread.start()
    _INCOMING_WATCH_WORKER_STARTED = True
    write_runtime_log(config, "incoming_watch_worker_started", {"thread": thread.name})


def incoming_realtime_watch_worker_loop(config: AppConfig) -> None:
    # Falo x Force 教學註解：
    # Realtime Watch 是本機工作站才有的進階模式。它用 watchdog 接作業系統資料夾事件；
    # 事件本身只當「提醒」，真正處理仍回到 incoming_watch_scan_once，維持同一條治理流程。
    observer = None
    observer_key = ""
    pending_events = []
    unavailable_logged = False

    while True:
        settings = read_runtime_settings(config)
        enabled = bool(settings.get("incoming_watch_enabled")) and str(settings.get("incoming_watch_mode", "polling")) == "realtime"
        folder = Path(str(settings.get("incoming_watch_folder") or simple_incoming_dir(config))).expanduser()
        if not folder.is_absolute():
            folder = (config.project_root / folder).resolve()
        recursive = bool(settings.get("incoming_watch_recursive", False))
        key = f"{folder}|{recursive}"

        if not enabled:
            if observer:
                observer.stop()
                observer.join(timeout=5)
                observer = None
                observer_key = ""
                write_runtime_log(config, "incoming_realtime_watch_stopped", {"reason": "disabled"})
            time.sleep(2)
            continue

        try:
            from watchdog.events import FileSystemEventHandler  # type: ignore
            from watchdog.observers import Observer  # type: ignore
        except Exception as exc:
            if not unavailable_logged:
                write_runtime_log(config, "incoming_realtime_watch_unavailable", {"error": str(exc), "hint": "Install watchdog to enable true folder events."})
                unavailable_logged = True
            time.sleep(30)
            continue

        if (observer is None or observer_key != key) and folder.exists() and folder.is_dir():
            if observer:
                observer.stop()
                observer.join(timeout=5)

            class IncomingEventHandler(FileSystemEventHandler):
                def on_created(self, event):  # type: ignore[no-untyped-def]
                    if not event.is_directory:
                        pending_events.append({"event": "created", "path": str(event.src_path), "ts": time.time()})

                def on_modified(self, event):  # type: ignore[no-untyped-def]
                    if not event.is_directory:
                        pending_events.append({"event": "modified", "path": str(event.src_path), "ts": time.time()})

                def on_moved(self, event):  # type: ignore[no-untyped-def]
                    if not event.is_directory:
                        pending_events.append({"event": "moved", "path": str(getattr(event, "dest_path", event.src_path)), "ts": time.time()})

            observer = Observer()
            observer.schedule(IncomingEventHandler(), str(folder), recursive=recursive)
            observer.start()
            observer_key = key
            write_runtime_log(config, "incoming_realtime_watch_started", {"folder": str(folder), "recursive": recursive})

        min_age = max(0, int(settings.get("incoming_watch_min_age_seconds") or 10))
        if pending_events:
            latest = max(float(item.get("ts", 0)) for item in pending_events)
            if time.time() - latest >= min_age:
                event_count = len(pending_events)
                event_paths = sorted({str(item.get("path", "")) for item in pending_events if item.get("path")})
                pending_events.clear()
                try:
                    result = incoming_watch_scan_once(config)
                    write_runtime_log(
                        config,
                        "incoming_realtime_watch_tick",
                        {
                            "event_count": event_count,
                            "event_paths": event_paths[:20],
                            "new_count": result.get("new_count", 0),
                            "ok": result.get("ok", False),
                        },
                    )
                except Exception as exc:
                    write_runtime_log(config, "incoming_realtime_watch_error", {"error": str(exc), "event_count": event_count})
        time.sleep(1)


def ensure_incoming_realtime_watch_worker(config: AppConfig) -> None:
    global _INCOMING_REALTIME_WORKER_STARTED
    if _INCOMING_REALTIME_WORKER_STARTED:
        return
    thread = threading.Thread(target=incoming_realtime_watch_worker_loop, args=(config,), daemon=True)
    thread.start()
    _INCOMING_REALTIME_WORKER_STARTED = True
    write_runtime_log(config, "incoming_realtime_watch_worker_started", {"thread": thread.name})


def _parse_notebooklm_json_output(output: bytes) -> Dict[str, object]:
    text = output.decode("utf-8", errors="replace").strip()
    if not text:
        return {}
    return json.loads(text)


def list_notebooks(config: AppConfig) -> Dict[str, object]:
    result = subprocess.run(
        [config.notebooklm_command, "list", "--json"],
        capture_output=True,
        timeout=180,
    )
    output = result.stdout + result.stderr
    if result.returncode != 0:
        payload = {
            "mode": "list_notebooks",
            "ok": False,
            "returncode": result.returncode,
            "notebooks": [],
            "error": output.decode("utf-8", errors="replace").strip()[:1000],
        }
        write_runtime_log(config, "list_notebooks_failed", {"returncode": result.returncode})
        return payload

    data = _parse_notebooklm_json_output(result.stdout)
    notebooks = data.get("notebooks", []) if isinstance(data, dict) else []
    payload = {
        "mode": "list_notebooks",
        "ok": True,
        "returncode": result.returncode,
        "count": len(notebooks),
        "notebooks": notebooks,
    }
    write_runtime_log(config, "list_notebooks", {"count": len(notebooks)})
    return payload


def create_notebook(config: AppConfig, title: str) -> Dict[str, object]:
    clean_title = title.strip()
    if not clean_title:
        return {
            "mode": "create_notebook",
            "ok": False,
            "title": "",
            "notebook": {},
            "error": "title is required",
        }

    result = subprocess.run(
        [config.notebooklm_command, "create", clean_title, "--json"],
        capture_output=True,
        timeout=180,
    )
    output = result.stdout + result.stderr
    if result.returncode != 0:
        payload = {
            "mode": "create_notebook",
            "ok": False,
            "title": clean_title,
            "returncode": result.returncode,
            "notebook": {},
            "error": output.decode("utf-8", errors="replace").strip()[:1000],
        }
        write_runtime_log(config, "create_notebook_failed", {"title": clean_title, "returncode": result.returncode})
        return payload

    data = _parse_notebooklm_json_output(result.stdout)
    notebook = data.get("notebook", data) if isinstance(data, dict) else {}
    payload = {
        "mode": "create_notebook",
        "ok": True,
        "title": clean_title,
        "returncode": result.returncode,
        "notebook": notebook,
    }
    if isinstance(notebook, dict) and notebook.get("id"):
        sync_projects_from_notebooks(config, [notebook])
    write_runtime_log(config, "create_notebook", {"title": clean_title, "notebook": notebook})
    return payload


def normalize_inbox_files(config: AppConfig) -> Dict[str, object]:
    outputs = []
    errors = []
    skipped = []

    for path in get_valid_files(config.source_dir):
        if path.suffix.lower() != ".xlsx":
            skipped.append({"name": path.name, "reason": "not xlsx"})
            continue
        try:
            for csv_path in xlsx_to_csvs(path, config.temp_dir):
                outputs.append(
                    {
                        "source": path.name,
                        "name": csv_path.name,
                        "path": str(csv_path),
                        "size_kb": round(csv_path.stat().st_size / 1024, 1),
                    }
                )
        except Exception as exc:
            errors.append({"source": path.name, "error": str(exc)})

    result = {
        "converted_count": len(outputs),
        "outputs": outputs,
        "errors": errors,
        "skipped": skipped,
    }
    write_runtime_log(config, "normalize_inbox_files", {"converted_count": len(outputs), "error_count": len(errors)})
    return result


def build_queue_manifest(config: AppConfig) -> Dict[str, object]:
    csv_files = sorted(config.temp_dir.glob("*.csv"), key=lambda path: path.name.lower())
    now = now_iso()
    tasks = []
    for index, path in enumerate(csv_files, 1):
        tasks.append(
            {
                "task_id": f"csv-{index:04d}",
                "target": "notebooklm_source_add",
                "status": "pending",
                "csv_name": path.name,
                "csv_path": str(path),
                "size_kb": round(path.stat().st_size / 1024, 1),
                "created_at": now,
                "note": "Prepared by AI NotebookLM Runtime Lab ETL queue.",
            }
        )

    manifest = {
        "app": "AI NotebookLM Runtime Lab",
        "kind": "etl_queue_manifest",
        "created_at": now,
        "task_count": len(tasks),
        "tasks": tasks,
    }
    queue_path = config.temp_dir / "etl_queue.json"
    queue_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "build_queue_manifest", {"task_count": len(tasks), "path": str(queue_path)})
    return manifest


def resolve_notebook_target(config: AppConfig, notebook_id: str, project_id: str = "") -> Dict[str, object]:
    if not project_id:
        project_id = str(get_active_project(config).get("active_project_id", ""))
    project = find_project(config, project_id) if project_id else {}
    resolved = str(notebook_id or "").strip()
    if project and not resolved:
        resolved = str(project.get("notebook_id", "")).strip()
    return {"notebook_id": resolved, "project_id": project_id, "project": project}


def preview_notebooklm_adapter(config: AppConfig, notebook_id: str, project_id: str = "") -> Dict[str, object]:
    target = resolve_notebook_target(config, notebook_id, project_id)
    notebook_id = target["notebook_id"]
    queue_path = config.temp_dir / "etl_queue.json"
    if not queue_path.exists():
        return {
            "mode": "dry_run",
            "notebook_id": notebook_id,
            "project_id": project_id,
            "project": target["project"],
            "command_count": 0,
            "commands": [],
            "errors": [{"message": "etl_queue.json not found. Build Queue Manifest first."}],
        }

    manifest = json.loads(queue_path.read_text(encoding="utf-8"))
    commands = []
    errors = []
    for task in manifest.get("tasks", []):
        if task.get("status") != "pending":
            continue
        csv_path = task.get("csv_path", "")
        if not csv_path or not Path(csv_path).exists():
            errors.append({"task_id": task.get("task_id", ""), "message": "CSV path not found"})
            continue
        args = ["source", "add", "-n", notebook_id, csv_path]
        commands.append(
            {
                "task_id": task.get("task_id", ""),
                "csv_name": task.get("csv_name", ""),
                "command": " ".join([config.notebooklm_command] + args),
                "args": args,
            }
        )

    return {
        "mode": "dry_run",
        "notebook_id": notebook_id,
        "project_id": project_id,
        "project": target["project"],
        "command_count": len(commands),
        "commands": commands,
        "errors": errors,
    }


def execute_notebooklm_adapter(config: AppConfig, notebook_id: str, confirm: bool, project_id: str = "") -> Dict[str, object]:
    target = resolve_notebook_target(config, notebook_id, project_id)
    notebook_id = target["notebook_id"]
    if not confirm:
        write_runtime_log(config, "adapter_execute_blocked", {"notebook_id": notebook_id, "project_id": project_id})
        return {
            "mode": "blocked",
            "notebook_id": notebook_id,
            "project_id": project_id,
            "project": target["project"],
            "executed_count": 0,
            "results": [],
            "errors": [{"message": "Execution requires confirm=yes."}],
        }

    queue_path = config.temp_dir / "etl_queue.json"
    if not queue_path.exists():
        return {
            "mode": "execute",
            "notebook_id": notebook_id,
            "project_id": project_id,
            "project": target["project"],
            "executed_count": 0,
            "results": [],
            "errors": [{"message": "etl_queue.json not found. Build Queue Manifest first."}],
        }
    if not notebook_id:
        return {
            "mode": "execute",
            "notebook_id": "",
            "project_id": project_id,
            "project": target["project"],
            "executed_count": 0,
            "results": [],
            "errors": [{"message": "Notebook target is required. Select a project or enter notebook_id."}],
        }

    manifest = json.loads(queue_path.read_text(encoding="utf-8"))
    executed = []
    errors = []
    for task in manifest.get("tasks", []):
        if task.get("status") != "pending":
            continue
        csv_path = task.get("csv_path", "")
        if not csv_path or not Path(csv_path).exists():
            task["status"] = "failed"
            task["error"] = "CSV path not found"
            errors.append({"task_id": task.get("task_id", ""), "message": task["error"]})
            continue
        args = ["source", "add", "-n", notebook_id, csv_path]
        result = subprocess.run(
            [config.notebooklm_command] + args,
            capture_output=True,
            timeout=180,
        )
        output = (result.stdout + result.stderr).decode("utf-8", errors="replace").strip()
        if result.returncode == 0:
            task["status"] = "uploaded"
            task["project_id"] = project_id
            task["notebook_id"] = notebook_id
        else:
            task["status"] = "failed"
            task["error"] = output[:500]
            errors.append({"task_id": task.get("task_id", ""), "message": output[:500]})
        task["executed_at"] = now_iso()
        task["returncode"] = result.returncode
        executed.append(
            {
                "task_id": task.get("task_id", ""),
                "csv_name": task.get("csv_name", ""),
                "status": task["status"],
                "returncode": result.returncode,
                "message": output[:300],
            }
        )

    manifest["updated_at"] = now_iso()
    queue_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_runtime_log(config, "execute_notebooklm_adapter", {"notebook_id": notebook_id, "project_id": project_id, "executed_count": len(executed), "error_count": len(errors)})
    return {
        "mode": "execute",
        "notebook_id": notebook_id,
        "project_id": project_id,
        "project": target["project"],
        "executed_count": len(executed),
        "results": executed,
        "errors": errors,
    }


def find_available_port(preferred_port: int, host: str = "127.0.0.1") -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        try:
            sock.bind((host, preferred_port))
            return preferred_port
        except OSError:
            pass

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind((host, 0))
        return int(sock.getsockname()[1])


def detect_lan_ip() -> str:
    # Falo x Force 教學註解：
    # 0.0.0.0 是「對所有網卡開放」，但使用者不能在瀏覽器輸入 0.0.0.0。
    # 所以啟動時順便推測一個區網 IP，方便同網段手機或其他電腦測試。
    with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
        try:
            sock.connect(("8.8.8.8", 80))
            return str(sock.getsockname()[0])
        except OSError:
            return "127.0.0.1"


def handle_simple_upload_post(config: AppConfig, handler: BaseHTTPRequestHandler) -> Dict[str, object]:
    content_type = handler.headers.get("Content-Type", "")
    if "multipart/form-data" not in content_type:
        return {
            "ok": False,
            "mode": "simple_upload",
            "project_id": "",
            "notebook_id": "",
            "results": [],
            "error": "Upload form must use multipart/form-data.",
        }
    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": content_type,
            "CONTENT_LENGTH": handler.headers.get("Content-Length", "0"),
        },
    )
    project_id = str(form.getfirst("project_id", "") or "").strip()
    conflict_policy = str(form.getfirst("conflict_policy", "rename") or "rename").strip()
    evidence_root = str(form.getfirst("evidence_root", "") or "").strip()
    file_fields = form["files"] if "files" in form else []
    if not isinstance(file_fields, list):
        file_fields = [file_fields]

    uploaded = []
    session_dir = config.source_dir / "_simple_uploads" / time_stamp("%Y%m%d-%H%M%S")
    for field in file_fields:
        if not getattr(field, "filename", ""):
            continue
        filename = Path(field.filename).name
        if not filename:
            continue
        target = unique_upload_path(session_dir, filename)
        with target.open("wb") as handle:
            while True:
                chunk = field.file.read(1024 * 1024)
                if not chunk:
                    break
                handle.write(chunk)
        uploaded.append({"original_name": filename, "path": str(target), "size_kb": round(target.stat().st_size / 1024, 1)})

    if not uploaded:
        return {
            "ok": False,
            "mode": "simple_upload",
            "project_id": project_id,
            "notebook_id": "",
            "results": [],
            "error": "No files received.",
        }
    return execute_simple_upload(config, uploaded, conflict_policy, project_id=project_id, evidence_root=evidence_root)


def unique_upload_path(folder: Path, filename: str) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / filename
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(1, 1000):
        candidate = folder / f"{stem}__local_{index}{suffix}"
        if not candidate.exists():
            return candidate
    stamp = time_stamp("%Y%m%d-%H%M%S")
    return folder / f"{stem}__local_{stamp}{suffix}"


def execute_folder_upload(config: AppConfig, params: Dict[str, list]) -> Dict[str, object]:
    folder_path = params.get("folder_path", [str(simple_incoming_dir(config))])[0]
    extensions = parse_extensions(params.get("types", []))
    recursive = params.get("recursive", [""])[0].lower() in {"1", "yes", "on", "true"}
    order = params.get("order", ["name"])[0]
    project_id = params.get("project_id", [""])[0].strip()
    conflict_policy = params.get("conflict_policy", ["rename"])[0].strip()
    evidence_root = params.get("evidence_root", [""])[0].strip()
    scan = scan_source_folder(config, folder_path, extensions, recursive, order)
    if not scan.get("ok"):
        return {"ok": False, "mode": "simple_upload", "project_id": project_id, "notebook_id": "", "results": [], "error": scan.get("error", "")}
    uploaded_files = [{"original_name": item["name"], "path": item["path"], "size_kb": item["size_kb"]} for item in scan.get("files", [])]
    if not uploaded_files:
        return {"ok": False, "mode": "simple_upload", "project_id": project_id, "notebook_id": "", "results": [], "error": "No files matched the scan settings."}
    return execute_simple_upload(config, uploaded_files, conflict_policy, project_id=project_id, evidence_root=evidence_root)


LOCAL_LOGIN_PAGE_HTML = """<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>FALO Local Web Gateway - Login</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
  <style>
    :root {
      --bg: #0b0f19;
      --card-bg: rgba(17, 24, 39, 0.75);
      --border: rgba(255, 255, 255, 0.08);
      --text: #f3f4f6;
      --text-muted: #9ca3af;
      --primary: #10b981;
      --primary-hover: #059669;
      --accent: #f97316;
      --accent-hover: #ea580c;
      --danger: #ef4444;
      --glow: rgba(16, 185, 129, 0.15);
    }
    body {
      margin: 0;
      background: radial-gradient(circle at top, #111827 0%, #030712 100%);
      color: var(--text);
      font-family: 'Outfit', -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      min-height: 100vh;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 20px;
      box-sizing: border-box;
    }
    .card {
      background: var(--card-bg);
      border: 1px solid var(--border);
      border-radius: 16px;
      padding: 30px;
      width: 100%;
      max-width: 400px;
      box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.3), 0 8px 10px -6px rgba(0, 0, 0, 0.3);
      backdrop-filter: blur(12px);
    }
    h2 {
      margin-top: 0;
      font-weight: 600;
      color: #fff;
      text-align: center;
      letter-spacing: 0.5px;
    }
    p.subtitle {
      color: var(--text-muted);
      font-size: 0.88rem;
      text-align: center;
      margin-bottom: 24px;
      line-height: 1.5;
    }
    .field {
      margin-bottom: 20px;
    }
    .field label {
      display: block;
      font-size: 0.9rem;
      font-weight: 600;
      margin-bottom: 8px;
      color: var(--text-muted);
    }
    .input-wrapper {
      position: relative;
      display: flex;
      align-items: center;
      width: 100%;
    }
    input {
      width: 100%;
      box-sizing: border-box;
      background: rgba(0, 0, 0, 0.35);
      border: 1px solid var(--border);
      border-radius: 8px;
      color: #fff;
      padding: 12px 12px;
      font-size: 1rem;
      outline: none;
      transition: border-color 0.2s, box-shadow 0.2s;
    }
    input:focus {
      border-color: var(--primary);
      box-shadow: 0 0 0 2px var(--glow);
    }
    .eye-btn {
      position: absolute;
      right: 12px;
      background: none;
      border: none;
      color: var(--text-muted);
      cursor: pointer;
      font-size: 1.1rem;
      padding: 4px;
      display: flex;
      align-items: center;
      justify-content: center;
      outline: none;
    }
    .eye-btn:hover {
      color: #fff;
    }
    button[type="submit"], .btn-submit {
      width: 100%;
      background: var(--primary);
      color: #fff;
      border: none;
      border-radius: 8px;
      padding: 12px;
      font-size: 1rem;
      font-weight: 600;
      cursor: pointer;
      transition: background-color 0.2s;
    }
    button[type="submit"]:hover, .btn-submit:hover {
      background: var(--primary-hover);
    }
    .divider {
      display: flex;
      align-items: center;
      text-align: center;
      margin: 20px 0;
      color: var(--text-muted);
      font-size: 0.8rem;
    }
    .divider::before, .divider::after {
      content: '';
      flex: 1;
      border-bottom: 1px solid var(--border);
    }
    .divider:not(:empty)::before {
      margin-right: .5em;
    }
    .divider:not(:empty)::after {
      margin-left: .5em;
    }
    .btn-anon {
      width: 100%;
      background: rgba(255, 255, 255, 0.05);
      color: #fff;
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 12px;
      font-size: 1rem;
      font-weight: 600;
      cursor: pointer;
      transition: all 0.2s;
    }
    .btn-anon:hover {
      background: rgba(255, 255, 255, 0.1);
      border-color: var(--accent);
      color: var(--accent);
    }
    .error-msg {
      color: var(--danger);
      font-weight: 600;
      text-align: center;
      margin-top: 15px;
      font-size: 0.9rem;
      min-height: 20px;
    }
    
    /* Modal styles */
    .modal-overlay {
      position: fixed;
      top: 0;
      left: 0;
      width: 100%;
      height: 100%;
      background: rgba(0, 0, 0, 0.6);
      backdrop-filter: blur(8px);
      display: flex;
      align-items: center;
      justify-content: center;
      z-index: 1000;
      opacity: 0;
      pointer-events: none;
      transition: opacity 0.3s ease;
    }
    .modal-overlay.active {
      opacity: 1;
      pointer-events: auto;
    }
    .modal-card {
      background: #111827;
      border: 1px solid var(--border);
      border-radius: 16px;
      padding: 30px;
      width: 100%;
      max-width: 380px;
      box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.5);
      transform: scale(0.9);
      transition: transform 0.3s ease;
    }
    .modal-overlay.active .modal-card {
      transform: scale(1);
    }
    .modal-title {
      font-size: 1.25rem;
      font-weight: 600;
      color: #fff;
      margin-top: 0;
      margin-bottom: 12px;
      text-align: center;
    }
    .modal-desc {
      color: var(--text-muted);
      font-size: 0.85rem;
      margin-bottom: 20px;
      text-align: center;
      line-height: 1.5;
    }
    .modal-note {
      background: rgba(249, 115, 22, 0.1);
      border-left: 3px solid var(--accent);
      padding: 8px 12px;
      border-radius: 4px;
      color: var(--accent);
      font-size: 0.8rem;
      margin-bottom: 20px;
    }
    .modal-actions {
      display: flex;
      gap: 10px;
    }
    .btn-cancel {
      flex: 1;
      background: rgba(255, 255, 255, 0.05);
      color: #fff;
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 12px;
      font-size: 1rem;
      font-weight: 600;
      cursor: pointer;
      transition: background-color 0.2s;
    }
    .btn-cancel:hover {
      background: rgba(255, 255, 255, 0.1);
    }
    .btn-verify {
      flex: 1;
      background: var(--accent);
      color: #fff;
      border: none;
      border-radius: 8px;
      padding: 12px;
      font-size: 1rem;
      font-weight: 600;
      cursor: pointer;
      transition: background-color 0.2s;
    }
    .btn-verify:hover {
      background: var(--accent-hover);
    }
  </style>
</head>
<body>
  <div class="card">
    <h2>地端管理閘道驗證</h2>
    <p class="subtitle">請登入您的使用者帳號與密碼，或選擇匿名登入。</p>
    <form onsubmit="handleCredentialsLogin(event)">
      <div class="field">
        <label for="username">使用者帳號 (Username)</label>
        <input type="text" id="username" placeholder="請輸入帳號" required autofocus>
      </div>
      <div class="field">
        <label for="password">安全密碼 (Password)</label>
        <div class="input-wrapper">
          <input type="password" id="password" placeholder="請輸入密碼" required style="padding-right: 42px;">
          <button type="button" class="eye-btn" onclick="togglePass()" title="顯示/隱藏密碼">👁️</button>
        </div>
      </div>
      <button type="submit">驗證並登入</button>
      <div id="error" class="error-msg"></div>
    </form>
    
    <div class="divider">或</div>
    
    <button class="btn-anon" onclick="openKeyModal()">👤 匿名登入</button>
  </div>
  
  <!-- Anonymous Key Modal -->
  <div id="key-modal" class="modal-overlay">
    <div class="modal-card">
      <h3 class="modal-title">匿名存取金鑰驗證</h3>
      <p class="modal-desc">請在下方輸入配發給您的存取金鑰進行匿名登入。</p>
      <div class="modal-note">
        <strong>備註：</strong>請提供你的一次性金鑰
      </div>
      <div class="field" style="margin-bottom: 24px;">
        <input type="text" id="key-token" placeholder="請輸入 6 位數字金鑰" maxlength="6" style="text-align: center; font-family: monospace; letter-spacing: 2px;">
      </div>
      <div class="modal-actions">
        <button class="btn-cancel" onclick="closeKeyModal()">取消</button>
        <button class="btn-verify" onclick="handleKeyLogin()">驗證登入</button>
      </div>
      <div id="modal-error" class="error-msg" style="margin-top: 15px;"></div>
    </div>
  </div>

  <script>
    function togglePass() {
      const input = document.getElementById('password');
      const btn = document.querySelector('.eye-btn');
      if (input.type === 'password') {
        input.type = 'text';
        btn.textContent = '🙈';
      } else {
        input.type = 'password';
        btn.textContent = '👁️';
      }
    }
    
    function openKeyModal() {
      document.getElementById('key-modal').classList.add('active');
      document.getElementById('key-token').focus();
      document.getElementById('modal-error').textContent = '';
      document.getElementById('error').textContent = '';
    }
    
    function closeKeyModal() {
      document.getElementById('key-modal').classList.remove('active');
      document.getElementById('key-token').value = '';
    }
    
    async function handleCredentialsLogin(e) {
      e.preventDefault();
      const user = document.getElementById('username').value;
      const pass = document.getElementById('password').value;
      const errEl = document.getElementById('error');
      errEl.textContent = '';
      try {
        const res = await fetch('/api/local-login', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ username: user, password: pass })
        });
        const data = await res.json();
        if (data.ok) {
          window.location.reload();
        } else {
          errEl.textContent = data.error || '帳號或密碼錯誤！';
        }
      } catch (err) {
        errEl.textContent = '連線失敗: ' + err;
      }
    }
    
    async function handleKeyLogin() {
      const keyVal = document.getElementById('key-token').value.trim();
      const errEl = document.getElementById('modal-error');
      errEl.textContent = '';
      if (!keyVal) {
        errEl.textContent = '請輸入金鑰！';
        return;
      }
      try {
        const res = await fetch('/api/local-login', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ key_token: keyVal })
        });
        const data = await res.json();
        if (data.ok) {
          window.location.reload();
        } else {
          errEl.textContent = data.error || '金鑰驗證失敗！';
        }
      } catch (err) {
        errEl.textContent = '連線失敗: ' + err;
      }
    }
  </script>
</body>
</html>
"""


LOCAL_403_HTML = """<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>403 Access Denied - FALO Local Web Gateway</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
  <style>
    :root {
      --bg: #0b0f19;
      --card-bg: rgba(17, 24, 39, 0.75);
      --border: rgba(255, 255, 255, 0.08);
      --text: #f3f4f6;
      --text-muted: #9ca3af;
      --danger: #f97316; /* Amber orange warning color */
      --primary: #10b981;
      --glow: rgba(249, 115, 22, 0.15);
    }
    body {
      margin: 0;
      background: radial-gradient(circle at top, #111827 0%, #030712 100%);
      color: var(--text);
      font-family: 'Outfit', -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      min-height: 100vh;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 20px;
      box-sizing: border-box;
    }
    .card {
      background: var(--card-bg);
      border: 1px solid var(--border);
      border-radius: 16px;
      padding: 40px 30px;
      width: 100%;
      max-width: 480px;
      box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.3), 0 8px 10px -6px rgba(0, 0, 0, 0.3);
      backdrop-filter: blur(12px);
      text-align: center;
    }
    .icon {
      font-size: 4rem;
      margin-bottom: 20px;
      animation: pulse 2s infinite ease-in-out;
    }
    @keyframes pulse {
      0%, 100% { transform: scale(1); opacity: 0.9; }
      50% { transform: scale(1.05); opacity: 1; filter: drop-shadow(0 0 10px rgba(249, 115, 22, 0.5)); }
    }
    h2 {
      margin-top: 0;
      font-weight: 600;
      color: #fff;
      letter-spacing: 0.5px;
    }
    p {
      color: var(--text-muted);
      font-size: 0.95rem;
      line-height: 1.6;
      margin-bottom: 30px;
    }
    .btn-group {
      display: flex;
      flex-direction: column;
      gap: 12px;
    }
    .btn {
      padding: 12px 24px;
      font-size: 1rem;
      font-weight: 600;
      border-radius: 8px;
      cursor: pointer;
      transition: all 0.2s;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
    }
    .btn-primary {
      background: var(--danger);
      color: #fff;
      border: none;
    }
    .btn-primary:hover {
      background: #ea580c;
      transform: translateY(-1px);
    }
    .btn-secondary {
      background: rgba(255, 255, 255, 0.05);
      color: #fff;
      border: 1px solid var(--border);
    }
    .btn-secondary:hover {
      background: rgba(255, 255, 255, 0.1);
      transform: translateY(-1px);
    }
  </style>
</head>
<body>
  <div class="card">
    <div class="icon">🔒</div>
    <h2>403 Access Denied</h2>
    <p>抱歉，您目前登入的身分沒有存取管理後台的權限。<br>此網頁僅開放給系統管理員 (Admin) 帳號存取。</p>
    <div class="btn-group">
      <button class="btn btn-primary" onclick="relogin()">🔑 使用管理員帳號登入</button>
      <a class="btn btn-secondary" href="/index.html">🏠 返回系統首頁</a>
    </div>
  </div>
  <script>
    async function relogin() {
      try {
        await fetch('/api/local-logout', { method: 'POST' });
      } catch (err) {}
      window.location.href = '/admin.html';
    }
  </script>
</body>
</html>
"""


def make_handler(config: AppConfig):
    class RuntimeHandler(BaseHTTPRequestHandler):
        def _network_allowed(self) -> bool:
            client_ip = str(self.client_address[0])
            return is_loopback_client(client_ip) or network_access_enabled(config)

        def _get_current_session(self) -> dict:
            cookie_header = self.headers.get("Cookie", "")
            import re
            match = re.search(r'falo_local_session=([^;]+)', cookie_header)
            if match:
                sid = match.group(1).strip()
                return SESSION_DB.get(sid)
            return None

        def _is_authenticated(self) -> bool:
            return self._get_current_session() is not None

        def _is_admin(self) -> bool:
            sess = self._get_current_session()
            return sess is not None and sess.get("role") == "admin"

        def _send_json_unauthorized(self) -> None:
            body = json.dumps({"ok": False, "error": "Unauthorized: Please login first."}, ensure_ascii=False).encode("utf-8")
            self.send_response(401)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_json_forbidden(self) -> None:
            body = json.dumps({"ok": False, "error": "Forbidden: You do not have permission to perform this action. Only administrators are allowed."}, ensure_ascii=False).encode("utf-8")
            self.send_response(403)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_html_403(self) -> None:
            body = LOCAL_403_HTML.encode("utf-8")
            self.send_response(403)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_network_blocked(self) -> None:
            body = (
                "<!doctype html><html lang=\"zh-Hant\"><head><meta charset=\"utf-8\">"
                "<title>Network Access Disabled</title></head><body>"
                "<h1>Network Access Disabled</h1>"
                "<p>Falo x Force runtime 目前只允許本機 localhost 存取。</p>"
                "<p>請在本機開啟 Portal，到 Logs / Governance 啟用 Same-network Access。</p>"
                "</body></html>"
            ).encode("utf-8")
            self.send_response(403)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_POST(self) -> None:
            global task_queue_manager
            if not self._network_allowed():
                self._send_network_blocked()
                return
            parsed = urlparse(self.path)
            if parsed.path.startswith("/v2/"):
                parsed = parsed._replace(path=parsed.path[3:])
            
            if parsed.path == "/api/local-login":
                try:
                    content_length = int(self.headers.get("Content-Length", 0))
                    post_data = self.rfile.read(content_length).decode("utf-8")
                    data = json.loads(post_data)
                    
                    registry = load_or_create_users(config)
                    
                    username = data.get("username", "").strip()
                    password = data.get("password", "").strip()
                    key_token = data.get("key_token", "").strip()
                    
                    session_id = uuid.uuid4().hex
                    
                    if key_token:
                        # Validate anonymous key
                        matched_key = None
                        for k in registry.get("anonymous_keys", []):
                            if k.get("key_token") == key_token:
                                matched_key = k
                                break
                        
                        if not matched_key:
                            body = json.dumps({"ok": False, "error": "金鑰無效！"}, ensure_ascii=False).encode("utf-8")
                            self.send_response(200)
                            self.send_header("Content-Type", "application/json; charset=utf-8")
                            self.send_header("Content-Length", str(len(body)))
                            self.end_headers()
                            self.wfile.write(body)
                            return
                            
                        if matched_key.get("status") != "active":
                            body = json.dumps({"ok": False, "error": "此金鑰已被停用！"}, ensure_ascii=False).encode("utf-8")
                            self.send_response(200)
                            self.send_header("Content-Type", "application/json; charset=utf-8")
                            self.send_header("Content-Length", str(len(body)))
                            self.end_headers()
                            self.wfile.write(body)
                            return
                        
                        # Update last_used_at
                        matched_key["last_used_at"] = now_iso()
                        
                        # Decouple key_token from key_id to allow rotation without losing session data/settings
                        key_id = matched_key.get("key_id")
                        if not key_id:
                            key_id = f"anon_key_{uuid.uuid4().hex[:8]}"
                            matched_key["key_id"] = key_id
                            
                        save_registry_directly(config, registry)
                        
                        # Set session with the persistent key_id
                        SESSION_DB[session_id] = {
                            "user_id": key_id,
                            "display_name": f"訪客 ({matched_key.get('alias', '未命名')})",
                            "role": "user",
                            "login_type": "key",
                            "key_token": key_token
                        }
                        
                        # Log it
                        write_runtime_log(config, "user_login_key", {
                            "key_token": key_token,
                            "alias": matched_key.get("alias"),
                            "session_id": session_id
                        })
                        
                        body = json.dumps({"ok": True}).encode("utf-8")
                        self.send_response(200)
                        self.send_header("Set-Cookie", f"falo_local_session={session_id}; Path=/; Max-Age=86400; HttpOnly")
                        self.send_header("Content-Type", "application/json; charset=utf-8")
                        self.send_header("Content-Length", str(len(body)))
                        self.end_headers()
                        self.wfile.write(body)
                        return
                    else:
                        # Credentials login
                        if not username:
                            body = json.dumps({"ok": False, "error": "請輸入使用者帳號！"}, ensure_ascii=False).encode("utf-8")
                            self.send_response(200)
                            self.send_header("Content-Type", "application/json; charset=utf-8")
                            self.send_header("Content-Length", str(len(body)))
                            self.end_headers()
                            self.wfile.write(body)
                            return
                            
                        matched_user = None
                        for u in registry.get("local_users", []):
                            if u.get("user_id") == username:
                                matched_user = u
                                break
                        
                        if not matched_user or matched_user.get("password") != password:
                            body = json.dumps({"ok": False, "error": "帳號或密碼錯誤！"}, ensure_ascii=False).encode("utf-8")
                            self.send_response(200)
                            self.send_header("Content-Type", "application/json; charset=utf-8")
                            self.send_header("Content-Length", str(len(body)))
                            self.end_headers()
                            self.wfile.write(body)
                            return
                        
                        # Set session
                        SESSION_DB[session_id] = {
                            "user_id": matched_user.get("user_id"),
                            "display_name": matched_user.get("display_name", username),
                            "role": matched_user.get("role", "user"),
                            "login_type": "credentials",
                            "key_token": None
                        }
                        
                        # Log it
                        write_runtime_log(config, "user_login_credentials", {
                            "user_id": username,
                            "display_name": matched_user.get("display_name"),
                            "role": matched_user.get("role"),
                            "session_id": session_id
                        })
                        
                        body = json.dumps({"ok": True}).encode("utf-8")
                        self.send_response(200)
                        self.send_header("Set-Cookie", f"falo_local_session={session_id}; Path=/; Max-Age=86400; HttpOnly")
                        self.send_header("Content-Type", "application/json; charset=utf-8")
                        self.send_header("Content-Length", str(len(body)))
                        self.end_headers()
                        self.wfile.write(body)
                        return
                except Exception as e:
                    body = json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False).encode("utf-8")
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                return

            if parsed.path == "/api/local-logout":
                cookie_header = self.headers.get("Cookie", "")
                import re
                match = re.search(r'falo_local_session=([^;]+)', cookie_header)
                if match:
                    sid = match.group(1).strip()
                    if sid in SESSION_DB:
                        del SESSION_DB[sid]
                body = json.dumps({"ok": True}).encode("utf-8")
                self.send_response(200)
                self.send_header("Set-Cookie", "falo_local_session=; Path=/; Max-Age=0; HttpOnly")
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return

            if not self._is_authenticated():
                self._send_json_unauthorized()
                return

            if parsed.path.startswith("/api/admin/"):
                if not self._is_admin():
                    self._send_json_forbidden()
                    return

            if parsed.path == "/api/admin/users/save":
                try:
                    content_length = int(self.headers.get("Content-Length", 0))
                    post_data = self.rfile.read(content_length).decode("utf-8")
                    data = json.loads(post_data)
                    
                    local_users = data.get("local_users")
                    anonymous_keys = data.get("anonymous_keys")
                    
                    registry = load_or_create_users(config)
                    if local_users is not None:
                        registry["local_users"] = local_users
                    if anonymous_keys is not None:
                        registry["anonymous_keys"] = anonymous_keys
                    
                    # Ensure there is always at least one admin!
                    has_admin = False
                    for u in registry.get("local_users", []):
                        if u.get("role") == "admin":
                            has_admin = True
                            break
                    if not has_admin:
                        body = json.dumps({"ok": False, "error": "必須保留至少一個 Admin 帳號！"}, ensure_ascii=False).encode("utf-8")
                        self.send_response(200)
                        self.send_header("Content-Type", "application/json; charset=utf-8")
                        self.send_header("Content-Length", str(len(body)))
                        self.end_headers()
                        self.wfile.write(body)
                        return
                    
                    # Ensure all users have passwords
                    for u in registry.get("local_users", []):
                        if not u.get("password"):
                            u["password"] = u.get("user_id", "123456")
                            
                    save_registry_directly(config, registry)
                    write_runtime_log(config, "admin_users_save", {
                        "user_count": len(registry.get("local_users", [])),
                        "key_count": len(registry.get("anonymous_keys", []))
                    })
                    body = json.dumps({"ok": True}).encode("utf-8")
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                except Exception as e:
                    body = json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False).encode("utf-8")
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                return

            elif parsed.path == "/api/admin/users/import":
                content_type = self.headers.get("Content-Type", "")
                if "multipart/form-data" not in content_type:
                    self._send_json({"ok": False, "error": "Must use multipart/form-data."})
                    return
                
                try:
                    form = cgi.FieldStorage(
                        fp=self.rfile,
                        headers=self.headers,
                        environ={
                            "REQUEST_METHOD": "POST",
                            "CONTENT_TYPE": content_type,
                            "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                        },
                    )
                    file_field = form["file"] if "file" in form else None
                    if file_field is None or not getattr(file_field, "filename", ""):
                        self._send_json({"ok": False, "error": "No file uploaded."})
                        return
                    
                    temp_dir = Path(config.project_root) / "data" / "temp"
                    temp_dir.mkdir(parents=True, exist_ok=True)
                    temp_path = temp_dir / f"import_users_{uuid.uuid4().hex[:8]}.xlsx"
                    with temp_path.open("wb") as handle:
                        while True:
                            chunk = file_field.file.read(8192)
                            if not chunk:
                                break
                            handle.write(chunk)
                    
                    import openpyxl
                    wb = openpyxl.load_workbook(temp_path)
                    registry = load_or_create_users(config)
                    
                    # Process 本地使用者 sheet
                    imported_users_new = 0
                    imported_users_updated = 0
                    if "本地使用者" in wb.sheetnames:
                        ws = wb["本地使用者"]
                        rows = list(ws.iter_rows(values_only=True))
                        if len(rows) > 1:
                            local_users = registry.get("local_users", [])
                            for row in rows[1:]:
                                if not row or len(row) < 1:
                                    continue
                                user_id = str(row[0] or "").strip()
                                if not user_id:
                                    continue
                                display_name = str(row[1] or "").strip() or user_id
                                password = str(row[2] or "").strip() or "123456"
                                role = str(row[3] or "").strip()
                                if role not in {"user", "document_manager", "admin"}:
                                    role = "user"
                                
                                found = False
                                for u in local_users:
                                    if u.get("user_id") == user_id:
                                        u["display_name"] = display_name
                                        u["password"] = password
                                        u["role"] = role
                                        imported_users_updated += 1
                                        found = True
                                        break
                                if not found:
                                    local_users.append({
                                        "user_id": user_id,
                                        "display_name": display_name,
                                        "password": password,
                                        "role": role
                                    })
                                    imported_users_new += 1
                            registry["local_users"] = local_users

                    # Process 匿名金鑰 sheet
                    imported_keys_new = 0
                    imported_keys_updated = 0
                    skipped_keys = 0
                    if "匿名金鑰" in wb.sheetnames:
                        ws = wb["匿名金鑰"]
                        rows = list(ws.iter_rows(values_only=True))
                        if len(rows) > 1:
                            anonymous_keys = registry.get("anonymous_keys", [])
                            for row in rows[1:]:
                                if not row or len(row) < 1:
                                    continue
                                token = str(row[0] or "").strip()
                                if not token:
                                    continue
                                import re
                                if not re.match(r"^[1-9]\d{5}$", token):
                                    skipped_keys += 1
                                    continue
                                
                                alias = str(row[1] or "").strip() or "未命名用途"
                                status = str(row[2] or "").strip()
                                if status not in {"active", "disabled"}:
                                    status = "active"
                                    
                                key_id = str(row[3] or "").strip() if len(row) > 3 else ""
                                created_at = str(row[4] or "").strip() if len(row) > 4 else ""
                                last_used_at = str(row[5] or "").strip() if len(row) > 5 else ""
                                
                                found = False
                                for k in anonymous_keys:
                                    if (key_id and k.get("key_id") == key_id) or (not key_id and k.get("key_token") == token):
                                        k["key_token"] = token
                                        k["alias"] = alias
                                        k["status"] = status
                                        if last_used_at:
                                            k["last_used_at"] = last_used_at
                                        imported_keys_updated += 1
                                        found = True
                                        break
                                if not found:
                                    if not key_id:
                                        key_id = f"anon_key_{uuid.uuid4().hex[:8]}"
                                    anonymous_keys.append({
                                        "key_id": key_id,
                                        "key_token": token,
                                        "alias": alias,
                                        "status": status,
                                        "created_at": created_at or now_iso(),
                                        "last_used_at": last_used_at or None
                                    })
                                    imported_keys_new += 1
                            registry["anonymous_keys"] = anonymous_keys
                            
                    # Ensure there is always at least one admin!
                    has_admin = False
                    for u in registry.get("local_users", []):
                        if u.get("role") == "admin":
                            has_admin = True
                            break
                    if not has_admin:
                        self._send_json({"ok": False, "error": "匯入的帳號中必須包含至少一個 Admin 帳號！"})
                        return
                    
                    save_registry_directly(config, registry)
                    
                    try:
                        temp_path.unlink()
                    except Exception:
                        pass
                        
                    msg = f"匯入完成！使用者：新增 {imported_users_new} 個，更新 {imported_users_updated} 個。金鑰：新增 {imported_keys_new} 個，更新 {imported_keys_updated} 個。"
                    if skipped_keys > 0:
                        msg += f"（已跳過 {skipped_keys} 組格式不合規的金鑰，金鑰必須為6位純數字且不能以0開頭）"
                    self._send_json({"ok": True, "message": msg})
                except Exception as e:
                    self._send_json({"ok": False, "error": str(e)})
                return

            elif parsed.path == "/api/admin/users/clear":
                sess = self._get_current_session()
                if not sess:
                    self._send_json({"ok": False, "error": "No active session"})
                    return
                current_user = sess.get("user_id")
                
                try:
                    registry = load_or_create_users(config)
                    filtered_users = [u for u in registry.get("local_users", []) if u.get("user_id") == current_user]
                    if not filtered_users:
                        filtered_users = [{"user_id": current_user, "display_name": sess.get("display_name", "Admin"), "role": "admin", "password": "admin123456"}]
                    registry["local_users"] = filtered_users
                    registry["anonymous_keys"] = []
                    
                    save_registry_directly(config, registry)
                    self._send_json({"ok": True})
                except Exception as e:
                    self._send_json({"ok": False, "error": str(e)})
                return

            elif parsed.path == "/api/admin/audit-logs/import":
                content_type = self.headers.get("Content-Type", "")
                if "multipart/form-data" not in content_type:
                    self._send_json({"ok": False, "error": "Must use multipart/form-data."})
                    return
                
                try:
                    form = cgi.FieldStorage(
                        fp=self.rfile,
                        headers=self.headers,
                        environ={
                            "REQUEST_METHOD": "POST",
                            "CONTENT_TYPE": content_type,
                            "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                        },
                    )
                    file_field = form["file"] if "file" in form else None
                    if file_field is None or not getattr(file_field, "filename", ""):
                        self._send_json({"ok": False, "error": "No file uploaded."})
                        return
                    
                    temp_dir = Path(config.project_root) / "data" / "temp"
                    temp_dir.mkdir(parents=True, exist_ok=True)
                    temp_path = temp_dir / f"import_logs_{uuid.uuid4().hex[:8]}.xlsx"
                    with temp_path.open("wb") as handle:
                        while True:
                            chunk = file_field.file.read(8192)
                            if not chunk:
                                break
                            handle.write(chunk)
                    
                    import openpyxl
                    wb = openpyxl.load_workbook(temp_path)
                    
                    sheet_name = "對話審計紀錄" if "對話審計紀錄" in wb.sheetnames else wb.sheetnames[0]
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    
                    imported_nlm_new = 0
                    imported_nlm_updated = 0
                    imported_gemini_new = 0
                    imported_gemini_updated = 0
                    
                    if len(rows) > 1:
                        # Parse headers
                        headers = [str(cell or "").strip() for cell in rows[0]]
                        idx_platform = 0
                        idx_cid = 1
                        idx_user_name = 2
                        idx_user_id = 3 if "帳號" in "".join(headers) else -1
                        
                        if idx_user_id == 3:
                            idx_remark = 4
                            idx_notebook_id = 5
                            idx_last_question = 6
                            idx_last_query_at = 7
                        else:
                            idx_remark = 3
                            idx_notebook_id = 4
                            idx_last_question = 5
                            idx_last_query_at = 6

                        nlm_file = Path(config.project_root) / "data" / "multichat_sessions.json"
                        gemini_file = Path(config.project_root) / "data" / "gemini_sessions.json"
                        
                        nlm_data = {"sessions": {}}
                        if nlm_file.exists():
                            try:
                                with open(nlm_file, "r", encoding="utf-8") as f:
                                    nlm_data = json.load(f)
                            except Exception:
                                pass
                        if "sessions" not in nlm_data:
                            nlm_data["sessions"] = {}
                            
                        gemini_data = {"sessions": {}}
                        if gemini_file.exists():
                            try:
                                with open(gemini_file, "r", encoding="utf-8") as f:
                                    gemini_data = json.load(f)
                            except Exception:
                                pass
                        if "sessions" not in gemini_data:
                            gemini_data["sessions"] = {}
                            
                        for row in rows[1:]:
                            if not row or len(row) < 2:
                                continue
                            platform = str(row[idx_platform] or "notebooklm").strip().lower()
                            cid = str(row[idx_cid] or "").strip()
                            if not cid:
                                continue
                            
                            user_name = str(row[idx_user_name] or "Unknown").strip()
                            user_id = str(row[idx_user_id] or "").strip() if (idx_user_id != -1 and len(row) > idx_user_id) else ""
                            remark = str(row[idx_remark] or "").strip() if len(row) > idx_remark else ""
                            notebook_id = str(row[idx_notebook_id] or "").strip() if len(row) > idx_notebook_id else ""
                            last_question = str(row[idx_last_question] or "").strip() if len(row) > idx_last_question else ""
                            last_query_at = str(row[idx_last_query_at] or "").strip() if (len(row) > idx_last_query_at and row[idx_last_query_at]) else now_iso()
                            
                            if platform == "notebooklm":
                                sessions = nlm_data["sessions"]
                                if cid in sessions:
                                    sessions[cid]["user_name"] = user_name
                                    sessions[cid]["user_id"] = user_id
                                    sessions[cid]["remark"] = remark
                                    sessions[cid]["notebook_id"] = notebook_id
                                    sessions[cid]["last_query_at"] = last_query_at
                                    imported_nlm_updated += 1
                                else:
                                    turns = [{"role": "user", "content": last_question}] if last_question else []
                                    sessions[cid] = {
                                        "user_id": user_id,
                                        "user_name": user_name,
                                        "notebook_id": notebook_id,
                                        "created_at": last_query_at,
                                        "last_query_at": last_query_at,
                                        "remark": remark,
                                        "turns": turns
                                    }
                                    imported_nlm_new += 1
                            else:
                                sessions = gemini_data["sessions"]
                                if cid in sessions:
                                    sessions[cid]["user_name"] = user_name
                                    sessions[cid]["user_id"] = user_id
                                    sessions[cid]["remark"] = remark
                                    sessions[cid]["last_query_at"] = last_query_at
                                    imported_gemini_updated += 1
                                else:
                                    turns = [{"role": "user", "content": last_question}] if last_question else []
                                    sessions[cid] = {
                                        "user_id": user_id,
                                        "user_name": user_name,
                                        "created_at": last_query_at,
                                        "last_query_at": last_query_at,
                                        "remark": remark,
                                        "turns": turns,
                                        "metadata": []
                                    }
                                    imported_gemini_new += 1
                                    
                        nlm_file.parent.mkdir(parents=True, exist_ok=True)
                        with open(nlm_file, "w", encoding="utf-8") as f:
                            json.dump(nlm_data, f, ensure_ascii=False, indent=2)
                        with open(gemini_file, "w", encoding="utf-8") as f:
                            json.dump(gemini_data, f, ensure_ascii=False, indent=2)
                            
                    try:
                        temp_path.unlink()
                    except Exception:
                        pass
                        
                    msg = f"審計紀錄匯入完成！NotebookLM：新增 {imported_nlm_new} 筆，更新 {imported_nlm_updated} 筆。Gemini：新增 {imported_gemini_new} 筆，更新 {imported_gemini_updated} 筆。"
                    self._send_json({"ok": True, "message": msg})
                except Exception as e:
                    self._send_json({"ok": False, "error": str(e)})
                return

            elif parsed.path == "/api/admin/audit-logs/clear":
                try:
                    nlm_file = Path(config.project_root) / "data" / "multichat_sessions.json"
                    with open(nlm_file, "w", encoding="utf-8") as f:
                        json.dump({"sessions": {}}, f, ensure_ascii=False, indent=2)
                    
                    gemini_file = Path(config.project_root) / "data" / "gemini_sessions.json"
                    with open(gemini_file, "w", encoding="utf-8") as f:
                        json.dump({"sessions": {}}, f, ensure_ascii=False, indent=2)
                        
                    self._send_json({"ok": True})
                except Exception as e:
                    self._send_json({"ok": False, "error": str(e)})
                return

            if parsed.path == "/api/admin/selectable-notebooks/save":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                try:
                    data = json.loads(post_data)
                except Exception:
                    from urllib.parse import parse_qs
                    params = parse_qs(post_data)
                    notebook_ids = params.get("notebook_ids", [])
                    if len(notebook_ids) == 1 and "," in notebook_ids[0]:
                        notebook_ids = [x.strip() for x in notebook_ids[0].split(",") if x.strip()]
                    data = {
                        "mode": params.get("mode", ["all"])[0].strip(),
                        "notebook_ids": notebook_ids
                    }
                
                mode = data.get("mode", "all").strip()
                notebook_ids = data.get("notebook_ids", [])
                if not isinstance(notebook_ids, list):
                    if isinstance(notebook_ids, str):
                        notebook_ids = [x.strip() for x in notebook_ids.split(",") if x.strip()]
                    else:
                        notebook_ids = []
                
                allowed_path = config_dir(config) / "selectable_books.json"
                allowed_path.parent.mkdir(parents=True, exist_ok=True)
                
                try:
                    allowed_path.write_text(json.dumps({
                        "app": "AI NotebookLM Runtime Lab",
                        "kind": "selectable_books",
                        "mode": mode,
                        "updated_at": now_iso(),
                        "notebook_ids": notebook_ids
                    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
                    self._send_json({"ok": True})
                except Exception as e:
                    self._send_json({"ok": False, "error": str(e)})
                return

            if parsed.path == "/api/session/remark":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                try:
                    data = json.loads(post_data)
                except Exception:
                    from urllib.parse import parse_qs
                    params = parse_qs(post_data)
                    data = {
                        "conversation_id": params.get("conversation_id", [""])[0].strip(),
                        "type": params.get("type", [""])[0].strip(),
                        "remark": params.get("remark", [""])[0].strip()
                    }
                
                conversation_id = data.get("conversation_id", "").strip()
                sess_type = data.get("type", "").strip() # 'nlm' or 'gemini'
                remark = data.get("remark", "").strip()
                
                if not conversation_id or not sess_type:
                    self._send_json({"ok": False, "error": "Conversation ID and type are required."})
                    return
                
                filename = "multichat_sessions.json" if sess_type == "nlm" else "gemini_sessions.json"
                sessions_file = Path(config.project_root) / "data" / filename
                
                sessions_data = {"sessions": {}}
                if sessions_file.exists():
                    try:
                        with open(sessions_file, "r", encoding="utf-8") as f:
                            sessions_data = json.load(f)
                    except Exception:
                        pass
                
                if "sessions" not in sessions_data:
                    sessions_data["sessions"] = {}
                
                if conversation_id not in sessions_data["sessions"]:
                    sessions_data["sessions"][conversation_id] = {
                        "conversation_id": conversation_id,
                        "turns": [],
                        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "last_query_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "user_name": "System"
                    }
                
                sessions_data["sessions"][conversation_id]["remark"] = remark
                
                try:
                    sessions_file.parent.mkdir(parents=True, exist_ok=True)
                    with open(sessions_file, "w", encoding="utf-8") as f:
                        json.dump(sessions_data, f, ensure_ascii=False, indent=2)
                    self._send_json({"ok": True})
                except Exception as e:
                    self._send_json({"ok": False, "error": str(e)})
                return

            if parsed.path == "/api/simple-upload":
                self._send_html(render_simple_upload_result(handle_simple_upload_post(config, self)))
            elif parsed.path == "/api/convert-meeting":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                from urllib.parse import parse_qs
                params = parse_qs(post_data)
                markdown_text = params.get("markdown", [""])[0].strip()

                if not markdown_text:
                    self._send_json({"ok": False, "error": "Markdown content is empty."})
                    return

                try:
                    import md_to_meeting_docx
                    docx_path, single_path, rwd_path, summary = md_to_meeting_docx.convert(markdown_text)
                    self._send_json({
                        "ok": True,
                        "docx": os.path.basename(docx_path) if docx_path else "",
                        "single_html": os.path.basename(single_path) if single_path else "",
                        "rwd_html": os.path.basename(rwd_path) if rwd_path else "",
                        "summary": summary
                    })
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    self._send_json({"ok": False, "error": str(e)})
            elif parsed.path == "/api/multichat/ask":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                from urllib.parse import parse_qs
                params = parse_qs(post_data)
                notebook_id = params.get("notebook_id", [""])[0].strip()
                notebook_title = params.get("notebook_title", [""])[0].strip()
                user_name = params.get("user_name", [""])[0].strip()
                conversation_id = params.get("conversation_id", [""])[0].strip()
                question = params.get("question", [""])[0].strip()

                sess = self._get_current_session()
                user_id = sess.get("user_id") if sess else "unknown_user"
                active_user_name = sess.get("display_name") if (sess and sess.get("display_name")) else user_name

                if not notebook_title:
                    try:
                        projects = load_projects(config)
                        for p in projects:
                            if p.get("notebook_id") == notebook_id:
                                notebook_title = p.get("notebook_title") or p.get("name") or ""
                                break
                    except Exception:
                        pass
                if not notebook_title:
                    notebook_title = f"ID: {notebook_id[:8]}"

                if not notebook_id or not active_user_name or not question:
                    self._send_json({"ok": False, "error": "Notebook ID, User Name, and Question are required."})
                    return

                if conversation_id and conversation_id != "new":
                    sessions_file = Path(config.project_root) / "data" / "multichat_sessions.json"
                    if sessions_file.exists():
                        try:
                            with open(sessions_file, "r", encoding="utf-8") as f:
                                data = json.load(f)
                            if "sessions" in data and conversation_id in data["sessions"]:
                                session_info = data["sessions"][conversation_id]
                                owner_id = session_info.get("user_id")
                                owner_name = session_info.get("user_name")
                                is_owner = (owner_id == user_id) or (not owner_id and owner_name == active_user_name)
                                if not is_owner:
                                    self._send_json({"ok": False, "error": "Access denied: cannot participate in a session belonging to another user."})
                                    return
                        except Exception:
                            pass

                if task_queue_manager is None:
                    self._send_json({"ok": False, "error": "Task queue manager is not initialized."})
                    return

                payload = {
                    "notebook_id": notebook_id,
                    "notebook_title": notebook_title,
                    "user_name": active_user_name,
                    "user_id": user_id,
                    "conversation_id": conversation_id,
                    "question": question
                }
                write_runtime_log(config, "multichat_ask", {
                    "user_name": active_user_name,
                    "user_id": user_id,
                    "notebook_id": notebook_id,
                    "notebook_title": notebook_title,
                    "question_summary": question[:50]
                })
                task_id = task_queue_manager.add_task("notebooklm", active_user_name, payload)
                detail = task_queue_manager.get_task_status_detail(task_id)
                self._send_json({
                    "ok": True,
                    "task_id": task_id,
                    "status": "pending",
                    "queue_position": detail.get("queue_position", 0),
                    "eta_seconds": detail.get("eta_seconds", 0)
                })
            elif parsed.path == "/api/gas/settings/save":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                from urllib.parse import parse_qs
                params = parse_qs(post_data)
                
                settings = read_gas_settings(config)
                web_app_url = params.get("web_app_url", [""])[0].strip()
                api_token = params.get("api_token", [""])[0].strip()
                poll_interval = params.get("poll_interval_seconds", ["300"])[0].strip()
                enabled = params.get("enabled", ["false"])[0].strip().lower() in {"true", "1", "on"}
                auto_poll = params.get("auto_poll_enabled", ["false"])[0].strip().lower() in {"true", "1", "on"}
                
                settings["web_app_url"] = web_app_url
                settings["api_token"] = api_token
                settings["poll_interval_seconds"] = int(poll_interval) if poll_interval.isdigit() else 300
                settings["enabled"] = enabled
                settings["auto_poll_enabled"] = auto_poll
                
                write_gas_settings(config, settings)
                self._send_json({"ok": True, "settings": settings})
            elif parsed.path == "/api/gas/push-host-info":
                from gas_adapter import push_host_info_to_gas, detect_wan_url
                res = push_host_info_to_gas(config, "manual")
                wan_url = detect_wan_url(RUNTIME_BIND_PORT)
                self._send_json({"ok": res.get("ok", False), "wan_url": wan_url, "response": res})
            elif parsed.path == "/api/gas/history/save":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                from urllib.parse import parse_qs
                params = parse_qs(post_data)
                url = params.get("url", [""])[0].strip()
                alias = params.get("alias", [""])[0].strip()
                action = params.get("action", ["save"])[0].strip()
                
                if not url:
                    self._send_json({"ok": False, "error": "URL is required."})
                    return
                
                history = get_gas_url_history(config)
                if action == "delete":
                    history = [item for item in history if item.get("url") != url]
                else:
                    found = False
                    for item in history:
                        if item.get("url") == url:
                            item["alias"] = alias
                            item["last_used_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
                            found = True
                            break
                    if not found:
                        history.append({
                            "url": url,
                            "alias": alias,
                            "added_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                            "last_used_at": time.strftime("%Y-%m-%d %H:%M:%S")
                        })
                
                save_gas_url_history(config, history)
                self._send_json({"ok": True})
            elif parsed.path == "/api/gas/history/clear":
                save_gas_url_history(config, [])
                self._send_json({"ok": True})
            elif parsed.path == "/api/gas/history/import":
                content_type = self.headers.get("Content-Type", "")
                if "multipart/form-data" not in content_type:
                    self._send_json({"ok": False, "error": "Must use multipart/form-data."})
                    return
                
                try:
                    form = cgi.FieldStorage(
                        fp=self.rfile,
                        headers=self.headers,
                        environ={
                            "REQUEST_METHOD": "POST",
                            "CONTENT_TYPE": content_type,
                            "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                        },
                    )
                    file_field = form["file"] if "file" in form else None
                    if file_field is None or not getattr(file_field, "filename", ""):
                        self._send_json({"ok": False, "error": "No file uploaded."})
                        return
                    
                    # Write file temporarily
                    temp_dir = Path(config.project_root) / "data" / "temp"
                    temp_dir.mkdir(parents=True, exist_ok=True)
                    temp_path = temp_dir / "import_temp.xlsx"
                    with temp_path.open("wb") as handle:
                        while True:
                            chunk = file_field.file.read(8192)
                            if not chunk:
                                break
                            handle.write(chunk)
                    
                    # Parse using openpyxl
                    import openpyxl
                    wb = openpyxl.load_workbook(temp_path)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        self._send_json({"ok": False, "error": "Excel sheet is empty or invalid."})
                        return
                    
                    # Assume row 0 is header: URL, Alias, Added At, Last Used At
                    history = get_gas_url_history(config)
                    imported_count = 0
                    updated_count = 0
                    
                    for row in rows[1:]:
                        if not row or len(row) < 1:
                            continue
                        url = str(row[0] or "").strip()
                        if not url or not url.startswith("http"):
                            continue
                            
                        alias = str(row[1] or "").strip() if len(row) > 1 else ""
                        added_at = str(row[2] or "").strip() if len(row) > 2 else time.strftime("%Y-%m-%d %H:%M:%S")
                        last_used_at = str(row[3] or "").strip() if len(row) > 3 else time.strftime("%Y-%m-%d %H:%M:%S")
                        
                        # Find existing
                        found = False
                        for item in history:
                            if item.get("url") == url:
                                # Update only if alias is different (upsert strategy)
                                if item.get("alias") != alias:
                                    item["alias"] = alias
                                    item["last_used_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
                                    updated_count += 1
                                found = True
                                break
                        if not found:
                            history.append({
                                "url": url,
                                "alias": alias,
                                "added_at": added_at or time.strftime("%Y-%m-%d %H:%M:%S"),
                                "last_used_at": last_used_at or time.strftime("%Y-%m-%d %H:%M:%S")
                            })
                            imported_count += 1
                            
                    save_gas_url_history(config, history)
                    
                    # Clean temp
                    try:
                        temp_path.unlink()
                    except Exception:
                        pass
                        
                    self._send_json({
                        "ok": True, 
                        "imported": imported_count, 
                        "updated": updated_count,
                        "message": f"匯入成功：新增 {imported_count} 筆，更新 {updated_count} 筆不同網址的備註。"
                    })
                except Exception as ex:
                    import traceback
                    traceback.print_exc()
                    self._send_json({"ok": False, "error": f"匯入失敗: {ex}"})
                return
            elif parsed.path == "/api/queue/clear":
                if task_queue_manager is None:
                    self._send_json({"ok": False, "error": "Task queue manager is not initialized."})
                    return
                task_queue_manager.clear_queue()
                self._send_json({"ok": True})
            elif parsed.path == "/api/multichat/delete":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                from urllib.parse import parse_qs
                params = parse_qs(post_data)
                conversation_id = params.get("conversation_id", [""])[0].strip()

                if not conversation_id:
                    self._send_json({"ok": False, "error": "Conversation ID is required."})
                    return

                sess = self._get_current_session()
                current_user_id = sess.get("user_id") if sess else None
                current_display_name = sess.get("display_name") if sess else None
                current_role = sess.get("role") if sess else "user"

                sessions_file = Path(config.project_root) / "data" / "multichat_sessions.json"
                if sessions_file.exists():
                    try:
                        with open(sessions_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        if "sessions" in data and conversation_id in data["sessions"]:
                            session_info = data["sessions"][conversation_id]
                            if current_role != "admin":
                                owner_id = session_info.get("user_id")
                                owner_name = session_info.get("user_name")
                                is_owner = (owner_id == current_user_id) or (not owner_id and owner_name == current_display_name)
                                if not is_owner:
                                    self._send_json({"ok": False, "error": "Access denied: cannot delete session belonging to another user."})
                                    return
                            del data["sessions"][conversation_id]
                            with open(sessions_file, "w", encoding="utf-8") as f:
                                json.dump(data, f, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                
                self._send_json({"ok": True})
            elif parsed.path == "/api/gemini/ask":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                from urllib.parse import parse_qs
                params = parse_qs(post_data)
                user_name = params.get("user_name", [""])[0].strip()
                question = params.get("question", [""])[0].strip()
                metadata = params.get("metadata", [""])[0].strip()
                model = params.get("model", [""])[0].strip()
                thinking = params.get("thinking", [""])[0].strip()

                sess = self._get_current_session()
                user_id = sess.get("user_id") if sess else "unknown_user"
                active_user_name = sess.get("display_name") if (sess and sess.get("display_name")) else user_name

                if not active_user_name or not question:
                    self._send_json({"ok": False, "error": "User Name and Question are required."})
                    return

                conversation_id = ""
                if metadata and metadata not in {"new", "[]", "null"}:
                    try:
                        meta_arr = json.loads(metadata)
                        if isinstance(meta_arr, list) and len(meta_arr) > 0:
                            conversation_id = meta_arr[0]
                    except Exception:
                        pass

                if conversation_id:
                    sessions_file = Path(config.project_root) / "data" / "gemini_sessions.json"
                    if sessions_file.exists():
                        try:
                            with open(sessions_file, "r", encoding="utf-8") as f:
                                data = json.load(f)
                            if "sessions" in data and conversation_id in data["sessions"]:
                                session_info = data["sessions"][conversation_id]
                                owner_id = session_info.get("user_id")
                                owner_name = session_info.get("user_name")
                                is_owner = (owner_id == user_id) or (not owner_id and owner_name == active_user_name)
                                if not is_owner:
                                    self._send_json({"ok": False, "error": "Access denied: cannot participate in a session belonging to another user."})
                                    return
                        except Exception:
                            pass

                if task_queue_manager is None:
                    self._send_json({"ok": False, "error": "Task queue manager is not initialized."})
                    return

                payload = {
                    "user_name": active_user_name,
                    "user_id": user_id,
                    "question": question,
                    "metadata": metadata,
                    "model": model,
                    "thinking": thinking
                }
                task_id = task_queue_manager.add_task("gemini", active_user_name, payload)
                detail = task_queue_manager.get_task_status_detail(task_id)
                self._send_json({
                    "ok": True,
                    "task_id": task_id,
                    "status": "pending",
                    "queue_position": detail.get("queue_position", 0),
                    "eta_seconds": detail.get("eta_seconds", 0)
                })
            elif parsed.path == "/api/gemini/delete":
                content_length = int(self.headers.get("Content-Length", 0))
                post_data = self.rfile.read(content_length).decode("utf-8")
                from urllib.parse import parse_qs
                params = parse_qs(post_data)
                conversation_id = params.get("conversation_id", [""])[0].strip()

                if not conversation_id:
                    self._send_json({"ok": False, "error": "Conversation ID is required."})
                    return

                sess = self._get_current_session()
                current_user_id = sess.get("user_id") if sess else None
                current_display_name = sess.get("display_name") if sess else None
                current_role = sess.get("role") if sess else "user"

                sessions_file = Path(config.project_root) / "data" / "gemini_sessions.json"
                if sessions_file.exists():
                    try:
                        with open(sessions_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        if "sessions" in data and conversation_id in data["sessions"]:
                            session_info = data["sessions"][conversation_id]
                            if current_role != "admin":
                                owner_id = session_info.get("user_id")
                                owner_name = session_info.get("user_name")
                                is_owner = (owner_id == current_user_id) or (not owner_id and owner_name == current_display_name)
                                if not is_owner:
                                    self._send_json({"ok": False, "error": "Access denied: cannot delete session belonging to another user."})
                                    return
                            del data["sessions"][conversation_id]
                            with open(sessions_file, "w", encoding="utf-8") as f:
                                json.dump(data, f, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                
                self._send_json({"ok": True})
            else:
                self.send_error(404, "Not found")


        def do_GET(self) -> None:
            import json
            global task_queue_manager
            if not self._network_allowed():
                self._send_network_blocked()
                return
            parsed = urlparse(self.path)
            if parsed.path.startswith("/v2/"):
                parsed = parsed._replace(path=parsed.path[3:])
            
            if parsed.path == "/api/status":
                self._send_json(build_status_payload(config))
                return

            if not self._is_authenticated():
                if parsed.path in {"/", "/index.html", "/index-old.html", "/admin.html"}:
                    self._send_html(LOCAL_LOGIN_PAGE_HTML)
                    return
                self._send_json_unauthorized()
                return

            if parsed.path == "/admin.html":
                if not self._is_admin():
                    self._send_html_403()
                    return

            if parsed.path.startswith("/api/admin/"):
                if not self._is_admin():
                    self._send_json_forbidden()
                    return

            if parsed.path == "/api/local-logout":
                cookie_header = self.headers.get("Cookie", "")
                import re
                match = re.search(r'falo_local_session=([^;]+)', cookie_header)
                if match:
                    sid = match.group(1).strip()
                    if sid in SESSION_DB:
                        del SESSION_DB[sid]
                body = json.dumps({"ok": True}).encode("utf-8")
                self.send_response(200)
                self.send_header("Set-Cookie", "falo_local_session=; Path=/; Max-Age=0; HttpOnly")
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return

            if parsed.path == "/api/local-session":
                sess = self._get_current_session()
                if sess:
                    body = json.dumps({
                        "ok": True,
                        "session": {
                            "user_id": sess.get("user_id"),
                            "display_name": sess.get("display_name"),
                            "role": sess.get("role"),
                            "login_type": sess.get("login_type")
                        }
                    }, ensure_ascii=False).encode("utf-8")
                else:
                    body = json.dumps({"ok": False, "error": "No active session"}, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return

            if parsed.path == "/api/admin/users":
                registry = load_or_create_users(config)
                body = json.dumps({
                    "ok": True,
                    "local_users": registry.get("local_users", []),
                    "anonymous_keys": registry.get("anonymous_keys", [])
                }, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return

            elif parsed.path == "/api/admin/users/export":
                registry = load_or_create_users(config)
                import io
                import openpyxl
                wb = openpyxl.Workbook()
                
                # Sheet 1: 本地使用者
                ws1 = wb.active
                ws1.title = "本地使用者"
                ws1.append(["帳號 (Username)", "顯示名稱 (Display Name)", "登入密碼 (Password)", "角色權限 (Role)"])
                for u in registry.get("local_users", []):
                    ws1.append([u.get("user_id", ""), u.get("display_name", ""), u.get("password", ""), u.get("role", "")])
                    
                # Sheet 2: 匿名金鑰
                ws2 = wb.create_sheet("匿名金鑰")
                ws2.append(["金鑰 Token (Access Key)", "用途別名 (Alias)", "啟用狀態 (Status)", "金鑰 ID (Key ID)", "建立時間", "最後使用時間"])
                for k in registry.get("anonymous_keys", []):
                    ws2.append([k.get("key_token", ""), k.get("alias", ""), k.get("status", ""), k.get("key_id", ""), k.get("created_at", ""), k.get("last_used_at", "")])
                
                stream = io.BytesIO()
                wb.save(stream)
                body = stream.getvalue()
                
                filename = f"accounts_keys_export_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f"attachment; filename={filename}")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return

            elif parsed.path == "/api/admin/audit-logs/export":
                import io
                import openpyxl
                import json
                
                # Load notebooklm sessions
                nlm_file = Path(config.project_root) / "data" / "multichat_sessions.json"
                audit_logs = []
                if nlm_file.exists():
                    try:
                        with open(nlm_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        for cid, info in data.get("sessions", {}).items():
                            last_q = ""
                            if info.get("turns"):
                                for turn in reversed(info["turns"]):
                                    if turn.get("role") == "user":
                                        last_q = turn.get("content", "")
                                        break
                            audit_logs.append({
                                "platform": "notebooklm",
                                "conversation_id": cid,
                                "user_name": info.get("user_name", "Unknown"),
                                "user_id": info.get("user_id", ""),
                                "notebook_id": f"{info.get('notebook_title')} ({info.get('notebook_id')})" if info.get("notebook_title") else info.get("notebook_id", ""),
                                "last_query_at": info.get("last_query_at", ""),
                                "last_question": last_q,
                                "remark": info.get("remark", "")
                            })
                    except Exception:
                        pass
                
                # Load gemini sessions
                gemini_file = Path(config.project_root) / "data" / "gemini_sessions.json"
                if gemini_file.exists():
                    try:
                        with open(gemini_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        for cid, info in data.get("sessions", {}).items():
                            last_q = ""
                            if info.get("turns"):
                                for turn in reversed(info["turns"]):
                                    if turn.get("role") == "user":
                                        last_q = turn.get("content", "")
                                        break
                            audit_logs.append({
                                "platform": "gemini",
                                "conversation_id": cid,
                                "user_name": info.get("user_name", "Unknown"),
                                "user_id": info.get("user_id", ""),
                                "notebook_id": "Shared Session (gemini.google.com)",
                                "last_query_at": info.get("last_query_at", ""),
                                "last_question": last_q,
                                "remark": info.get("remark", "")
                            })
                    except Exception:
                        pass
                
                # Sort by last_query_at descending
                audit_logs.sort(key=lambda x: x["last_query_at"], reverse=True)
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "對話審計紀錄"
                ws.append(["平台 (Platform)", "Session ID", "同仁姓名 (Display Name)", "帳號 (User ID)", "會話標記 / 備註 (Remark)", "筆記本名稱/ID (Notebook ID)", "最後提問內容 (Last Question)", "最後活動時間 (Last Query At)"])
                for log in audit_logs:
                    ws.append([
                        log["platform"],
                        log["conversation_id"],
                        log["user_name"],
                        log["user_id"],
                        log["remark"],
                        log["notebook_id"],
                        log["last_question"],
                        log["last_query_at"]
                    ])
                
                stream = io.BytesIO()
                wb.save(stream)
                body = stream.getvalue()
                
                filename = f"audit_logs_export_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f"attachment; filename={filename}")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return

            if parsed.path in {"/", "/index.html"}:
                self._send_file(PROJECT_ROOT / "index.html", "text/html; charset=utf-8")
            elif parsed.path == "/index-old.html":
                self._send_file(PROJECT_ROOT / "index-old.html", "text/html; charset=utf-8")
            elif parsed.path == "/admin.html":
                self._send_file(PROJECT_ROOT / "admin.html", "text/html; charset=utf-8")
            elif parsed.path == "/api/meeting-download":
                params = parse_qs(parsed.query)
                filename = params.get("filename", [""])[0].strip()
                filename = os.path.basename(filename)
                if not filename:
                    self.send_error(400, "Bad Request: missing filename")
                    return
                
                import md_to_meeting_docx
                file_path = Path(md_to_meeting_docx.OUTPUT_DIR) / filename
                if not file_path.exists() or not file_path.is_file():
                    self.send_error(404, "File not found")
                    return
                
                if filename.endswith(".docx"):
                    ctype = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                elif filename.endswith(".html"):
                    ctype = "text/html; charset=utf-8"
                else:
                    ctype = "application/octet-stream"
                self._send_file(file_path, ctype)
            elif parsed.path == "/api/gas/wan-url":
                from gas_adapter import detect_wan_url
                wan_url = detect_wan_url(RUNTIME_BIND_PORT)
                self._send_json({"ok": True, "wan_url": wan_url})
            elif parsed.path == "/api/gas/settings":
                settings = read_gas_settings(config)
                self._send_json({"ok": True, "settings": settings})
            elif parsed.path == "/api/admin/selectable-notebooks":
                projects = load_projects(config)
                notebook_map = {}
                for p in projects:
                    nb_id = p.get("notebook_id")
                    nb_title = p.get("notebook_title") or p.get("name") or "(無名稱)"
                    if nb_id:
                        notebook_map[nb_id] = {
                            "id": nb_id,
                            "title": nb_title,
                            "created_at": p.get("notebook_created_at") or p.get("created_at") or "",
                            "source": p.get("source") or "sync"
                        }
                allowed_path = config_dir(config) / "selectable_books.json"
                allowed_ids = []
                mode = "all"
                if allowed_path.exists():
                    try:
                        allowed_data = json.loads(allowed_path.read_text(encoding="utf-8"))
                        allowed_ids = allowed_data.get("notebook_ids", [])
                        mode = allowed_data.get("mode", "all")
                    except Exception:
                        pass
                notebooks_list = []
                for nb_id, nb in notebook_map.items():
                    nb["selected"] = (nb_id in allowed_ids)
                    notebooks_list.append(nb)
                notebooks_list.sort(key=lambda x: x["title"].lower())
                self._send_json({"ok": True, "mode": mode, "notebooks": notebooks_list})
            elif parsed.path == "/api/gas/history":
                history = get_gas_url_history(config)
                self._send_json({"ok": True, "history": history})
            elif parsed.path == "/api/gas/history/export":
                import openpyxl
                from openpyxl import Workbook
                history = get_gas_url_history(config)
                
                wb = Workbook()
                ws = wb.active
                ws.title = "GAS URL History"
                
                ws.append(["URL", "Alias", "Added At", "Last Used At"])
                for item in history:
                    ws.append([
                        item.get("url", ""),
                        item.get("alias", ""),
                        item.get("added_at", ""),
                        item.get("last_used_at", "")
                    ])
                
                export_dir = Path(config.project_root) / "data" / "exports"
                export_dir.mkdir(parents=True, exist_ok=True)
                export_path = export_dir / "gas_url_history.xlsx"
                wb.save(export_path)
                
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="gas_url_history.xlsx"')
                self.send_header("Content-Length", str(export_path.stat().st_size))
                self.end_headers()
                
                with export_path.open("rb") as f:
                    self.wfile.write(f.read())
                return
            elif parsed.path == "/api/status":
                self._send_json(build_status_payload(config))
            elif parsed.path == "/api/runtime-state":
                self._send_json(build_runtime_state(config))
            elif parsed.path == "/api/normalize":
                self._send_html(render_normalize_result(normalize_inbox_files(config)))
            elif parsed.path == "/api/build-queue":
                self._send_html(render_queue_result(build_queue_manifest(config)))
            elif parsed.path == "/api/adapter-preview":
                params = parse_qs(parsed.query)
                notebook_id = params.get("notebook_id", ["PASTE_NOTEBOOK_ID_HERE"])[0].strip()
                project_id = params.get("project_id", [""])[0].strip()
                self._send_html(render_adapter_preview(preview_notebooklm_adapter(config, notebook_id, project_id)))
            elif parsed.path == "/api/notebooks":
                params = parse_qs(parsed.query)
                if params.get("format", [""])[0] == "json":
                    res_data = list_notebooks(config)
                    if res_data.get("ok") and isinstance(res_data.get("notebooks"), list):
                        allowed_path = config_dir(config) / "selectable_books.json"
                        original_count = len(res_data["notebooks"])
                        res_data["admin_restricted"] = False
                        if allowed_path.exists():
                            try:
                                allowed_data = json.loads(allowed_path.read_text(encoding="utf-8"))
                                mode = allowed_data.get("mode", "all")
                                allowed_ids = allowed_data.get("notebook_ids", [])
                                
                                if mode == "none" and original_count > 0:
                                    res_data["notebooks"] = []
                                    res_data["count"] = 0
                                    res_data["admin_restricted"] = True
                                elif mode == "custom" and isinstance(allowed_ids, list):
                                    filtered = [nb for nb in res_data["notebooks"] if nb.get("id") in allowed_ids]
                                    if len(filtered) < original_count:
                                        res_data["admin_restricted"] = True
                                    res_data["notebooks"] = filtered
                                    res_data["count"] = len(filtered)
                            except Exception as e:
                                write_runtime_log(config, "filter_notebooks_error", {"error": str(e)})
                    self._send_json(res_data)
                else:
                    self._send_html(render_notebook_list(list_notebooks(config)))
            elif parsed.path == "/api/projects":
                params = parse_qs(parsed.query)
                search = params.get("search", [""])[0]
                sort = params.get("sort", ["updated_at"])[0]
                page = int(params.get("page", ["1"])[0] or 1)
                self._send_html(render_project_manager(list_projects(config, search=search, sort=sort, page=page, page_size=500)))
            elif parsed.path == "/api/projects/sync":
                notebooks = list_notebooks(config)
                if notebooks.get("ok"):
                    result = sync_projects_from_notebooks(config, notebooks.get("notebooks", []))
                else:
                    result = {"mode": "project_sync", "project_count": 0, "added_count": 0, "projects": [], "error": notebooks.get("error", "")}
                self._send_html(render_project_sync_result(result))
            elif parsed.path == "/api/projects/active":
                params = parse_qs(parsed.query)
                project_id = params.get("project_id", [""])[0].strip()
                self._send_html(render_active_project_result(set_active_project(config, project_id)))
            elif parsed.path == "/api/simple-upload/create-project":
                params = parse_qs(parsed.query)
                title = params.get("title", [""])[0]
                self._send_html(render_create_or_use_result(create_or_use_notebook(config, title)))
            elif parsed.path == "/api/folder-scan":
                params = parse_qs(parsed.query)
                folder_path = params.get("folder_path", [str(simple_incoming_dir(config))])[0]
                extensions = parse_extensions(params.get("types", []))
                recursive = params.get("recursive", [""])[0].lower() in {"1", "yes", "on", "true"}
                order = params.get("order", ["name"])[0]
                self._send_html(render_folder_scan_preview(scan_source_folder(config, folder_path, extensions, recursive, order), params))
            elif parsed.path == "/api/folder-upload":
                params = parse_qs(parsed.query)
                confirm = params.get("confirm", [""])[0].lower() == "yes"
                if not confirm:
                    self._send_html(render_simple_upload_result({"ok": False, "mode": "simple_upload", "project_id": "", "notebook_id": "", "results": [], "error": "Folder upload requires confirm=yes from Scan Preview."}))
                else:
                    self._send_html(render_simple_upload_result(execute_folder_upload(config, params)))
            elif parsed.path == "/api/command-queue":
                params = parse_qs(parsed.query)
                if read_runtime_settings(config).get("command_auto_run"):
                    queue_command_packages(config)
                self._send_html(render_command_queue(list_command_packages(config), params))
            elif parsed.path == "/api/command-queue/sample":
                self._send_html(render_command_result("Sample Command Package", create_sample_command_package(config)))
            elif parsed.path == "/api/command-queue/queue":
                self._send_html(render_command_result("Queue Command Packages", queue_command_packages(config)))
            elif parsed.path == "/api/command-queue/execute":
                params = parse_qs(parsed.query)
                command_id = params.get("command_id", [""])[0].strip()
                limit = int(params.get("limit", ["1"])[0] or 1)
                self._send_html(render_command_result("Execute Command Queue", execute_command_queue(config, command_id=command_id, limit=limit)))
            elif parsed.path == "/api/command-queue/tick":
                params = parse_qs(parsed.query)
                limit = int(params.get("limit", ["1"])[0] or 1)
                self._send_html(render_command_result("Command Auto Tick", command_auto_tick(config, limit=limit)))
            elif parsed.path == "/api/command-queue/auto":
                params = parse_qs(parsed.query)
                enabled = params.get("enabled", [""])[0].lower() in {"1", "yes", "on", "true"}
                self._send_html(render_command_result("Command Auto Mode", set_command_auto_run(config, enabled)))
            elif parsed.path == "/api/command-queue/archive":
                params = parse_qs(parsed.query)
                command_id = params.get("command_id", [""])[0].strip()
                self._send_html(render_command_result("Archive Command", archive_command_package(config, command_id)))
            elif parsed.path in {"/api/gas-poll", "/api/gas-poll-execute"}:
                from gas_adapter import poll_gas_once

                params = parse_qs(parsed.query)
                execute_value = params.get("execute", [""])[0].strip().lower()
                execute_override = True if parsed.path == "/api/gas-poll-execute" else None
                if execute_value:
                    execute_override = execute_value in {"1", "yes", "on", "true"}
                wake_context = {
                    "poll_origin": params.get("poll_origin", ["manual_or_local"])[0],
                    "trigger_source": params.get("trigger_source", [""])[0],
                    "trigger_mode": params.get("trigger_mode", [""])[0],
                    "cloud_event_type": params.get("cloud_event_type", [""])[0],
                    "cloud_task_id": params.get("task_id", [""])[0],
                    "execute_override": execute_override,
                    "from": self.client_address[0] if self.client_address else "",
                }
                result = poll_gas_once(config, wake_context=wake_context, execute_override=execute_override)
                if params.get("response", [""])[0] == "json":
                    self._send_json(result)
                else:
                    title = "GAS Pull & Upload to NotebookLM" if execute_override is True else "GAS Pull from Cloud"
                    self._send_html(render_command_result(title, result))
            elif parsed.path == "/api/gas-test":
                from gas_adapter import test_gas_connection

                self._send_html(render_command_result("GAS Test Connection", test_gas_connection(config)))
            elif parsed.path == "/api/gas-settings":
                params = parse_qs(parsed.query)
                if params:
                    result = update_gas_settings_from_params(config, params)
                else:
                    result = read_gas_settings_result(config)
                self._send_html(render_command_result("GAS Settings", result))
            elif parsed.path == "/api/gas-settings/default":
                self._send_html(render_command_result("GAS Safe Default", apply_gas_safe_default(config)))
            elif parsed.path == "/api/gas-settings/toggle":
                params = parse_qs(parsed.query)
                self._send_html(render_command_result("GAS Toggle", toggle_gas_setting(config, params.get("key", [""])[0])))
            elif parsed.path == "/api/incoming-watch":
                params = parse_qs(parsed.query)
                execute = params.get("execute", [""])[0].lower() in {"1", "yes", "on", "true"}
                self._send_html(render_command_result("Incoming Watch Scan", incoming_watch_scan_once(config, execute=execute)))
            elif parsed.path == "/api/incoming-watch-settings":
                params = parse_qs(parsed.query)
                if params:
                    result = update_incoming_watch_settings_from_params(config, params)
                else:
                    result = incoming_watch_settings_result(config)
                self._send_html(render_command_result("Incoming Watch Settings", result))
            elif parsed.path == "/api/network-access":
                params = parse_qs(parsed.query)
                enabled = params.get("enabled", [""])[0].lower() in {"1", "yes", "on", "true"}
                self._send_html(render_command_result("Network Access", set_network_access(config, enabled)))
            elif parsed.path == "/api/notebooks/create":
                params = parse_qs(parsed.query)
                title = params.get("title", [""])[0]
                self._send_html(render_notebook_create_result(create_notebook(config, title)))
            elif parsed.path == "/api/logs":
                params = parse_qs(parsed.query)
                search = params.get("search", [""])[0]
                event = params.get("event", [""])[0]
                page = int(params.get("page", ["1"])[0] or 1)
                self._send_html(render_log_cms(query_runtime_logs(config, search=search, event=event, page=page)))
            elif parsed.path == "/api/adapter-execute":
                params = parse_qs(parsed.query)
                notebook_id = params.get("notebook_id", [""])[0].strip()
                project_id = params.get("project_id", [""])[0].strip()
                confirm = params.get("confirm", [""])[0].strip().lower() == "yes"
                self._send_html(render_execute_result(execute_notebooklm_adapter(config, notebook_id, confirm, project_id=project_id)))
            elif parsed.path == "/api/export-state-json":
                self._send_file(export_state_json(config), "application/json; charset=utf-8")
            elif parsed.path == "/api/export-excel":
                self._send_file(export_excel_report(config), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            elif parsed.path == "/api/export-local-logs-json":
                self._send_file(export_local_logs_json(config), "application/json; charset=utf-8")
            elif parsed.path == "/api/export-local-logs-excel":
                self._send_file(export_local_logs_excel(config), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            elif parsed.path == "/api/export-local-tasks-json":
                self._send_file(export_local_tasks_json(config), "application/json; charset=utf-8")
            elif parsed.path == "/api/export-local-tasks-excel":
                self._send_file(export_local_tasks_excel(config), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            elif parsed.path == "/api/clear-local-logs":
                self._send_html(render_command_result("Clear Local Logs", clear_local_logs(config)))
            elif parsed.path == "/api/clear-local-tasks":
                self._send_html(render_clear_result(clear_local_tasks(config)))
            elif parsed.path == "/api/clear":
                params = parse_qs(parsed.query)
                scope = params.get("scope", ["temp"])[0].strip()
                self._send_html(render_clear_result(clear_local_working_data(config, scope)))
            elif parsed.path == "/api/import-queue-json":
                self._send_html(render_command_result("Import Disabled", {"ok": False, "mode": "import_disabled", "error": "Import is intentionally disabled in v1.01 MVP. Export-only keeps the teaching runtime safer."}))
            elif parsed.path == "/api/import-queue-excel":
                self._send_html(render_command_result("Import Disabled", {"ok": False, "mode": "import_disabled", "error": "Import is intentionally disabled in v1.01 MVP. Export-only keeps the teaching runtime safer."}))
            elif parsed.path == "/docs/refactor_notes.html":
                self._send_file(PROJECT_ROOT / "docs" / "refactor_notes.html", "text/html; charset=utf-8")
            elif parsed.path == "/docs/student_guide.html":
                self._send_file(PROJECT_ROOT / "docs" / "student_guide.html", "text/html; charset=utf-8")
            elif parsed.path == "/docs/student_guide.md":
                self._send_file(PROJECT_ROOT / "docs" / "student_guide.md", "text/plain; charset=utf-8")
            elif parsed.path == "/api/queue/status":
                params = parse_qs(parsed.query)
                task_id = params.get("task_id", [""])[0].strip()
                if not task_id:
                    self._send_json({"ok": False, "error": "Task ID is required."})
                    return
                if task_queue_manager is None:
                    self._send_json({"ok": False, "error": "Task queue manager is not initialized."})
                    return
                detail = task_queue_manager.get_task_status_detail(task_id)
                if not detail:
                    self._send_json({"ok": False, "error": "Task not found."})
                    return
                self._send_json({"ok": True, **detail})
            elif parsed.path == "/api/queue/active":
                if task_queue_manager is None:
                    self._send_json({"ok": False, "error": "Task queue manager is not initialized."})
                    return
                active_list = task_queue_manager.get_active_tasks()
                self._send_json({"ok": True, "queue": active_list})
            elif parsed.path == "/api/multichat/sessions":
                params = parse_qs(parsed.query)
                is_audit = params.get("audit", [""])[0].strip() == "true"
                
                sess = self._get_current_session()
                current_user_id = sess.get("user_id") if sess else None
                current_display_name = sess.get("display_name") if sess else None
                current_role = sess.get("role") if sess else "user"
                
                show_all = (current_role == "admin" and is_audit)

                sessions_file = Path(config.project_root) / "data" / "multichat_sessions.json"
                import json
                sessions_list = []
                if sessions_file.exists():
                    try:
                        with open(sessions_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        for cid, info in data.get("sessions", {}).items():
                            owner_id = info.get("user_id")
                            owner_name = info.get("user_name")
                            if not show_all:
                                is_owner = (owner_id == current_user_id) or (not owner_id and owner_name == current_display_name)
                                if not is_owner:
                                    continue
                            last_q = ""
                            if info.get("turns"):
                                for turn in reversed(info["turns"]):
                                    if turn.get("role") == "user":
                                        last_q = turn.get("content", "")
                                        break
                            sessions_list.append({
                                "conversation_id": cid,
                                "user_id": owner_id,
                                "user_name": owner_name or "Unknown",
                                "notebook_id": info.get("notebook_id", ""),
                                "notebook_title": info.get("notebook_title", ""),
                                "created_at": info.get("created_at", ""),
                                "last_query_at": info.get("last_query_at", ""),
                                "last_question": last_q,
                                "remark": info.get("remark", "")
                            })
                    except Exception:
                        pass
                sessions_list.sort(key=lambda x: x["last_query_at"], reverse=True)
                self._send_json({"ok": True, "sessions": sessions_list})
            elif parsed.path == "/api/multichat/history":
                params = parse_qs(parsed.query)
                conversation_id = params.get("conversation_id", [""])[0].strip()
                if not conversation_id:
                    self._send_json({"ok": False, "error": "Conversation ID is required."})
                    return
                sess = self._get_current_session()
                current_user_id = sess.get("user_id") if sess else None
                current_display_name = sess.get("display_name") if sess else None
                current_role = sess.get("role") if sess else "user"

                sessions_file = Path(config.project_root) / "data" / "multichat_sessions.json"
                import json
                turns = []
                if sessions_file.exists():
                    try:
                        with open(sessions_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        if "sessions" in data and conversation_id in data["sessions"]:
                            session_info = data["sessions"][conversation_id]
                            if current_role != "admin":
                                owner_id = session_info.get("user_id")
                                owner_name = session_info.get("user_name")
                                is_owner = (owner_id == current_user_id) or (not owner_id and owner_name == current_display_name)
                                if not is_owner:
                                    self._send_json({"ok": False, "error": "Access denied: this session belongs to another user."})
                                    return
                            turns = session_info.get("turns", [])
                    except Exception:
                        pass
                self._send_json({"ok": True, "turns": turns})
            elif parsed.path == "/api/gemini/sessions":
                params = parse_qs(parsed.query)
                is_audit = params.get("audit", [""])[0].strip() == "true"
                
                sess = self._get_current_session()
                current_user_id = sess.get("user_id") if sess else None
                current_display_name = sess.get("display_name") if sess else None
                current_role = sess.get("role") if sess else "user"
                
                show_all = (current_role == "admin" and is_audit)

                sessions_file = Path(config.project_root) / "data" / "gemini_sessions.json"
                import json
                sessions_list = []
                if sessions_file.exists():
                    try:
                        with open(sessions_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        for cid, info in data.get("sessions", {}).items():
                            owner_id = info.get("user_id")
                            owner_name = info.get("user_name")
                            if not show_all:
                                is_owner = (owner_id == current_user_id) or (not owner_id and owner_name == current_display_name)
                                if not is_owner:
                                    continue
                            last_q = ""
                            if info.get("turns"):
                                for turn in reversed(info["turns"]):
                                    if turn.get("role") == "user":
                                        last_q = turn.get("content", "")
                                        break
                            sessions_list.append({
                                "conversation_id": cid,
                                "user_id": owner_id,
                                "user_name": owner_name or "Unknown",
                                "created_at": info.get("created_at", ""),
                                "last_query_at": info.get("last_query_at", ""),
                                "last_question": last_q,
                                "metadata": json.dumps(info.get("metadata", [])),
                                "remark": info.get("remark", "")
                            })
                    except Exception:
                        pass
                sessions_list.sort(key=lambda x: x["last_query_at"], reverse=True)
                self._send_json({"ok": True, "sessions": sessions_list})
            elif parsed.path == "/api/gemini/history":
                params = parse_qs(parsed.query)
                conversation_id = params.get("conversation_id", [""])[0].strip()
                if not conversation_id:
                    self._send_json({"ok": False, "error": "Conversation ID is required."})
                    return
                sess = self._get_current_session()
                current_user_id = sess.get("user_id") if sess else None
                current_display_name = sess.get("display_name") if sess else None
                current_role = sess.get("role") if sess else "user"

                sessions_file = Path(config.project_root) / "data" / "gemini_sessions.json"
                import json
                turns = []
                if sessions_file.exists():
                    try:
                        with open(sessions_file, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        if "sessions" in data and conversation_id in data["sessions"]:
                            session_info = data["sessions"][conversation_id]
                            if current_role != "admin":
                                owner_id = session_info.get("user_id")
                                owner_name = session_info.get("user_name")
                                is_owner = (owner_id == current_user_id) or (not owner_id and owner_name == current_display_name)
                                if not is_owner:
                                    self._send_json({"ok": False, "error": "Access denied: this session belongs to another user."})
                                    return
                            turns = session_info.get("turns", [])
                    except Exception:
                        pass
                self._send_json({"ok": True, "turns": turns})
            else:
                self.send_error(404, "Not found")


        def log_message(self, format: str, *args: object) -> None:
            print(f"[portal] {self.address_string()} - {format % args}")

        def _send_json(self, payload: Dict[str, object]) -> None:
            body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_html(self, html: str) -> None:
            body = html.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_file(self, path: Path, content_type: str) -> None:
            if not path.exists():
                self.send_error(404, "Not found")
                return
            body = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    return RuntimeHandler


def render_index(payload: Dict[str, object]) -> str:
    environment = payload["environment"]
    files = payload["files"]
    runtime_identity = payload.get("runtime_identity", {})
    config = load_or_create_config(PROJECT_ROOT)
    active = get_active_project(config)
    active_project = active.get("project", {})
    active_name = active_project.get("name", "尚未選擇 Project")
    active_project_id = active.get("active_project_id", "")
    active_notebook_id = active_project.get("notebook_id", "")
    env_rows = "\n".join(
        f"""
        <tr>
          <td><span class="badge {'ok' if item['ok'] else 'warn'}">{'OK' if item['ok'] else 'NEEDS ACTION'}</span></td>
          <td>{escape_html(key)}</td>
          <td>{escape_html(item['message'])}<br><code>{escape_html(item['path'])}</code></td>
        </tr>
        """
        for key, item in environment.items()
    )
    file_rows = "\n".join(
        f"""
        <tr>
          <td>{escape_html(item['name'])}</td>
          <td>{item['size_kb']}</td>
          <td>{escape_html(item['suffix'] or '(no ext)')}</td>
        </tr>
        """
        for item in files
    ) or '<tr><td colspan="3" class="muted">data/inbox 目前沒有可處理檔案</td></tr>'

    projects = list_projects(config, sort="name", page_size=100).get("projects", [])
    project_options = "\n".join(
        f'<option value="{escape_html(project.get("project_id", ""))}" {"selected" if project.get("project_id", "") == active_project_id else ""}>{escape_html(project.get("name", ""))} · {escape_html(project.get("notebook_id", ""))}</option>'
        for project in projects
    ) or '<option value="">No project yet - sync projects first</option>'
    command_queue = list_command_packages(config)
    command_counts = command_queue.get("counts", {})
    auto_run_label = "ON" if command_queue.get("auto_run") else "OFF"
    runtime_settings = read_runtime_settings(config)
    allow_network_access = bool(runtime_settings.get("allow_network_access", True))
    incoming_watch_enabled = bool(runtime_settings.get("incoming_watch_enabled", False))
    incoming_watch_mode = str(runtime_settings.get("incoming_watch_mode", "polling"))
    incoming_watch_auto_queue = bool(runtime_settings.get("incoming_watch_auto_queue", True))
    incoming_watch_auto_execute = bool(runtime_settings.get("incoming_watch_auto_execute", False))
    incoming_watch_folder = str(runtime_settings.get("incoming_watch_folder") or simple_incoming_dir(config))
    incoming_watch_types = parse_extensions(list(runtime_settings.get("incoming_watch_file_types") or DEFAULT_SIMPLE_TYPES))
    incoming_watch_recursive = bool(runtime_settings.get("incoming_watch_recursive", False))
    realtime_engine = incoming_realtime_engine_status()
    lan_ip = detect_lan_ip()
    gas_settings = read_gas_settings(config)
    gas_url = str(gas_settings.get("web_app_url", ""))
    gas_enabled = bool(gas_settings.get("enabled"))
    gas_auto_poll = bool(gas_settings.get("auto_poll_enabled"))
    gas_auto_execute = bool(gas_settings.get("auto_execute"))
    gas_token = str(gas_settings.get("api_token", ""))
    gas_token_label = "(set)" if gas_token else "(missing)"
    gas_download_root = str(gas_settings.get("local_download_root", "data/gas_downloads"))
    identity_rows = "\n".join(
        f"<tr><td>{escape_html(label)}</td><td><code>{escape_html(value)}</code></td></tr>"
        for label, value in [
            ("Computer", runtime_identity.get("computer_name", "")),
            ("OS", runtime_identity.get("os", "")),
            ("Host / Port", f"{runtime_identity.get('bind_host', '')}:{runtime_identity.get('port', '')}"),
            ("Local URL", runtime_identity.get("local_url", "")),
            ("LAN URL", runtime_identity.get("lan_url", "") or "(local only)"),
            ("Python", runtime_identity.get("python_version", "")),
            ("Project Path", runtime_identity.get("project_root", "")),
        ]
    )

    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>AI NotebookLM Runtime Lab Portal</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Noto Sans TC", sans-serif; color: #202124; background: #f6f7f4; }}
    main {{ max-width: 1080px; margin: 0 auto; padding: 32px 20px 56px; }}
    header {{ border-bottom: 4px solid #1f6f5b; padding-bottom: 18px; margin-bottom: 24px; }}
    h1 {{ margin: 0 0 8px; font-size: 32px; }}
    h2 {{ margin-top: 28px; color: #174f43; }}
    .meta, .muted {{ color: #667085; }}
    .grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; }}
    .panel {{ background: white; border: 1px solid #d8ddd8; border-radius: 8px; padding: 16px; }}
    .identity-panel {{ background: #fff; border: 1px solid #b8d9cd; border-left: 6px solid #315a7d; border-radius: 8px; padding: 16px; margin-top: 18px; }}
    .identity-panel h2 {{ margin-top: 0; }}
    .identity-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
    .status-panel {{ background: #edf7f1; border: 1px solid #9fd6b5; border-left: 6px solid #208454; border-radius: 8px; padding: 16px; margin-top: 18px; }}
    .status-panel.warn {{ background: #fff8e6; border-color: #f0c36d; border-left-color: #b87900; }}
    .section-band {{ border-top: 3px solid #d8ddd8; padding-top: 18px; margin-top: 28px; }}
    .tabbar {{ display: flex; gap: 8px; flex-wrap: wrap; margin: 22px 0; border-bottom: 2px solid #d8ddd8; padding-bottom: 10px; }}
    .tab-button {{ padding: 10px 14px; border: 0; border-radius: 6px; background: #e8f1ed; color: #174f43; font: inherit; font-weight: 700; cursor: pointer; }}
    .tab-button.active {{ background: #1f6f5b; color: #fff; }}
    .tab-section {{ display: none; }}
    .tab-section.active {{ display: block; }}
    .upload-card {{ background: #fff; border: 1px solid #d8ddd8; border-left: 6px solid #315a7d; border-radius: 8px; padding: 18px; }}
    input[type=file] {{ width: 100%; padding: 18px; border: 2px dashed #9bb7aa; border-radius: 8px; background: #f8fbf9; }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    th, td {{ border: 1px solid #d8ddd8; padding: 10px; vertical-align: top; text-align: left; }}
    th {{ background: #e8f1ed; color: #174f43; }}
    code {{ font-family: Menlo, Consolas, monospace; font-size: 13px; }}
    .badge {{ display: inline-block; min-width: 92px; text-align: center; padding: 3px 7px; border-radius: 999px; font-size: 12px; font-weight: 700; }}
    .ok {{ background: #ddf5e8; color: #176c3a; }}
    .warn {{ background: #fff0d5; color: #9a5b00; }}
    .button {{ display: inline-block; margin-top: 10px; padding: 9px 12px; border: 0; border-radius: 6px; background: #1f6f5b; color: white; text-decoration: none; font: inherit; cursor: pointer; }}
    .button.secondary {{ background: #315a7d; }}
    .button.gold {{ background: #a46700; }}
    .button.toggle-on {{ background: #1f6f5b; }}
    .button.toggle-off {{ background: #8f2f2f; }}
    .danger {{ background: #8f2f2f !important; }}
    .muted-invert {{ opacity: 0.82; font-size: 0.92em; }}
    .button.is-running {{ background: #a46700 !important; color: #fff; }}
    .button.is-done {{ background: #208454 !important; color: #fff; }}
    .button.is-failed {{ background: #9d2f2f !important; color: #fff; }}
    .formrow {{ display: flex; gap: 8px; flex-wrap: wrap; align-items: center; margin-top: 10px; }}
    .brandline {{ color: #1f6f5b; font-weight: 800; margin-top: 4px; }}
    .watermark {{ position: fixed; right: 18px; bottom: 12px; color: rgba(31, 111, 91, 0.22); font-size: 18px; font-weight: 800; pointer-events: none; }}
    .type-grid {{ display: flex; gap: 12px; flex-wrap: wrap; margin: 8px 0 14px; }}
    .subpanel {{ background: #f8fbf9; border: 1px solid #d8ddd8; border-radius: 8px; padding: 14px; margin-top: 14px; }}
    input[type=text], input[type=password], input[type=number] {{ min-width: 420px; padding: 8px; border: 1px solid #cfd8d3; border-radius: 6px; }}
    input[type=checkbox] {{ transform: scale(1.15); margin-right: 6px; }}
    select {{ min-width: 360px; padding: 8px; border: 1px solid #cfd8d3; border-radius: 6px; background: white; }}
    label.inline {{ display: inline-flex; align-items: center; gap: 4px; margin: 6px 14px 6px 0; }}
    nav.tabs {{ display: flex; gap: 8px; flex-wrap: wrap; margin: 18px 0; }}
    @media (max-width: 760px) {{ .grid {{ grid-template-columns: 1fr; }} }}
  </style>
  <script>
    function remember(action) {{
      const log = JSON.parse(localStorage.getItem('falo_etl_browser_log') || '[]');
      log.unshift({{ action, at: new Date().toISOString() }});
      localStorage.setItem('falo_etl_browser_log', JSON.stringify(log.slice(0, 50)));
      localStorage.setItem('falo_etl_last_action', action);
    }}
    function clearBrowserCache() {{
      localStorage.removeItem('falo_etl_browser_log');
      localStorage.removeItem('falo_etl_last_action');
      alert('Browser cache cleared for this portal.');
    }}
    function goImport(kind) {{
      const input = document.getElementById(kind + '_path');
      const path = encodeURIComponent(input.value.trim());
      if (!path) {{ alert('Please enter a local path.'); return false; }}
      remember(kind);
      window.location.href = '/api/' + kind + '?path=' + path;
      return false;
    }}
    function switchTab(tabId, button) {{
      document.querySelectorAll('.tab-section').forEach(item => item.classList.remove('active'));
      document.querySelectorAll('.tab-button').forEach(item => item.classList.remove('active'));
      document.getElementById(tabId).classList.add('active');
      button.classList.add('active');
      localStorage.setItem('falo_etl_active_tab', tabId);
    }}
    window.addEventListener('DOMContentLoaded', () => {{
      const saved = localStorage.getItem('falo_etl_active_tab') || 'tab-simple-upload';
      const button = document.querySelector(`[data-tab="${{saved}}"]`) || document.querySelector('[data-tab="tab-simple-upload"]');
      if (button) switchTab(button.dataset.tab, button);
      loadSessions();
    }});
    function toggleTypes(formId, checked) {{
      document.querySelectorAll(`#${{formId}} input[name="types"]`).forEach(item => item.checked = checked);
    }}
    function faloAction(el, label) {{
      if (!el || el.dataset.skipState === 'yes') return true;
      const running = label || el.dataset.runningLabel || 'Running...';
      el.dataset.originalText = el.textContent.trim();
      el.textContent = running;
      el.classList.add('is-running');
      if (el.tagName === 'BUTTON' && el.type !== 'submit') el.disabled = true;
      return true;
    }}
    document.addEventListener('click', event => {{
      const el = event.target.closest('a.button, button.button');
      if (!el) return;
      if (el.dataset.noActionState === 'yes') return;
      if (el.getAttribute('href') === '#') return;
      if (el.classList.contains('danger') && !el.dataset.confirmedState) return;
      faloAction(el);
    }});

    async function convertMeeting() {{
      const btn = document.getElementById('btn_convert');
      const md = document.getElementById('markdown_content').value.trim();
      if (!md) {{ alert('請貼入 Markdown 會議記錄內容。'); return; }}
      
      btn.disabled = true;
      const originalText = btn.textContent;
      btn.textContent = 'Converting...';
      btn.classList.add('is-running');
      
      const resultPanel = document.getElementById('converter_result_panel');
      const msgDiv = document.getElementById('converter_message');
      const linksDiv = document.getElementById('download_links');
      const summaryPre = document.getElementById('converter_summary');
      
      resultPanel.style.display = 'none';
      
      try {{
        const response = await fetch('/api/convert-meeting', {{
          method: 'POST',
          headers: {{
            'Content-Type': 'application/x-www-form-urlencoded'
          }},
          body: 'markdown=' + encodeURIComponent(md)
        }});
        const res = await response.json();
        resultPanel.style.display = 'block';
        if (res.ok) {{
          msgDiv.style.background = '#edf7f1';
          msgDiv.style.color = '#208454';
          msgDiv.style.border = '1px solid #9fd6b5';
          msgDiv.innerHTML = '<strong>轉換成功！</strong> 請點擊下方連結下載對應格式的檔案。';
          
          let html = '';
          if (res.docx) {{
            html += `<a class="button" href="/api/meeting-download?filename=${{encodeURIComponent(res.docx)}}">Download Word (.docx)</a> `;
          }}
          if (res.single_html) {{
            html += `<a class="button secondary" href="/api/meeting-download?filename=${{encodeURIComponent(res.single_html)}}" target="_blank">View Single HTML</a> `;
          }}
          if (res.rwd_html) {{
            html += `<a class="button secondary" href="/api/meeting-download?filename=${{encodeURIComponent(res.rwd_html)}}" target="_blank">View RWD HTML</a> `;
          }}
          linksDiv.innerHTML = html;
          summaryPre.textContent = res.summary || '無摘要資訊';
        }} else {{
          msgDiv.style.background = '#fff0f0';
          msgDiv.style.color = '#9d2f2f';
          msgDiv.style.border = '1px solid #f9c5c5';
          msgDiv.innerHTML = '<strong>轉換失敗！</strong> ' + (res.error || '未知錯誤');
          linksDiv.innerHTML = '';
          summaryPre.textContent = '';
        }}
      }} catch (err) {{
        resultPanel.style.display = 'block';
        msgDiv.style.background = '#fff0f0';
        msgDiv.style.color = '#9d2f2f';
        msgDiv.style.border = '1px solid #f9c5c5';
        msgDiv.innerHTML = '<strong>請求錯誤：</strong> ' + err.message;
        linksDiv.innerHTML = '';
        summaryPre.textContent = '';
      }} finally {{
        btn.disabled = false;
        btn.textContent = originalText;
        btn.classList.remove('is-running');
      }}
    }}

    async function loadSessions() {{
      try {{
        const response = await fetch('/api/multichat/sessions');
        const res = await response.json();
        if (res.ok) {{
          const select = document.getElementById('chat_session_select');
          select.innerHTML = '<option value="new">-- 開啟全新對話 (Create New Chat) --</option>';
          
          res.sessions.forEach(s => {{
            const opt = document.createElement('option');
            opt.value = s.conversation_id;
            opt.textContent = `${{s.user_name}} (${{s.created_at}}) - ${{s.last_question.substring(0, 30)}}...`;
            select.appendChild(opt);
          }});
          
          const tbody = document.getElementById('admin_sessions_table_body');
          if (res.sessions.length === 0) {{
            tbody.innerHTML = '<tr><td colspan="5" style="text-align:center; padding:20px; color:#66706b;">尚無地端對話紀錄</td></tr>';
          }} else {{
            let html = '';
            res.sessions.forEach(s => {{
              html += `
                <tr>
                  <td style="padding:10px; border:1px solid #d8ddd8;"><strong>${{escapeHtml(s.user_name)}}</strong></td>
                  <td style="padding:10px; border:1px solid #d8ddd8;"><code style="font-size:12px;">${{s.conversation_id}}</code></td>
                  <td style="padding:10px; border:1px solid #d8ddd8; font-size:14px; max-width:300px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap;" title="${{escapeHtml(s.last_question)}}">${{escapeHtml(s.last_question)}}</td>
                  <td style="padding:10px; border:1px solid #d8ddd8; font-size:13px; color:#66706b;">${{s.last_query_at}}</td>
                  <td style="padding:10px; border:1px solid #d8ddd8;">
                    <a class="button secondary" style="padding:4px 8px; font-size:12px;" href="#" onclick="selectSessionFromTable('${{s.conversation_id}}', '${{escapeHtml(s.user_name)}}'); return false;">載入對話</a>
                    <a class="button danger" style="padding:4px 8px; font-size:12px; background:#d9534f;" href="#" onclick="if(confirm('確定要刪除此地端對話紀錄嗎？雲端不會自動刪除。')){{deleteSession('${{s.conversation_id}}');}} return false;">刪除</a>
                  </td>
                </tr>
              `;
            }});
            tbody.innerHTML = html;
          }}
        }}
      }} catch (err) {{
        console.error('Failed to load sessions:', err);
      }}
    }}
    
    function escapeHtml(string) {{
      const map = {{
        '&': '&amp;',
        '<': '&lt;',
        '>': '&gt;',
        '"': '&quot;',
        "'": '&#039;'
      }};
      return String(string).replace(/[&<>"']/g, function(m) {{ return map[m]; }});
    }}

    async function selectSessionFromTable(convId, userName) {{
      document.getElementById('chat_session_select').value = convId;
      document.getElementById('chat_user_name').value = userName;
      await loadSessionHistory(convId);
    }}

    async function onSessionSelectChange() {{
      const select = document.getElementById('chat_session_select');
      const convId = select.value;
      if (convId === 'new') {{
        document.getElementById('chat_box_conversation').innerHTML = '<div style="color:#66706b; text-align:center; margin-top:100px;">請選擇對話或輸入問題開始問答</div>';
      }} else {{
        await loadSessionHistory(convId);
      }}
    }}

    async function loadSessionHistory(convId) {{
      const chatBox = document.getElementById('chat_box_conversation');
      chatBox.innerHTML = '<div style="color:#66706b; text-align:center; margin-top:100px;">載入對話歷史中...</div>';
      try {{
        const response = await fetch(`/api/multichat/history?conversation_id=${{encodeURIComponent(convId)}}`);
        const res = await response.json();
        if (res.ok && res.turns) {{
          if (res.turns.length === 0) {{
            chatBox.innerHTML = '<div style="color:#66706b; text-align:center; margin-top:100px;">無對話內容</div>';
            return;
          }}
          let html = '';
          res.turns.forEach(t => {{
            const isUser = t.role === 'user';
            const bg = isUser ? '#e8f1ed' : '#ffffff';
            const align = isUser ? 'flex-end' : 'flex-start';
            const color = isUser ? '#174f43' : '#2c302e';
            const border = isUser ? '1px solid #9fd6b5' : '1px solid #d8ddd8';
            html += `
              <div style="display:flex; justify-content:${{align}}; margin-bottom:10px;">
                <div style="max-width:85%; background:${{bg}}; color:${{color}}; border:${{border}}; padding:10px 14px; border-radius:10px; font-size:14px; white-space:pre-wrap; box-shadow: 0 2px 4px rgba(0,0,0,0.02);">
                  <div style="font-size:11px; color:#66706b; margin-bottom:4px; font-weight:bold;">${{isUser ? 'USER' : 'NOTEBOOKLM'}} - ${{t.timestamp}}</div>
                  ${{escapeHtml(t.content)}}
                </div>
              </div>
            `;
          }});
          chatBox.innerHTML = html;
          chatBox.scrollTop = chatBox.scrollHeight;
        }}
      }} catch (err) {{
        chatBox.innerHTML = `<div style="color:#9d2f2f; text-align:center; margin-top:100px;">載入失敗: ${{escapeHtml(err.message)}}</div>`;
      }}
    }}

    async function sendChatMessage() {{
      const btn = document.getElementById('btn_send_chat');
      const notebookId = document.getElementById('chat_notebook_id').value.trim();
      const userName = document.getElementById('chat_user_name').value.trim();
      const sessionSelect = document.getElementById('chat_session_select');
      const conversationId = sessionSelect.value;
      const question = document.getElementById('chat_question').value.trim();

      if (!notebookId || !userName || !question) {{
        alert('請填寫 Notebook ID、姓名與提問內容。');
        return;
      }}

      btn.disabled = true;
      const originalText = btn.textContent;
      btn.textContent = 'Querying NotebookLM (5~8s)...';
      btn.classList.add('is-running');

      const chatBox = document.getElementById('chat_box_conversation');
      if (chatBox.querySelector('div[style*="text-align:center"]')) {{
        chatBox.innerHTML = '';
      }}

      const tempUserBubble = document.createElement('div');
      tempUserBubble.style.display = 'flex';
      tempUserBubble.style.justifyContent = 'flex-end';
      tempUserBubble.style.marginBottom = '10px';
      tempUserBubble.innerHTML = `
        <div style="max-width:85%; background:#e8f1ed; color:#174f43; border:1px solid #9fd6b5; padding:10px 14px; border-radius:10px; font-size:14px; white-space:pre-wrap; opacity:0.7;">
          <div style="font-size:11px; color:#66706b; margin-bottom:4px; font-weight:bold;">USER (傳送中...)</div>
          ${{escapeHtml(question)}}
        </div>
      `;
      chatBox.appendChild(tempUserBubble);
      chatBox.scrollTop = chatBox.scrollHeight;

      try {{
        const response = await fetch('/api/multichat/ask', {{
          method: 'POST',
          headers: {{
            'Content-Type': 'application/x-www-form-urlencoded'
          }},
          body: `notebook_id=${{encodeURIComponent(notebookId)}}&user_name=${{encodeURIComponent(userName)}}&conversation_id=${{encodeURIComponent(conversationId)}}&question=${{encodeURIComponent(question)}}`
        }});
        
        const res = await response.json();
        if (res.ok) {{
          document.getElementById('chat_question').value = '';
          await loadSessions();
          document.getElementById('chat_session_select').value = res.conversation_id;
          await loadSessionHistory(res.conversation_id);
        }} else {{
          alert('發問失敗：' + (res.error || '未知錯誤'));
          tempUserBubble.remove();
        }}
      }} catch (err) {{
        alert('請求錯誤：' + err.message);
        tempUserBubble.remove();
      }} finally {{
        btn.disabled = false;
        btn.textContent = originalText;
        btn.classList.remove('is-running');
      }}
    }}

    async function deleteSession(convId) {{
      try {{
        const response = await fetch('/api/multichat/delete', {{
          method: 'POST',
          headers: {{
            'Content-Type': 'application/x-www-form-urlencoded'
          }},
          body: `conversation_id=${{encodeURIComponent(convId)}}`
        }});
        const res = await response.json();
        if (res.ok) {{
          const select = document.getElementById('chat_session_select');
          if (select.value === convId) {{
            select.value = 'new';
            document.getElementById('chat_box_conversation').innerHTML = '<div style="color:#66706b; text-align:center; margin-top:100px;">請選擇對話或輸入問題開始問答</div>';
          }}
          await loadSessions();
        }}
      }} catch (err) {{
        alert('刪除失敗：' + err.message);
      }}
    }}
  </script>

</head>
<body>
<main>
  <header>
    <h1>AI NotebookLM Runtime Lab</h1>
    <div class="meta">NotebookLM ETL Runtime Lab · {APP_VERSION} · HTML portal · Python runtime · 0.0.0.0 capable service</div>
    <div class="brandline">{APP_WATERMARK} · Teaching Runtime MVP</div>
  </header>
  <section class="grid">
    <div class="panel"><strong>UI</strong><br>HTML 是 portal，不是核心。</div>
    <div class="panel"><strong>Runtime</strong><br>Python service 才是真核心。</div>
    <div class="panel"><strong>ETL</strong><br>NotebookLM 前面的整理層才是價值點。</div>
  </section>
  <section class="identity-panel">
    <h2>Runtime Identity</h2>
    <p class="muted">Win / Mac 混合部署時，先確認現在是哪一台 worker 在執行。Falo x Force 教學原則：先辨識主機，再執行任務。</p>
    <div class="identity-grid">
      <table>{identity_rows}</table>
      <div class="panel">
        <strong>Worker Role</strong><br>
        {escape_html(runtime_identity.get("worker_role", "local_runtime_worker"))}<br><br>
        <strong>用途</strong><br>
        這台負責接收 command queue、GAS polling、NotebookLM upload 與本機 log。<br><br>
        <a class="button secondary" href="/api/status">Open JSON Status</a>
      </div>
    </div>
  </section>
  <section class="status-panel {'warn' if not active_project_id else ''}">
    <strong>Active Project</strong><br>
    {escape_html(active_name)}<br>
    <span class="muted">project_id: <code>{escape_html(active_project_id or '(not selected)')}</code></span><br>
    <span class="muted">notebook_id: <code>{escape_html(active_notebook_id or '(not selected)')}</code></span>
  </section>
  <div class="tabbar">
    <button class="tab-button active" data-tab="tab-simple-upload" onclick="switchTab('tab-simple-upload', this)">Tab 1 Simple Upload</button>
    <button class="tab-button" data-tab="tab-etl" onclick="switchTab('tab-etl', this)">Tab 2 Excel / CSV ETL</button>
    <button class="tab-button" data-tab="tab-projects" onclick="switchTab('tab-projects', this)">Tab 3 Projects</button>
    <button class="tab-button" data-tab="tab-governance" onclick="switchTab('tab-governance', this)">Tab 4 Logs / Governance</button>
    <button class="tab-button" data-tab="tab-command-queue" onclick="switchTab('tab-command-queue', this)">Tab 5 Command Queue</button>
    <button class="tab-button" data-tab="tab-gas-cloud" onclick="switchTab('tab-gas-cloud', this)">Tab 6 GAS Cloud</button>
    <button class="tab-button" data-tab="tab-incoming-watch" onclick="switchTab('tab-incoming-watch', this)">Tab 7 Incoming Watch</button>
    <button class="tab-button" data-tab="tab-meeting-converter" onclick="switchTab('tab-meeting-converter', this)">Tab 8 Meeting Converter</button>
    <button class="tab-button" data-tab="tab-multi-chat" onclick="switchTab('tab-multi-chat', this)">Tab 9 Multi-User Chat</button>
  </div>
  <section id="tab-simple-upload" class="tab-section active">
    <h2>Simple Upload</h2>
    <p>這是主要入口：選 Project、選檔案、處理重名策略、上傳到 NotebookLM。</p>
    <form class="upload-card" action="/api/simple-upload" method="post" enctype="multipart/form-data">
      <label><strong>Target Project</strong></label><br>
      <select name="project_id">{project_options}</select>
      <a class="button secondary" href="/api/projects">Manage Projects</a>
      <div class="formrow">
        <input id="quick_notebook_title" type="text" placeholder="New Notebook title">
        <a class="button gold" href="#" onclick="const title=encodeURIComponent(document.getElementById('quick_notebook_title').value.trim()); if(!title){{alert('Please enter a notebook title.'); return false;}} window.location.href='/api/simple-upload/create-project?title='+title; return false;">Create & Use</a>
      </div>
      <p class="muted">沒有另外填 notebook_id；系統會從 Project metadata 取得目的地。</p>
      <label><strong>Files</strong></label><br>
      <input type="file" name="files" multiple>
      <p><strong>Duplicate policy</strong></p>
      <label><input type="radio" name="conflict_policy" value="rename" checked> Rename and upload</label>
      <label><input type="radio" name="conflict_policy" value="skip"> Skip duplicates</label>
      <label><input type="radio" name="conflict_policy" value="replace"> Replace if possible</label>
      <label><input type="radio" name="conflict_policy" value="upload_anyway"> Upload anyway</label>
      <br>
      <button class="button" type="submit" onclick="remember('simple_upload')">Upload Files</button>
    </form>
    <div class="upload-card">
      <h3>Scan Server Folder</h3>
      <p class="muted">預設掃描專案內 Source Pool；也可指定 server 上任何可讀資料夾。先 Preview，再 Upload。</p>
      <form id="folderScanForm" action="/api/folder-scan" method="get">
        <label><strong>Folder path</strong></label><br>
        <input type="text" name="folder_path" value="{escape_html(str(simple_incoming_dir(config)))}">
        <p><strong>File types</strong></p>
        <div class="formrow">
          <button class="button secondary" type="button" onclick="toggleTypes('folderScanForm', true)">Select All</button>
          <button class="button secondary" type="button" onclick="toggleTypes('folderScanForm', false)">Clear All</button>
        </div>
        <div class="type-grid">
          {''.join(f'<label><input type="checkbox" name="types" value="{ext}" checked> {ext}</label>' for ext in DEFAULT_SIMPLE_TYPES)}
        </div>
        <label><input type="checkbox" name="recursive" value="yes"> Include subfolders</label>
        <label>Order <select name="order"><option value="name">Name A-Z</option><option value="modified_asc">Modified old first</option><option value="modified_desc">Modified new first</option></select></label>
        <p><strong>Duplicate policy</strong></p>
        <label><input type="radio" name="conflict_policy" value="rename" checked> Rename and upload</label>
        <label><input type="radio" name="conflict_policy" value="skip"> Skip duplicates</label>
        <label><input type="radio" name="conflict_policy" value="replace"> Replace if possible</label>
        <label><input type="radio" name="conflict_policy" value="upload_anyway"> Upload anyway</label>
        <p><strong>Evidence copy</strong></p>
        <input type="text" name="evidence_root" value="{escape_html(str(simple_evidence_dir(config)))}">
        <input type="hidden" name="project_id" value="{escape_html(active_project_id)}">
        <br>
        <button class="button" type="submit">Scan Preview</button>
      </form>
    </div>
  </section>
  <section id="tab-etl" class="tab-section">
  <section id="dashboard">
    <h2>Environment</h2>
    <table>
      <thead><tr><th>Status</th><th>Item</th><th>Detail</th></tr></thead>
      <tbody>{env_rows}</tbody>
    </table>
  </section>
  <section>
    <h2>Inbox Files</h2>
    <table>
      <thead><tr><th>Name</th><th>Size KB</th><th>Type</th></tr></thead>
      <tbody>{file_rows}</tbody>
    </table>
  </section>
  <section>
    <h2>Runtime Notes</h2>
    <p>這裡是進階 ETL：Excel 拆 CSV、建立 queue、預覽 adapter、匯出入閉環。</p>
    <a class="button" href="/api/normalize" onclick="remember('normalize')">Run XLSX Normalize</a>
    <a class="button" href="/api/build-queue" onclick="remember('build_queue')">Build Queue Manifest</a>
    <a class="button" href="/api/adapter-preview" onclick="remember('adapter_preview')">Preview NotebookLM Adapter</a>
    <a class="button" href="/api/notebooks" onclick="remember('list_notebooks')">List Notebooks</a>
    <a class="button" href="/api/projects" onclick="remember('projects')">Project Manager</a>
    <a class="button" href="/api/projects/sync" onclick="remember('project_sync')">Sync Notebooks to Projects</a>
    <a class="button" href="/api/logs" onclick="remember('log_cms')">Open Log CMS</a>
    <a class="button" href="/api/export-state-json" onclick="remember('export_state_json')">Export State JSON</a>
    <a class="button" href="/api/export-excel" onclick="remember('export_excel')">Export Excel Report</a>
    <a class="button" href="/docs/refactor_notes.html">Open Refactor Notes</a>
    <a class="button" href="/docs/student_guide.html">Open Student Guide</a>
    <a class="button" href="/api/status">Open JSON Status</a>
    <a class="button" href="/api/runtime-state">Open Runtime State</a>
  </section>
  </section>
  <section id="tab-projects" class="tab-section">
  <section>
    <h2>Notebook Controls</h2>
    <p>先列出既有 Notebook 取得 ID；需要新案件時，再建立新的 Notebook。</p>
    <div class="formrow">
      <input id="notebook_title" type="text" placeholder="New Notebook title">
      <a class="button" href="#" onclick="const title=encodeURIComponent(document.getElementById('notebook_title').value.trim()); if(!title){{alert('Please enter a title.'); return false;}} remember('create_notebook'); window.location.href='/api/notebooks/create?title='+title; return false;">Create Notebook</a>
      <a class="button secondary" href="/api/projects">Manage Projects</a>
      <a class="button gold" href="/api/projects/sync">Sync NotebookLM</a>
    </div>
  </section>
  <section class="section-band">
    <h2>Advanced Project Adapter</h2>
    <p>給 runtime debug：使用 Active Project 或手動 notebook_id 預覽 adapter，不是一般上傳入口。</p>
    <form class="formrow" action="/api/adapter-preview" method="get">
      <input type="text" name="notebook_id" placeholder="Manual notebook_id for debug only">
      <button class="button" type="submit">Preview Manual ID</button>
    </form>
  </section>
  </section>
  <section id="tab-governance" class="tab-section">
  <section class="status-panel {'warn' if not allow_network_access else ''}">
    <h2>Network Access</h2>
    <p>server 預設綁 <code>0.0.0.0</code>；這個開關決定是否接受同網段非本機連線。localhost 永遠保留。</p>
    <div class="grid">
      <div class="panel"><strong>Same-network Access</strong><br>{'ON' if allow_network_access else 'OFF'}</div>
      <div class="panel"><strong>Local URL</strong><br><code>http://127.0.0.1:8765</code></div>
      <div class="panel"><strong>LAN URL</strong><br><code>http://{escape_html(lan_ip)}:8765</code></div>
    </div>
    <div class="formrow">
      <a class="button" href="/api/network-access?enabled=yes">Enable Same-network Access</a>
      <a class="button danger" href="/api/network-access?enabled=no">Disable Same-network Access</a>
    </div>
  </section>
  <section id="closed-loop">
    <h2>Closed Loop Controls</h2>
    <p>本版先只做「清除」與「匯出」。匯入與修正屬於高風險全覆蓋操作，等資料模型穩定後再開，避免教學或 Win365 部署時誤覆蓋。</p>
    <h3>Local Logs</h3>
    <div class="formrow">
      <a class="button secondary" href="/api/export-local-logs-json">Export Local Logs JSON</a>
      <a class="button secondary" href="/api/export-local-logs-excel">Export Local Logs Excel</a>
      <a class="button danger" href="/api/clear-local-logs" onclick="return confirm('Clear local runtime logs only? Tasks and evidence stay.')">Clear Local Logs</a>
    </div>
    <h3>Local Tasks</h3>
    <div class="formrow">
      <a class="button secondary" href="/api/export-local-tasks-json">Export Local Tasks JSON</a>
      <a class="button secondary" href="/api/export-local-tasks-excel">Export Local Tasks Excel</a>
      <a class="button danger" href="/api/clear-local-tasks" onclick="return confirm('Clear local command packages and ETL queue? Logs, config, auth, downloads, and evidence stay.')">Clear Local Tasks</a>
    </div>
    <h3>Reset</h3>
    <a class="button danger" href="/api/clear?scope=temp" onclick="return confirm('Reset current run? This clears temp CSV and queue, but preserves inbox, config, and auth.')">Reset Current Run</a>
    <a class="button danger" href="/api/clear?scope=logs" onclick="return confirm('Clear runtime logs?')">Clear Logs</a>
    <a class="button danger" href="#" onclick="clearBrowserCache(); return false;">Clear Browser Memory</a>
    <details>
      <summary>Advanced clear actions</summary>
      <a class="button danger" href="/api/clear?scope=queue" onclick="return confirm('Clear queue manifest only?')">Clear Queue Only</a>
      <a class="button danger" href="/api/clear?scope=working" onclick="return confirm('Clear temp files and runtime logs? Inbox, config, and NotebookLM auth will be preserved.')">Reset Runtime Workspace</a>
    </details>
  </section>
  </section>
  <section id="tab-command-queue" class="tab-section">
    <h2>Command Queue</h2>
    <p>本機指令包模式：先把 JSON 放進 inbox，Runtime 驗證後排隊，再用手動或 auto mode 依序執行。這是 API / GAS / 遠端主機之前的教學與 dev 底座。</p>
    <div class="grid">
      <div class="panel"><strong>Inbox</strong><br>{command_counts.get('inbox', 0)} packages</div>
      <div class="panel"><strong>Queued</strong><br>{command_counts.get('queued', 0)} packages</div>
      <div class="panel"><strong>Auto Mode</strong><br>{auto_run_label}</div>
    </div>
    <div class="formrow">
      <a class="button gold" href="/api/command-queue/sample">Create Sample Package</a>
      <a class="button" href="/api/command-queue/queue">Validate & Queue Inbox</a>
      <a class="button secondary" href="/api/command-queue/execute?limit=1">Execute Next</a>
      <a class="button secondary" href="/api/command-queue/execute?limit=10">Execute Up To 10</a>
      <a class="button secondary" href="/api/command-queue/tick?limit=1">Auto Tick Once</a>
      <a class="button secondary" href="/api/gas-poll?execute=no">GAS Pull Only</a>
      <a class="button gold" href="/api/gas-poll?execute=yes">GAS Pull & Upload</a>
      <a class="button" href="/api/command-queue">Open Queue Manager</a>
    </div>
    <div class="formrow">
      <a class="button {'danger' if command_queue.get('auto_run') else ''}" href="/api/command-queue/auto?enabled={'no' if command_queue.get('auto_run') else 'yes'}">Turn Auto Mode {'OFF' if command_queue.get('auto_run') else 'ON'}</a>
    </div>
    <table>
      <thead><tr><th>Stage</th><th>Folder</th></tr></thead>
      <tbody>
        <tr><td>Inbox</td><td><code>{escape_html(command_stage_dir(config, 'inbox'))}</code></td></tr>
        <tr><td>Queued</td><td><code>{escape_html(command_stage_dir(config, 'queued'))}</code></td></tr>
        <tr><td>Completed / Failed</td><td><code>{escape_html(command_stage_dir(config, 'completed'))}</code> / <code>{escape_html(command_stage_dir(config, 'failed'))}</code></td></tr>
      </tbody>
    </table>
  </section>
  <section id="tab-incoming-watch" class="tab-section">
    <h2>Incoming Folder Watcher</h2>
    <p>監督本機資料夾：可用定時掃描，也可用系統事件 realtime watch。兩種地端模式都會先複製成 watch batch、建立 command package，再依設定排隊或執行。</p>
    <div class="grid">
      <div class="panel"><strong>Watch</strong><br>{'ON' if incoming_watch_enabled else 'OFF'}</div>
      <div class="panel"><strong>Mode</strong><br>{escape_html(incoming_watch_mode)}</div>
      <div class="panel"><strong>Interval</strong><br>{escape_html(runtime_settings.get('incoming_watch_interval_seconds', DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS))} sec</div>
      <div class="panel"><strong>Realtime Engine</strong><br>{'READY' if realtime_engine.get('available') else 'UNAVAILABLE'}</div>
      <div class="panel"><strong>Auto Queue</strong><br>{'ON' if incoming_watch_auto_queue else 'OFF'}</div>
      <div class="panel"><strong>Auto Execute</strong><br>{'ON' if incoming_watch_auto_execute else 'OFF'}</div>
    </div>
    <form id="incomingWatchForm" class="subpanel" action="/api/incoming-watch-settings" method="get">
      <h3>Local Watch Settings</h3>
      <p class="muted">預設 600 秒掃一次；檔案需穩定至少 10 秒，避免複製到一半就被上傳。手動測試可按下方 Scan Once。</p>
      <p class="danger-text">紅字提醒：本機 watcher 會先複製一份到 watch batch 再建立 command package；若檔案尚未被複製完成就刪除或搬移，該次任務可能不會建立。已進入 completed 後再整理來源資料夾最安全。</p>
      <label class="inline"><input type="checkbox" name="incoming_watch_enabled" value="yes" {'checked' if incoming_watch_enabled else ''}>啟用 incoming watcher</label>
      <label>模式
        <select name="incoming_watch_mode">
          <option value="polling" {'selected' if incoming_watch_mode == 'polling' else ''}>Polling Mode - 定時掃描</option>
          <option value="realtime" {'selected' if incoming_watch_mode == 'realtime' else ''}>Realtime Watch Mode - 系統事件偵測</option>
        </select>
      </label>
      <label class="inline"><input type="checkbox" name="incoming_watch_auto_queue" value="yes" {'checked' if incoming_watch_auto_queue else ''}>新檔自動排隊</label>
      <label class="inline"><input type="checkbox" name="incoming_watch_auto_execute" value="yes" {'checked' if incoming_watch_auto_execute else ''}>新檔自動執行</label>
      <label class="inline"><input type="checkbox" name="incoming_watch_recursive" value="yes" {'checked' if incoming_watch_recursive else ''}>包含子資料夾</label>
      <div class="formrow">
        <label>秒數<br><input type="number" name="incoming_watch_interval_seconds" min="{MIN_INCOMING_WATCH_INTERVAL_SECONDS}" value="{escape_html(runtime_settings.get('incoming_watch_interval_seconds', DEFAULT_INCOMING_WATCH_INTERVAL_SECONDS))}"></label>
        <label>檔案穩定秒數<br><input type="number" name="incoming_watch_min_age_seconds" min="0" value="{escape_html(runtime_settings.get('incoming_watch_min_age_seconds', 10))}"></label>
        <label>每次最多新檔<br><input type="number" name="incoming_watch_max_files_per_scan" min="1" max="100" value="{escape_html(runtime_settings.get('incoming_watch_max_files_per_scan', 10))}"></label>
      </div>
      <label>Incoming folder<br><input type="text" name="incoming_watch_folder" value="{escape_html(incoming_watch_folder)}"></label>
      <p><strong>File types</strong></p>
      <div class="formrow">
        <button class="button secondary" type="button" onclick="toggleTypes('incomingWatchForm', true)">Select All</button>
        <button class="button secondary" type="button" onclick="toggleTypes('incomingWatchForm', false)">Clear All</button>
      </div>
      <div class="type-grid">
        {''.join(f'<label><input type="checkbox" name="types" value="{ext}" {"checked" if ext in incoming_watch_types else ""}> {ext}</label>' for ext in DEFAULT_SIMPLE_TYPES)}
      </div>
      <button class="button" type="submit">Save Watch Settings</button>
    </form>
    <div class="formrow">
      <a class="button" href="/api/incoming-watch">Scan Once</a>
      <a class="button secondary" href="/api/incoming-watch?execute=yes">Scan Once + Execute</a>
      <a class="button secondary" href="/api/command-queue">Open Command Queue</a>
      <a class="button secondary" href="/api/logs?event=incoming_watch">Open Watch Logs</a>
    </div>
    <table>
      <tbody>
        <tr><th>Incoming folder</th><td><code>{escape_html(incoming_watch_folder)}</code></td></tr>
        <tr><th>Watch batches</th><td><code>{escape_html(simple_watch_batches_dir(config))}</code></td></tr>
        <tr><th>Watch state</th><td><code>{escape_html(incoming_watch_state_path(config))}</code></td></tr>
        <tr><th>Realtime engine</th><td><code>{escape_html(json.dumps(realtime_engine, ensure_ascii=False))}</code></td></tr>
      </tbody>
    </table>
  </section>
  <section id="tab-gas-cloud" class="tab-section">
    <h2>GAS Cloud Worker</h2>
    <p>這裡整合 Google Sheet / GAS 雲端任務中控。本機不開 ngrok，而是由 Python runtime 主動去雲端領任務。</p>
    <div class="grid">
      <div class="panel"><strong>GAS Enabled</strong><br>{'ON' if gas_enabled else 'OFF'}</div>
      <div class="panel"><strong>Auto Poll</strong><br>{'ON' if gas_auto_poll else 'OFF'}</div>
      <div class="panel"><strong>Auto Execute</strong><br>{'ON' if gas_auto_execute else 'OFF'}</div>
      <div class="panel"><strong>Poll Seconds</strong><br>{escape_html(gas_settings.get('poll_interval_seconds', DEFAULT_GAS_POLL_INTERVAL_SECONDS))}</div>
      <div class="panel"><strong>Token</strong><br>{escape_html(gas_token_label)}</div>
    </div>
    <form class="subpanel" action="/api/gas-settings" method="get">
      <h3>Local Polling Settings</h3>
      <p class="muted">Falo x Force 教學設計：雲端只放任務，本機依秒數主動取件。預設 600 秒，避免 NotebookLM / GAS / Drive 太密集時互相干擾；教學測試時可手動調短。</p>
      <p class="danger-text">紅字提醒：雲端 incoming 檔案在 task 尚未 completed 前，請不要刪除。系統用 Google Drive file_id 防重複；但本機 worker 仍需要原始檔案下載，上傳前刪除會造成 queued / queued_local 任務失敗。</p>
      <p class="muted">三個開關拆開：先能連線，再決定要不要自動檢查，最後才決定要不要自動執行。</p>
      <div class="formrow">
        {render_toggle_button('啟用 GAS 連線', 'enabled', gas_enabled)}
        {render_toggle_button('每 N 秒自動檢查', 'auto_poll_enabled', gas_auto_poll)}
        {render_toggle_button('抓到任務後自動執行', 'auto_execute', gas_auto_execute)}
      </div>
      <div class="formrow">
        <label>秒數<br><input type="number" name="poll_interval_seconds" min="{MIN_GAS_POLL_INTERVAL_SECONDS}" value="{escape_html(gas_settings.get('poll_interval_seconds', DEFAULT_GAS_POLL_INTERVAL_SECONDS))}"></label>
        <label>每次最多任務<br><input type="number" name="max_tasks_per_poll" min="1" max="50" value="{escape_html(gas_settings.get('max_tasks_per_poll', DEFAULT_GAS_MAX_TASKS_PER_POLL))}"></label>
      </div>
      <div class="formrow">
        <label>GAS Web App URL<br><input type="text" name="web_app_url" value="{escape_html(gas_url)}" placeholder="https://script.google.com/macros/s/AKfycbw9X3Y6MQ2XpvsS9BXuCZeZsVkrbT1VL0JkDkotrbs-omYG8OpuWpAl1fowiJa_QW1i/exec"></label>
      </div>
      <div class="formrow">
        <label>API Token<br><input id="gas_api_token" type="password" name="api_token" placeholder="留空代表沿用目前 token"></label>
        <button class="button gold" type="button" onclick="document.getElementById('gas_api_token').value='123456'; faloAction(this, 'Filled'); return false;">Use Default Token</button>
        <label>下載暫存資料夾<br><input type="text" name="local_download_root" value="{escape_html(gas_download_root)}"></label>
      </div>
      <div class="formrow">
        <a class="button gold" href="/api/gas-settings/default" data-running-label="Applying...">Apply Safe Default: 600s / 3 tasks</a>
        <button class="button" type="submit">Save GAS Settings</button>
      </div>
      <p class="muted">Safe Default 是 Falo 建議的治理預設；Save GAS Settings 則保存目前表單自訂值。</p>
      <p class="muted">治理預設只會開啟 GAS 連線；自動檢查與自動執行預設關閉，避免 demo 或教學時一按就連續批次執行。</p>
    </form>
    <table>
      <tbody>
        <tr><th>Web App URL</th><td><code>{escape_html(gas_url or '(not set)')}</code></td></tr>
        <tr><th>Local download root</th><td><code>{escape_html(gas_download_root)}</code></td></tr>
        <tr><th>Max tasks per poll</th><td><code>{escape_html(gas_settings.get('max_tasks_per_poll', DEFAULT_GAS_MAX_TASKS_PER_POLL))}</code></td></tr>
        <tr><th>Auto execute</th><td><code>{escape_html(gas_settings.get('auto_execute', False))}</code></td></tr>
      </tbody>
    </table>
    <div class="formrow">
      <a class="button gold" href="/api/gas-test">Test GAS Token Connection</a>
      <a class="button secondary" href="/api/gas-poll?execute=no">Pull from GAS Only</a>
      <a class="button gold" href="/api/gas-poll?execute=yes">Pull & Upload to NotebookLM</a>
      <a class="button secondary" href="/api/command-queue">Open Local Queue</a>
      <a class="button secondary" href="/api/logs?event=gas">Open GAS Logs</a>
      {f'<a class="button gold" href="{escape_html(gas_url)}" target="_blank">Open GAS Web App</a>' if gas_url else ''}
    <p class="muted">操作順序：先在 GAS Web App 產生 timestamp .md task，再回這裡按 Poll GAS Once。成功後會下載雲端檔案並轉成本機 command package。</p>
  </section>
  <section id="tab-meeting-converter" class="tab-section">
    <h2>Meeting Minutes Converter</h2>
    <p>將 NotebookLM 產生的結構化 Markdown 會議記錄轉換為集團標準格式的 Word 檔及 HTML。</p>
    
    <div class="upload-card">
      <form id="meetingConverterForm" onsubmit="event.preventDefault(); convertMeeting();">
        <label><strong>Pasted Markdown Content</strong></label><br>
        <textarea id="markdown_content" style="width: 100%; height: 350px; padding: 12px; border: 1px solid #cfd8d3; border-radius: 6px; font-family: monospace;" placeholder="在此貼上 NotebookLM 整理的會議記錄 Markdown (包含 ## MEETING_INFO, ## DISCUSSION 等區段)..."></textarea>
        <br><br>
        <button class="button" type="submit" id="btn_convert">Convert to Document</button>
      </form>
    </div>

    <div id="converter_result_panel" class="upload-card" style="display: none; margin-top: 18px;">
      <h3>Conversion Result</h3>
      <div id="converter_message" style="margin-bottom: 14px; padding: 10px; border-radius: 6px;"></div>
      <div class="formrow" id="download_links">
        <!-- Download links will be inserted here dynamically -->
      </div>
      <br>
      <h4>Parsing Summary</h4>
      <pre id="converter_summary" style="background: #f4f6f4; padding: 12px; border-radius: 6px; overflow-x: auto; font-family: monospace; white-space: pre-wrap;"></pre>
    </div>
  </section>
  <section id="tab-multi-chat" class="tab-section">
    <h2>Multi-User Chat Gateway</h2>
    <p>透過隔離的對話 ID，讓不同同仁在同一個筆記本下進行獨立的問答，互不干涉且零 Token 費用。</p>
    
    <div class="grid-2">
      <div class="upload-card">
        <form id="multiChatForm" onsubmit="event.preventDefault(); sendChatMessage();">
          <label><strong>Target Notebook ID</strong></label><br>
          <input type="text" id="chat_notebook_id" style="width:100%; padding:8px; border:1px solid #cfd8d3; border-radius:6px;" value="{escape_html(active_notebook_id or '73085c64-dea5-4945-9226-949023b0ac9b')}">
          <br><br>
          
          <label><strong>User Name (Submitter)</strong></label><br>
          <input type="text" id="chat_user_name" style="width:100%; padding:8px; border:1px solid #cfd8d3; border-radius:6px;" placeholder="請輸入您的姓名/代號 (如: PM-A)" value="PM-A">
          <br><br>
          
          <label><strong>Conversation Session</strong></label><br>
          <select id="chat_session_select" style="width:100%; padding:8px; border:1px solid #cfd8d3; border-radius:6px;" onchange="onSessionSelectChange();">
            <option value="new">-- 開啟全新對話 (Create New Chat) --</option>
          </select>
          <br><br>
          
          <label><strong>Enter Question</strong></label><br>
          <textarea id="chat_question" style="width: 100%; height: 120px; padding: 12px; border: 1px solid #cfd8d3; border-radius: 6px; font-family: inherit;" placeholder="請輸入您想對 NotebookLM 提問的問題..."></textarea>
          <br><br>
          
          <button class="button" type="submit" id="btn_send_chat">Send Question</button>
        </form>
      </div>

      <div class="upload-card" style="display:flex; flex-direction:column; max-height:480px;">
        <h3>Chat Box</h3>
        <div id="chat_box_conversation" style="flex-grow:1; overflow-y:auto; border:1px solid #cfd8d3; border-radius:6px; background:#f9faf9; padding:12px; margin-bottom:12px; min-height:300px;">
          <div style="color:#66706b; text-align:center; margin-top:100px;">請選擇對話或輸入問題開始問答</div>
        </div>
      </div>
    </div>

    <div class="upload-card" style="margin-top:18px;">
      <h3>Admin Session Audit (管理者審計面板)</h3>
      <p class="muted">地端儲存的所有隔離對話清單。管理員可以直接點擊切換檢視歷史，或刪除對話。</p>
      <div style="overflow-x:auto;">
        <table style="width:100%; border-collapse:collapse; background:white;">
          <thead>
            <tr style="background:#e8f1ed; color:#174f43;">
              <th style="padding:10px; border:1px solid #d8ddd8; text-align:left;">使用者姓名</th>
              <th style="padding:10px; border:1px solid #d8ddd8; text-align:left;">對話 ID (Session ID)</th>
              <th style="padding:10px; border:1px solid #d8ddd8; text-align:left;">最後發問內容</th>
              <th style="padding:10px; border:1px solid #d8ddd8; text-align:left;">最後發問時間</th>
              <th style="padding:10px; border:1px solid #d8ddd8; text-align:left;">操作</th>
            </tr>
          </thead>
          <tbody id="admin_sessions_table_body">
            <tr><td colspan="5" style="text-align:center; padding:20px; color:#66706b;">載入中...</td></tr>
          </tbody>
        </table>
      </div>
    </div>
  </section>
</main>
<div class="watermark">{APP_WATERMARK}</div>
</body>
</html>"""


def render_clear_result(result: Dict[str, object]) -> str:
    removed = "".join(f"<li><code>{escape_html(item)}</code></li>" for item in result.get("removed", [])) or "<li>無</li>"
    error = result.get("error", "")
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Clear Result</title>{simple_style()}</head>
<body><main><h1>Clear Result</h1>
<p>Scope: <code>{escape_html(result.get('scope', ''))}</code> · removed: {result.get('removed_count', 0)}</p>
<p>{escape_html(error)}</p>
<ul>{removed}</ul>
<a class="button" href="/">Back to Portal</a>
</main></body></html>"""


def render_import_result(title: str, result: Dict[str, object]) -> str:
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>{escape_html(title)}</title>{simple_style()}</head>
<body><main><h1>{escape_html(title)}</h1>
<p>Imported tasks: {result.get('imported_count', result.get('task_count', 0))}</p>
<p>Source: <code>{escape_html(result.get('imported_from', ''))}</code></p>
<p>Queue path: <code>{escape_html(PROJECT_ROOT / 'data' / 'temp' / 'etl_queue.json')}</code></p>
<a class="button" href="/">Back to Portal</a>
</main></body></html>"""


def render_simple_upload_result(result: Dict[str, object]) -> str:
    rows = []
    for item in result.get("results", []):
        status = item.get("status", "")
        row_class = "ok-row" if status == "uploaded" else "warn-row" if status == "skipped" else "error-row"
        source = item.get("source", {})
        if not isinstance(source, dict):
            source = {}
        source_id = source.get("id") or source.get("source_id") or ""
        rows.append(
            f"""<tr class="{row_class}">
            <td>{escape_html(item.get('filename', ''))}</td>
            <td>{escape_html(item.get('target_name', ''))}</td>
            <td>{escape_html(item.get('action', ''))}</td>
            <td>{escape_html(status)}</td>
            <td><code>{escape_html(source_id)}</code></td>
            <td>{escape_html(item.get('message', ''))}</td>
            </tr>"""
        )
    body = "".join(rows) or "<tr><td colspan='6'>No files processed.</td></tr>"
    if result.get("error"):
        box_class = "result-error"
        title = "Upload Blocked"
    elif result.get("failed_count", 0):
        box_class = "result-error"
        title = "Upload Finished With Errors"
    else:
        box_class = "result-ok"
        title = "Upload Complete"
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Simple Upload Result</title>{simple_style()}
<style>
  .ok-row {{ background: #eaf8ef; }}
  .warn-row {{ background: #fff8e6; }}
  .error-row {{ background: #fdecec; }}
</style></head>
<body><main>
<div class="{box_class}">
  <h1>{title}</h1>
  <p>{escape_html(result.get('error', ''))}</p>
  <p>Project: <strong>{escape_html(result.get('project', {}).get('name', '')) if isinstance(result.get('project', {}), dict) else ''}</strong></p>
  <p>Notebook ID: <code>{escape_html(result.get('notebook_id', ''))}</code></p>
  <p>Uploaded: {result.get('uploaded_count', 0)} · Skipped: {result.get('skipped_count', 0)} · Failed: {result.get('failed_count', 0)}</p>
</div>
<table><thead><tr><th>Original File</th><th>Uploaded Name</th><th>Action</th><th>Status</th><th>Source ID</th><th>Message</th></tr></thead><tbody>{body}</tbody></table>
<a class="button" href="/">Back to Portal</a>
<a class="button secondary" href="/api/logs?event=simple_upload">Open Upload Log</a>
</main></body></html>"""


def render_create_or_use_result(result: Dict[str, object]) -> str:
    project = result.get("project", {})
    if not isinstance(project, dict):
        project = {}
    box = "result-ok" if result.get("ok") else "result-error"
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Create & Use Notebook</title>{simple_style()}</head>
<body><main><div class="{box}">
<h1>Create & Use Notebook</h1>
<p>{escape_html(result.get('message', result.get('error', '')))}</p>
<p>Project: <strong>{escape_html(project.get('name', ''))}</strong></p>
<p>Notebook ID: <code>{escape_html(project.get('notebook_id', ''))}</code></p>
</div>
<a class="button" href="/">Back to Simple Upload</a>
<a class="button secondary" href="/api/projects">Project Manager</a>
</main></body></html>"""


def render_folder_scan_preview(scan: Dict[str, object], params: Dict[str, list]) -> str:
    rows = []
    for index, item in enumerate(scan.get("files", []), 1):
        rows.append(
            f"""<tr>
            <td>{index}</td>
            <td>{escape_html(item.get('name', ''))}</td>
            <td>{escape_html(item.get('suffix', ''))}</td>
            <td>{item.get('size_kb', '')}</td>
            <td><code>{escape_html(item.get('path', ''))}</code></td>
            </tr>"""
        )
    body = "".join(rows) or "<tr><td colspan='5'>No files matched.</td></tr>"
    query_parts = []
    for key, values in params.items():
        for value in values:
            query_parts.append(f"{key}={quote_url(value)}")
    query_parts.append("confirm=yes")
    upload_url = "/api/folder-upload?" + "&".join(query_parts)
    box = "result-ok" if scan.get("ok") else "result-error"
    error = f"<p>{escape_html(scan.get('error', ''))}</p>" if scan.get("error") else ""
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Folder Scan Preview</title>{simple_style()}</head>
<body><main><div class="{box}">
<h1>Folder Scan Preview</h1>
{error}
<p>Folder: <code>{escape_html(scan.get('folder', ''))}</code></p>
<p>Matched files: {scan.get('count', 0)}</p>
</div>
<table><thead><tr><th>#</th><th>Name</th><th>Type</th><th>Size KB</th><th>Path</th></tr></thead><tbody>{body}</tbody></table>
<a class="button" href="{upload_url}">Confirm Upload Folder</a>
<a class="button secondary" href="/">Back to Portal</a>
</main></body></html>"""


def render_notebook_list(result: Dict[str, object]) -> str:
    rows = []
    for item in result.get("notebooks", []):
        rows.append(
            "<tr>"
            f"<td>{escape_html(item.get('title', ''))}</td>"
            f"<td><code>{escape_html(item.get('id', ''))}</code></td>"
            f"<td>{escape_html(str(item.get('is_owner', '')))}</td>"
            f"<td>{escape_html(item.get('created_at', ''))}</td>"
            "</tr>"
        )
    body = "".join(rows) or "<tr><td colspan='4'>No notebooks returned.</td></tr>"
    error = ""
    if not result.get("ok"):
        error = f"<p class='danger-text'>{escape_html(result.get('error', ''))}</p>"
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Notebook List</title>{simple_style()}</head>
<body><main><h1>Notebook List</h1>
<p>Count: {result.get('count', 0)}</p>
{error}
<table><thead><tr><th>Title</th><th>ID</th><th>Owner</th><th>Created At</th></tr></thead><tbody>{body}</tbody></table>
<a class="button" href="/">Back to Portal</a>
</main></body></html>"""


def render_notebook_create_result(result: Dict[str, object]) -> str:
    notebook = result.get("notebook", {})
    if not isinstance(notebook, dict):
        notebook = {}
    error = ""
    if not result.get("ok"):
        error = f"<p class='danger-text'>{escape_html(result.get('error', ''))}</p>"
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Create Notebook</title>{simple_style()}</head>
<body><main><h1>Create Notebook</h1>
<p>Status: <strong>{'OK' if result.get('ok') else 'FAILED'}</strong></p>
{error}
<table>
<tr><th>Title</th><td>{escape_html(notebook.get('title', result.get('title', '')))}</td></tr>
<tr><th>ID</th><td><code>{escape_html(notebook.get('id', ''))}</code></td></tr>
<tr><th>Created At</th><td>{escape_html(notebook.get('created_at', ''))}</td></tr>
</table>
<a class="button" href="/">Back to Portal</a>
<a class="button" href="/api/notebooks">List Notebooks</a>
</main></body></html>"""


def render_project_manager(result: Dict[str, object]) -> str:
    projects_json = json.dumps(result.get("projects", []), ensure_ascii=False).replace("</", "<\\/")
    active = get_active_project(load_or_create_config(PROJECT_ROOT))
    active_project_id = active.get("active_project_id", "")
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Project Manager</title>{simple_style()}
<style>
  .toolbar {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; margin: 16px 0; background: #ffffff; border: 1px solid #d8ddd8; border-radius: 8px; padding: 12px; }}
  .toolbar input {{ min-width: 300px; }}
  .view-tabs button {{ background: #e8f1ed; color: #174f43; }}
  .view-tabs button.active {{ background: #1f6f5b; color: white; }}
  tr.active-row, .card.active-card, .compact-row.active-row {{ background: #e4f6ec !important; outline: 2px solid #208454; }}
  th.sortable {{ cursor: pointer; user-select: none; }}
  th.sortable:hover {{ background: #d8ebe3; }}
  .cards {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(260px, 1fr)); gap: 12px; margin-top: 16px; }}
  .card {{ background: #fff; border: 1px solid #d8ddd8; border-radius: 8px; padding: 14px; }}
  .card:hover, tr:hover, .compact-row:hover {{ background: #f0f7f4; }}
  .compact-row {{ display: flex; justify-content: space-between; gap: 12px; background: white; border: 1px solid #d8ddd8; padding: 8px 10px; }}
  .notice {{ background: #edf7f1; border-left: 6px solid #208454; padding: 12px; border-radius: 8px; }}
  .flash {{ transition: background .25s ease; }}
  .flash.clicked {{ background: #ccebd9 !important; }}
</style></head>
<body><main><h1>Project Manager</h1>
<p class="notice">NotebookLM 專案管理：即時搜尋、點欄位排序、可切換呈現方式。點 <strong>Set Active</strong> 後，Portal 前面的 Active Project 會同步更新。</p>
<div class="toolbar">
  <input id="searchBox" type="text" placeholder="Instant search: title / id / tag">
  <label>每頁 <select id="pageSize"><option>10</option><option selected>25</option><option>50</option><option>100</option></select></label>
  <span class="view-tabs">
    <button class="button active" type="button" data-view="table">Table</button>
    <button class="button" type="button" data-view="card">Cards</button>
    <button class="button" type="button" data-view="compact">Compact</button>
  </span>
  <a class="button gold" href="/api/projects/sync">Sync Notebooks</a>
  <a class="button secondary" href="/">Portal</a>
</div>
<p id="summary"></p>
<div id="projectMount"></div>
<div class="formrow">
  <button class="button secondary" type="button" onclick="changePage(-1)">Prev</button>
  <button class="button secondary" type="button" onclick="changePage(1)">Next</button>
</div>
<script>
const projects = {projects_json};
const activeProjectId = {json.dumps(active_project_id, ensure_ascii=False)};
let view = 'table';
let sortKey = 'name';
let sortDir = 1;
let page = 1;

function textOf(project) {{
  return JSON.stringify(project).toLowerCase();
}}
function filtered() {{
  const q = document.getElementById('searchBox').value.trim().toLowerCase();
  let rows = projects.filter(p => !q || textOf(p).includes(q));
  rows.sort((a, b) => String(a[sortKey] || '').localeCompare(String(b[sortKey] || ''), 'zh-Hant') * sortDir);
  return rows;
}}
function flash(el) {{
  el.classList.add('clicked');
  setTimeout(() => el.classList.remove('clicked'), 650);
}}
function setActive(projectId, el) {{
  flash(el);
  el.textContent = 'Setting...';
  window.location.href = '/api/projects/active?project_id=' + encodeURIComponent(projectId);
}}
function preview(projectId, el) {{
  flash(el);
  window.location.href = '/api/adapter-preview?project_id=' + encodeURIComponent(projectId);
}}
function actions(project) {{
  const id = project.project_id || '';
  return `<button class="button flash" onclick="setActive('${{id}}', this)">Set Active</button> <button class="button secondary flash" onclick="preview('${{id}}', this)">Preview</button>`;
}}
function rowClass(project) {{
  return project.project_id === activeProjectId ? 'active-row' : '';
}}
function renderTable(rows) {{
  const body = rows.map(project => `<tr class="${{rowClass(project)}}"><td><strong>${{escapeJs(project.name)}}</strong><br><span class="muted">${{escapeJs(project.project_id)}}</span></td><td><code>${{escapeJs(project.notebook_id)}}</code></td><td>${{escapeJs(project.status)}}</td><td>${{escapeJs(project.is_owner)}}</td><td>${{escapeJs(project.notebook_created_at)}}</td><td>${{actions(project)}}</td></tr>`).join('');
  return `<table><thead><tr><th class="sortable" onclick="sortBy('name')">Project</th><th>Notebook ID</th><th class="sortable" onclick="sortBy('status')">Status</th><th class="sortable" onclick="sortBy('is_owner')">Owner</th><th class="sortable" onclick="sortBy('notebook_created_at')">Created</th><th>Actions</th></tr></thead><tbody>${{body || '<tr><td colspan="6">No projects matched.</td></tr>'}}</tbody></table>`;
}}
function renderCards(rows) {{
  return `<div class="cards">${{rows.map(project => `<div class="card ${{project.project_id === activeProjectId ? 'active-card' : ''}}"><strong>${{escapeJs(project.name)}}</strong><p class="muted"><code>${{escapeJs(project.notebook_id)}}</code></p><p>Status: ${{escapeJs(project.status)}} · Owner: ${{escapeJs(project.is_owner)}}</p><p>${{escapeJs(project.notebook_created_at)}}</p>${{actions(project)}}</div>`).join('') || '<p>No projects matched.</p>'}}</div>`;
}}
function renderCompact(rows) {{
  return rows.map(project => `<div class="compact-row ${{rowClass(project)}}"><span><strong>${{escapeJs(project.name)}}</strong> <span class="muted">${{escapeJs(project.notebook_created_at)}}</span></span><span>${{actions(project)}}</span></div>`).join('') || '<p>No projects matched.</p>';
}}
function escapeJs(value) {{
  return String(value ?? '').replaceAll('&','&amp;').replaceAll('<','&lt;').replaceAll('>','&gt;').replaceAll('"','&quot;');
}}
function render() {{
  const size = Number(document.getElementById('pageSize').value);
  const rows = filtered();
  const totalPages = Math.max(1, Math.ceil(rows.length / size));
  page = Math.min(page, totalPages);
  const shown = rows.slice((page - 1) * size, page * size);
  document.getElementById('summary').textContent = `Total ${{rows.length}} · Page ${{page}} / ${{totalPages}} · Sort ${{sortKey}} ${{sortDir > 0 ? '↑' : '↓'}}`;
  document.getElementById('projectMount').innerHTML = view === 'card' ? renderCards(shown) : view === 'compact' ? renderCompact(shown) : renderTable(shown);
}}
function sortBy(key) {{
  if (sortKey === key) sortDir *= -1;
  else {{ sortKey = key; sortDir = 1; }}
  render();
}}
function changePage(delta) {{
  page = Math.max(1, page + delta);
  render();
}}
document.getElementById('searchBox').addEventListener('input', () => {{ page = 1; render(); }});
document.getElementById('pageSize').addEventListener('change', () => {{ page = 1; render(); }});
document.querySelectorAll('.view-tabs button').forEach(btn => btn.addEventListener('click', () => {{
  document.querySelectorAll('.view-tabs button').forEach(item => item.classList.remove('active'));
  btn.classList.add('active');
  view = btn.dataset.view;
  render();
}}));
render();
</script>
</main></body></html>"""


def render_project_sync_result(result: Dict[str, object]) -> str:
    error = f"<p class='danger-text'>{escape_html(result.get('error', ''))}</p>" if result.get("error") else ""
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Project Sync</title>{simple_style()}</head>
<body><main><h1>Project Sync</h1>
{error}
<p>Projects: {result.get('project_count', 0)} · Added: {result.get('added_count', 0)}</p>
<a class="button" href="/api/projects">Open Project Manager</a>
<a class="button" href="/">Back to Portal</a>
</main></body></html>"""


def render_active_project_result(result: Dict[str, object]) -> str:
    project = result.get("project", {})
    if not isinstance(project, dict):
        project = {}
    status_class = "result-ok" if result.get("ok") else "result-warn"
    title = "Active Project Updated" if result.get("ok") else "Active Project Not Updated"
    error = f"<p>{escape_html(result.get('error', ''))}</p>" if result.get("error") else ""
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>{title}</title>{simple_style()}</head>
<body><main><div class="{status_class}"><h1>{title}</h1>
{error}
<p><strong>{escape_html(project.get('name', ''))}</strong></p>
<p>project_id: <code>{escape_html(result.get('active_project_id', ''))}</code></p>
<p>notebook_id: <code>{escape_html(result.get('active_notebook_id', ''))}</code></p>
</div>
<a class="button" href="/">Back to Portal</a>
<a class="button secondary" href="/api/projects">Project Manager</a>
<a class="button" href="/api/adapter-preview">Preview Upload</a>
</main></body></html>"""


def render_log_cms(result: Dict[str, object]) -> str:
    rows = []
    for record in result.get("logs", []):
        rows.append(
            "<tr>"
            f"<td>{escape_html(record.get('ts', ''))}</td>"
            f"<td>{escape_html(record.get('event', ''))}</td>"
            f"<td><code>{escape_html(json.dumps(record.get('payload', {}), ensure_ascii=False))}</code></td>"
            f"<td>{escape_html(record.get('_line', ''))}</td>"
            "</tr>"
        )
    body = "".join(rows) or "<tr><td colspan='4'>No logs matched.</td></tr>"
    search = escape_html(result.get("search", ""))
    event = escape_html(result.get("event", ""))
    page = int(result.get("page", 1))
    next_page = page + 1
    prev_page = max(1, page - 1)
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Log CMS</title>{simple_style()}</head>
<body><main><h1>Log CMS</h1>
<p>Runtime log 的人看管理頁：搜尋、事件過濾、分頁。</p>
<form class="formrow" action="/api/logs" method="get">
  <input type="text" name="search" value="{search}" placeholder="Search payload / project / notebook">
  <input type="text" name="event" value="{event}" placeholder="Filter event">
  <button class="button" type="submit">Search</button>
  <a class="button" href="/">Portal</a>
</form>
<p>Total: {result.get('total', 0)} · Page: {page}</p>
<a class="button" href="/api/logs?search={search}&event={event}&page={prev_page}">Prev</a>
<a class="button" href="/api/logs?search={search}&event={event}&page={next_page}">Next</a>
<table><thead><tr><th>Time</th><th>Event</th><th>Payload</th><th>Line</th></tr></thead><tbody>{body}</tbody></table>
</main></body></html>"""


def render_command_queue(result: Dict[str, object], params: Dict[str, list]) -> str:
    stage_filter = params.get("stage", ["all"])[0]
    search = params.get("search", [""])[0].strip().lower()
    rows = []
    for item in result.get("commands", []):
        if stage_filter != "all" and str(item.get("stage", "")) != stage_filter:
            continue
        haystack = " ".join(
            [
                str(item.get("command_id", "")),
                str(item.get("cloud_task_id", "")),
                str(item.get("trigger_source", "")),
                str(item.get("trigger_mode", "")),
                str(item.get("cloud_event_type", "")),
                str(item.get("target_project_id", "")),
                " ".join(item.get("source_names", [])),
            ]
        ).lower()
        if search and search not in haystack:
            continue
        valid = bool(item.get("valid"))
        stage = str(item.get("stage", ""))
        badge = "ok" if valid and stage in {"queued", "completed"} else "warn" if valid else "error"
        action = ""
        if stage == "queued":
            action = f'<a class="button secondary" data-running-label="Executing..." href="/api/command-queue/execute?command_id={quote_url(item.get("command_id", ""))}">Execute</a>'
        if stage in {"queued", "completed", "failed"}:
            action += f' <a class="button gold" data-running-label="Archiving..." href="/api/command-queue/archive?command_id={quote_url(item.get("command_id", ""))}">Archive</a>'
        source_names = item.get("source_names", [])
        source_text = "<br>".join(escape_html(name) for name in source_names[:3]) or "<span class='muted'>No source metadata</span>"
        if len(source_names) > 3:
            source_text += f"<br><span class='muted'>+{len(source_names) - 3} more</span>"
        source_meta = (
            f"<br><span class='muted'>source: {escape_html(item.get('trigger_source', ''))}</span>"
            f"<br><span class='muted'>mode: {escape_html(item.get('trigger_mode', ''))}</span>"
            f"<br><code>{escape_html(item.get('cloud_event_type', ''))}</code>"
        )
        rows.append(
            "<tr>"
            f"<td><span class='badge {badge}'>{escape_html(stage)}</span></td>"
            f"<td><strong>{escape_html(item.get('command_id', ''))}</strong><br><span class='muted'>{escape_html(item.get('filename', ''))}</span></td>"
            f"<td>{source_text}{source_meta}<br><code>{escape_html(item.get('cloud_task_id', ''))}</code><br><span class='muted'>{escape_html(item.get('created_at', ''))}</span></td>"
            f"<td>{escape_html(item.get('command_type', ''))}</td>"
            f"<td>{escape_html(item.get('submitter', ''))}<br><span class='muted'>{escape_html(item.get('role', ''))}</span></td>"
            f"<td><code>{escape_html(item.get('target_project_id', ''))}</code></td>"
            f"<td>{escape_html('; '.join(item.get('errors', [])))}</td>"
            f"<td>{action}</td>"
            "</tr>"
        )
    body = "".join(rows) or "<tr><td colspan='8'>No command packages matched.</td></tr>"
    counts = result.get("counts", {})
    roles = result.get("users", {}).get("roles", {}) if isinstance(result.get("users", {}), dict) else {}
    role_rows = "".join(
        f"<tr><td>{escape_html(role)}</td><td>{escape_html(meta.get('label', ''))}</td><td>{escape_html(meta.get('description', ''))}</td><td><code>{escape_html(', '.join(meta.get('allowed_actions', [])))}</code></td></tr>"
        for role, meta in roles.items()
        if isinstance(meta, dict)
    )
    auto = bool(result.get("auto_run"))
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>Command Queue</title>{simple_style()}
<style>
  .badge {{ display: inline-block; min-width: 72px; text-align: center; padding: 3px 7px; border-radius: 999px; font-size: 12px; font-weight: 700; }}
  .badge.ok {{ background: #ddf5e8; color: #176c3a; }}
  .badge.warn {{ background: #fff0d5; color: #9a5b00; }}
  .badge.error {{ background: #fdecec; color: #8f2f2f; }}
  .summary-grid {{ display: grid; grid-template-columns: repeat(6, 1fr); gap: 8px; margin: 16px 0; }}
  .summary-box {{ background: white; border: 1px solid #d8ddd8; border-radius: 8px; padding: 12px; }}
  .filterbar {{ background: white; border: 1px solid #d8ddd8; border-radius: 8px; padding: 12px; margin: 12px 0; }}
  @media (max-width: 900px) {{ .summary-grid {{ grid-template-columns: repeat(2, 1fr); }} }}
</style></head>
<body><main><h1>Command Queue</h1>
<p>本機指令包 queue：適合教學與 dev。先驗證，再排隊，再執行；未來 API、GAS、遠端主機都可以把任務轉成同一種 JSON。</p>
{f'<div class="result-warn"><strong>Auto Mode ON</strong><br>此頁會用本機 tick 依序執行 queued 指令包。若要暫停，請先關閉 Auto Mode。</div>' if auto else ''}
<div class="summary-grid">
  {''.join(f"<div class='summary-box'><strong>{stage}</strong><br>{counts.get(stage, 0)}</div>" for stage in COMMAND_STAGES)}
</div>
<div class="formrow">
  <a class="button gold" data-running-label="Creating..." href="/api/command-queue/sample">Create Sample Package</a>
  <a class="button" data-running-label="Validating..." href="/api/command-queue/queue">Validate & Queue Inbox</a>
  <a class="button secondary" data-running-label="Executing..." href="/api/command-queue/execute?limit=1">Execute Next</a>
  <a class="button secondary" data-running-label="Executing..." href="/api/command-queue/execute?limit=10">Execute Up To 10</a>
  <a class="button secondary" data-running-label="Ticking..." href="/api/command-queue/tick?limit=1">Auto Tick Once</a>
  <a class="button {'gold' if auto else ''}" href="/api/command-queue/auto?enabled={'no' if auto else 'yes'}">Auto Mode: {'ON' if auto else 'OFF'}</a>
  <a class="button secondary" href="/">Portal</a>
</div>
<form class="filterbar" action="/api/command-queue" method="get">
  <label>Stage
    <select name="stage">
      {''.join(f'<option value="{stage}" {"selected" if stage_filter == stage else ""}>{stage}</option>' for stage in ["all"] + COMMAND_STAGES)}
    </select>
  </label>
  <label>Search
    <input type="text" name="search" value="{escape_html(search)}" placeholder="source filename / cloud task id / command id">
  </label>
  <button class="button" type="submit" data-running-label="Filtering...">Filter</button>
  <a class="button secondary" href="/api/command-queue?stage=queued">Queued Only</a>
</form>
<h2>Packages</h2>
<table><thead><tr><th>Stage</th><th>Command</th><th>Source / Cloud Task</th><th>Type</th><th>Submitter</th><th>Project</th><th>Validation</th><th>Action</th></tr></thead><tbody>{body}</tbody></table>
<h2>User Roles MVP</h2>
<table><thead><tr><th>Role</th><th>Label</th><th>Description</th><th>Allowed Actions</th></tr></thead><tbody>{role_rows}</tbody></table>
<h2>Folders</h2>
<table><tbody>
  <tr><th>Inbox</th><td><code>{escape_html(command_stage_dir(load_or_create_config(PROJECT_ROOT), 'inbox'))}</code></td></tr>
  <tr><th>Queued</th><td><code>{escape_html(command_stage_dir(load_or_create_config(PROJECT_ROOT), 'queued'))}</code></td></tr>
  <tr><th>Audit</th><td><code>{escape_html(command_audit_path(load_or_create_config(PROJECT_ROOT)))}</code></td></tr>
</tbody></table>
{f'<script>setTimeout(() => {{ window.location.href = "/api/command-queue/tick?limit=1"; }}, 5000);</script>' if auto and counts.get('queued', 0) else ''}
</main></body></html>"""


def render_command_result(title: str, result: Dict[str, object]) -> str:
    box = "result-ok" if result.get("ok", True) else "result-error"
    summary_items = [
        ("Mode", result.get("mode", "")),
        ("OK", result.get("ok", "")),
        ("Task Count", result.get("task_count", "")),
        ("Created Commands", len(result.get("created_commands", [])) if isinstance(result.get("created_commands", []), list) else ""),
        ("Executed Count", result.get("executed_count", "")),
        ("Queued", result.get("queued", {}).get("queued_count", "") if isinstance(result.get("queued", {}), dict) else ""),
        ("Failed", result.get("queued", {}).get("failed_count", "") if isinstance(result.get("queued", {}), dict) else ""),
    ]
    summary_rows = "".join(
        f"<tr><th>{escape_html(label)}</th><td>{escape_html(value)}</td></tr>"
        for label, value in summary_items
        if value != ""
    )
    rows = ""
    if result.get("results"):
        rows = "<h2>Results</h2><table><thead><tr><th>Command</th><th>Status</th><th>OK</th><th>Path / Error</th></tr></thead><tbody>"
        for item in result.get("results", []):
            rows += (
                "<tr>"
                f"<td>{escape_html(item.get('command_id', ''))}</td>"
                f"<td>{escape_html(item.get('status', ''))}</td>"
                f"<td>{escape_html(item.get('ok', ''))}</td>"
                f"<td><code>{escape_html(item.get('path', item.get('error', '')))}</code></td>"
                "</tr>"
            )
        rows += "</tbody></table>"
    return f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><title>{escape_html(title)}</title>{simple_style()}</head>
<body><main>
<div class="{box}">
  <h1>{escape_html(title)}</h1>
  <p>Falo x Force action result：先看摘要，詳細 JSON 保留在下方方便工程除錯。</p>
</div>
<table><tbody>{summary_rows or '<tr><td>No summary.</td></tr>'}</tbody></table>
{rows}
<details><summary>Raw JSON Detail</summary><pre>{escape_html(json.dumps(result, ensure_ascii=False, indent=2))}</pre></details>
<a class="button" href="/api/command-queue">Open Command Queue</a>
<a class="button secondary" href="/">Portal</a>
</main></body></html>"""


def simple_style() -> str:
    return """<style>
    body { margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Noto Sans TC", sans-serif; color: #202124; background: #f6f7f4; }
    main { max-width: 1080px; margin: 0 auto; padding: 32px 20px 56px; }
    h1 { margin: 0 0 8px; font-size: 32px; }
    table { width: 100%; border-collapse: collapse; background: white; margin-top: 16px; }
    th, td { border: 1px solid #d8ddd8; padding: 10px; vertical-align: top; text-align: left; }
    th { background: #e8f1ed; color: #174f43; }
    code { font-family: Menlo, Consolas, monospace; font-size: 13px; }
    input, select { padding: 8px; border: 1px solid #cfd8d3; border-radius: 6px; }
    .formrow { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; margin-top: 10px; }
    .muted { color: #667085; }
    .danger-text { color: #8f2f2f; }
    .result-ok { background: #e4f6ec; border-left: 7px solid #208454; padding: 18px; border-radius: 8px; }
    .result-warn { background: #fff8e6; border-left: 7px solid #b87900; padding: 18px; border-radius: 8px; }
    .result-error { background: #fdecec; border-left: 7px solid #9d2f2f; padding: 18px; border-radius: 8px; }
    .button { display: inline-block; margin-top: 10px; padding: 9px 12px; border: 0; border-radius: 6px; background: #1f6f5b; color: white; text-decoration: none; font: inherit; cursor: pointer; }
    .button.secondary { background: #315a7d; }
    .button.gold { background: #a46700; }
    .button.is-running { background: #a46700 !important; color: #fff; }
    .button.is-done { background: #208454 !important; color: #fff; }
    .button.is-failed { background: #9d2f2f !important; color: #fff; }
    pre { white-space: pre-wrap; background: #eef4f1; border-radius: 8px; padding: 14px; overflow: auto; }
    </style>
    <script>
    function faloAction(el, label) {
      if (!el || el.dataset.noActionState === 'yes') return true;
      const running = label || el.dataset.runningLabel || 'Running...';
      el.dataset.originalText = el.textContent.trim();
      el.textContent = running;
      el.classList.add('is-running');
      if (el.tagName === 'BUTTON' && el.type !== 'submit') el.disabled = true;
      return true;
    }
    document.addEventListener('click', event => {
      const el = event.target.closest('a.button, button.button');
      if (!el) return;
      if (el.getAttribute('href') === '#') return;
      faloAction(el);
    });
    </script>"""


def render_adapter_preview(result: Dict[str, object]) -> str:
    command_rows = "\n".join(
        f"""
        <tr>
          <td>{escape_html(item['task_id'])}</td>
          <td>{escape_html(item['csv_name'])}</td>
          <td><code>{escape_html(item['command'])}</code></td>
        </tr>
        """
        for item in result["commands"]
    ) or '<tr><td colspan="3" class="muted">沒有可預覽命令。請先 normalize 並建立 queue。</td></tr>'
    error_rows = "\n".join(
        f"<li>{escape_html(item.get('task_id', ''))} {escape_html(item['message'])}</li>"
        for item in result["errors"]
    ) or "<li>無</li>"

    project = result.get("project", {})
    if not isinstance(project, dict):
        project = {}
    notice_class = "result-ok" if result["command_count"] and not result["errors"] else "result-warn"
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>NotebookLM Adapter Preview</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Noto Sans TC", sans-serif; color: #202124; background: #f6f7f4; }}
    main {{ max-width: 1080px; margin: 0 auto; padding: 32px 20px 56px; }}
    h1 {{ margin: 0 0 8px; font-size: 32px; }}
    h2 {{ margin-top: 28px; color: #174f43; }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    th, td {{ border: 1px solid #d8ddd8; padding: 10px; vertical-align: top; text-align: left; }}
    th {{ background: #e8f1ed; color: #174f43; }}
    code {{ font-family: Menlo, Consolas, monospace; font-size: 13px; word-break: break-all; }}
    .button {{ display: inline-block; margin-top: 18px; padding: 9px 12px; border-radius: 6px; background: #1f6f5b; color: white; text-decoration: none; }}
    .muted {{ color: #667085; }}
    .result-ok {{ background: #e4f6ec; border-left: 7px solid #208454; padding: 16px; border-radius: 8px; }}
    .result-warn {{ background: #fff8e6; border-left: 7px solid #b87900; padding: 16px; border-radius: 8px; }}
  </style>
</head>
<body>
<main>
  <h1>NotebookLM Adapter Preview</h1>
  <div class="{notice_class}">
    <p><strong>Dry-run only.</strong> 已預覽 {result['command_count']} 個命令；此頁不會真的上傳 NotebookLM。</p>
    <p>Project: <strong>{escape_html(project.get('name', '(active/manual target)'))}</strong></p>
    <p>Notebook ID: <code>{escape_html(result['notebook_id'])}</code></p>
  </div>
  <h2>Command Preview</h2>
  <table>
    <thead><tr><th>Task ID</th><th>CSV</th><th>Command</th></tr></thead>
    <tbody>{command_rows}</tbody>
  </table>
  <h2>Errors</h2>
  <ul>{error_rows}</ul>
  <a class="button" href="/">Back to Portal</a>
</main>
</body>
</html>"""


def render_execute_result(result: Dict[str, object]) -> str:
    rows = "\n".join(
        f"""
        <tr>
          <td>{escape_html(item['task_id'])}</td>
          <td>{escape_html(item['csv_name'])}</td>
          <td>{escape_html(item['status'])}</td>
          <td>{item['returncode']}</td>
          <td>{escape_html(item['message'])}</td>
        </tr>
        """
        for item in result["results"]
    ) or '<tr><td colspan="5" class="muted">沒有執行任何 task。</td></tr>'
    error_rows = "\n".join(
        f"<li>{escape_html(item.get('task_id', ''))} {escape_html(item['message'])}</li>"
        for item in result["errors"]
    ) or "<li>無</li>"

    if result["mode"] == "blocked":
        result_class = "result-warn"
        status_text = "Blocked - needs confirmation"
    elif result["errors"]:
        result_class = "result-error"
        status_text = "Finished with errors"
    else:
        result_class = "result-ok"
        status_text = "Finished successfully"
    project = result.get("project", {})
    if not isinstance(project, dict):
        project = {}
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>NotebookLM Adapter Execute</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Noto Sans TC", sans-serif; color: #202124; background: #f6f7f4; }}
    main {{ max-width: 1080px; margin: 0 auto; padding: 32px 20px 56px; }}
    h1 {{ margin: 0 0 8px; font-size: 32px; }}
    h2 {{ margin-top: 28px; color: #174f43; }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    th, td {{ border: 1px solid #d8ddd8; padding: 10px; vertical-align: top; text-align: left; }}
    th {{ background: #e8f1ed; color: #174f43; }}
    code {{ font-family: Menlo, Consolas, monospace; font-size: 13px; word-break: break-all; }}
    .button {{ display: inline-block; margin-top: 18px; padding: 9px 12px; border-radius: 6px; background: #1f6f5b; color: white; text-decoration: none; }}
    .muted {{ color: #667085; }}
    .result-ok {{ background: #e4f6ec; border-left: 7px solid #208454; padding: 16px; border-radius: 8px; }}
    .result-warn {{ background: #fff8e6; border-left: 7px solid #b87900; padding: 16px; border-radius: 8px; }}
    .result-error {{ background: #fdecec; border-left: 7px solid #9d2f2f; padding: 16px; border-radius: 8px; }}
  </style>
</head>
<body>
<main>
  <h1>NotebookLM Adapter Execute</h1>
  <div class="{result_class}">
    <p><strong>{status_text}</strong></p>
    <p>Mode: <code>{escape_html(result['mode'])}</code> · executed: {result['executed_count']}</p>
    <p>Project: <strong>{escape_html(project.get('name', '(active/manual target)'))}</strong></p>
    <p>Notebook ID: <code>{escape_html(result['notebook_id'])}</code></p>
  </div>
  <h2>Results</h2>
  <table>
    <thead><tr><th>Task ID</th><th>CSV</th><th>Status</th><th>Return Code</th><th>Message</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
  <h2>Errors</h2>
  <ul>{error_rows}</ul>
  <p class="muted">一般操作請先 Set Active Project，再執行。手填 notebook_id 只保留給 debug。</p>
  <a class="button" href="/">Back to Portal</a>
</main>
</body>
</html>"""


def render_normalize_result(result: Dict[str, object]) -> str:
    output_rows = "\n".join(
        f"""
        <tr>
          <td>{escape_html(item['source'])}</td>
          <td>{escape_html(item['name'])}</td>
          <td>{item['size_kb']}</td>
          <td><code>{escape_html(item['path'])}</code></td>
        </tr>
        """
        for item in result["outputs"]
    ) or '<tr><td colspan="4" class="muted">沒有產生 CSV。請確認 inbox 內是否有 .xlsx。</td></tr>'
    error_rows = "\n".join(
        f"<li><strong>{escape_html(item['source'])}</strong>: {escape_html(item['error'])}</li>"
        for item in result["errors"]
    ) or "<li>無</li>"

    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>XLSX Normalize Result</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Noto Sans TC", sans-serif; color: #202124; background: #f6f7f4; }}
    main {{ max-width: 1080px; margin: 0 auto; padding: 32px 20px 56px; }}
    h1 {{ margin: 0 0 8px; font-size: 32px; }}
    h2 {{ margin-top: 28px; color: #174f43; }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    th, td {{ border: 1px solid #d8ddd8; padding: 10px; vertical-align: top; text-align: left; }}
    th {{ background: #e8f1ed; color: #174f43; }}
    code {{ font-family: Menlo, Consolas, monospace; font-size: 13px; }}
    .button {{ display: inline-block; margin-top: 18px; padding: 9px 12px; border-radius: 6px; background: #1f6f5b; color: white; text-decoration: none; }}
    .muted {{ color: #667085; }}
  </style>
</head>
<body>
<main>
  <h1>XLSX Normalize Result</h1>
  <p>已產生 {result['converted_count']} 個 CSV。這一步只做 ETL normalize，尚未上傳 NotebookLM。</p>
  <h2>CSV Outputs</h2>
  <table>
    <thead><tr><th>Source</th><th>CSV</th><th>Size KB</th><th>Path</th></tr></thead>
    <tbody>{output_rows}</tbody>
  </table>
  <h2>Errors</h2>
  <ul>{error_rows}</ul>
  <a class="button" href="/">Back to Portal</a>
</main>
</body>
</html>"""


def render_queue_result(result: Dict[str, object]) -> str:
    rows = "\n".join(
        f"""
        <tr>
          <td>{escape_html(item['task_id'])}</td>
          <td>{escape_html(item['status'])}</td>
          <td>{escape_html(item['csv_name'])}</td>
          <td>{item['size_kb']}</td>
          <td><code>{escape_html(item['csv_path'])}</code></td>
        </tr>
        """
        for item in result["tasks"]
    ) or '<tr><td colspan="5" class="muted">沒有找到 CSV。請先執行 XLSX Normalize。</td></tr>'

    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Queue Manifest Result</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Noto Sans TC", sans-serif; color: #202124; background: #f6f7f4; }}
    main {{ max-width: 1080px; margin: 0 auto; padding: 32px 20px 56px; }}
    h1 {{ margin: 0 0 8px; font-size: 32px; }}
    h2 {{ margin-top: 28px; color: #174f43; }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    th, td {{ border: 1px solid #d8ddd8; padding: 10px; vertical-align: top; text-align: left; }}
    th {{ background: #e8f1ed; color: #174f43; }}
    code {{ font-family: Menlo, Consolas, monospace; font-size: 13px; }}
    .button {{ display: inline-block; margin-top: 18px; padding: 9px 12px; border-radius: 6px; background: #1f6f5b; color: white; text-decoration: none; }}
    .muted {{ color: #667085; }}
  </style>
</head>
<body>
<main>
  <h1>Queue Manifest Result</h1>
  <p>已建立 {result['task_count']} 個 pending task。這一步只建立 ETL queue，不執行 NotebookLM 上傳。</p>
  <h2>Pending Tasks</h2>
  <table>
    <thead><tr><th>Task ID</th><th>Status</th><th>CSV</th><th>Size KB</th><th>Path</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
  <p>Manifest path: <code>{escape_html(PROJECT_ROOT / 'data' / 'temp' / 'etl_queue.json')}</code></p>
  <a class="button" href="/">Back to Portal</a>
</main>
</body>
</html>"""


def escape_html(value: object) -> str:
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def render_toggle_button(label: str, key: str, enabled: bool) -> str:
    state_label = "ON" if enabled else "OFF"
    state_class = "toggle-on" if enabled else "toggle-off"
    action_label = "Turn OFF" if enabled else "Turn ON"
    return (
        f'<a class="button {state_class}" href="/api/gas-settings/toggle?key={quote_url(key)}" '
        f'data-running-label="Switching...">{escape_html(label)}：{state_label} '
        f'<span class="muted-invert">({action_label})</span></a>'
    )


def quote_url(value: object) -> str:
    return quote_plus(str(value))


def run_server(port: int, open_browser: bool, host: str = "0.0.0.0") -> None:
    global RUNTIME_BIND_HOST, RUNTIME_BIND_PORT, task_queue_manager
    RUNTIME_BIND_HOST = host
    config = load_or_create_config(PROJECT_ROOT)
    ensure_gas_auto_worker(config)
    ensure_incoming_watch_worker(config)
    ensure_incoming_realtime_watch_worker(config)
    
    # Start task queue manager
    task_queue_manager = TaskQueueManager(config)
    task_queue_manager.start()
    
    actual_port = find_available_port(port, host)
    RUNTIME_BIND_PORT = actual_port
    server = ThreadingHTTPServer((host, actual_port), make_handler(config))
    local_url = f"http://127.0.0.1:{actual_port}"
    bind_url = f"http://{host}:{actual_port}"
    print(f"AI NotebookLM Runtime Lab portal: {local_url}")
    if host == "0.0.0.0":
        print(f"LAN portal: http://{detect_lan_ip()}:{actual_port}")
        print("Binding host: 0.0.0.0, same-network devices may connect if macOS firewall allows Python.")
    else:
        print(f"Binding host: {bind_url}")
    print("Close this terminal window or press Ctrl+C to stop.")
    if open_browser:
        webbrowser.open(local_url)
    server.serve_forever()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--host", default="0.0.0.0", help="Bind host. Default 0.0.0.0; use 127.0.0.1 for local-only socket binding.")
    parser.add_argument("--no-open", action="store_true")
    args = parser.parse_args()
    run_server(args.port, not args.no_open, args.host)


if __name__ == "__main__":
    main()
